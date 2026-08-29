#!/usr/bin/env python3
"""
Global Shutter Camera + YOLO26 TensorRT + rolling-window QR validation.

A single-purpose build of run.py: rolling-window mode only. Every QR the
detector finds anywhere in the frame is decoded once and offered, on its own,
to a window of the next few sheet rows. Because every payload in the sheet is
unique a code identifies itself — which row and which QR DATA column it
belongs to — so the order codes happen to decode in never matters.

Nothing is tolerated. A row that reaches the end of the window still missing
a code does not get written off: the line stops and the row is held open at
the head of the window. The operator winds the coil back, the same labels
pass the camera again, and their codes tick off exactly as they would have the
first time. Pressing START is the adjudication:

    row filled in while winding back  ->  cleared, the run carries on
    row still short                   ->  LABEL HAS ISSUE, recorded as a
                                          defect, and the window moves past it

Nothing is read while the machine is stopped. Pressing START does not energise
the relay straight away: for --start-delay seconds the camera reads the labels
standing in front of it while the coil is still, and only then does the relay
go on. That read-in is what validates the position the coil is actually in,
and it is also the second look a held row gets — starting the web first would
drag those labels out of frame before they could be checked.

What this build does NOT do, compared to run.py: no trigger-line mode, no
label grouping, and so no check of *where* a label sat. Four correct codes in
the wrong four positions tick four correct cells and the row passes. Use
run.py --mode row if position matters.

Usage:
    python3 run_window.py                                  # defaults
    python3 run_window.py --xlsx validation.xlsx
    python3 run_window.py --window-size 4                  # shorter rewind
    python3 run_window.py --no-display                     # headless
    python3 run_window.py --no-relay                       # vision only
    python3 run_window.py --dump-crops bad/                # save failed crops
    python3 run_window.py --debug                          # print every read

On start-up relay 0 is switched ON to run the winding machine, and it is
switched OFF whenever a row will not validate.
"""

import argparse
import os
import threading
import time
import cv2

from utils.crops import LabelSaver
from utils.qr import decode_qr, pick_qr_for_label
from utils.relay import RelayController
from utils.results import ResultLog
from utils.trt_engine import YOLO26TRT
from utils.ui import ControlPanel
from utils.utils import preprocess, postprocess, draw_detections
from utils.validation import ValidationSheet, normalize

# ── Stream config (same defaults as cam_view.py) ────────────────────────────
DEFAULT_CAM_INDEX = 0
DEFAULT_WIDTH     = 2592
DEFAULT_HEIGHT    = 1944
DEFAULT_FPS       = 60       # MJPG supports 60fps at full 2592x1944; YUYV only
                              # goes to 35fps at that size (see --list-formats-ext)
DEFAULT_FORMAT    = "MJPG"
DISPLAY_MAX_W     = 1280     # imshow window is capped to this width so a full
DISPLAY_MAX_H     = 960      # -res frame doesn't overflow the screen
DEFAULT_ROTATE    = 270      # fixed rotation applied to every frame: 0/90/180/270

# ── Inference config ─────────────────────────────────────────────────────────
DEFAULT_IMGSZ      = 640
DEFAULT_CONF_THRES = 0.25   # applies to any class without its own threshold
DEFAULT_CONF_LABEL = None   # None -> fall back to DEFAULT_CONF_THRES
DEFAULT_CONF_QR    = None

# ── QR decode config ─────────────────────────────────────────────────────────
DEFAULT_LABEL_CLASS = "label"      # class whose crossing triggers a decode
DEFAULT_QR_CLASS    = "qr_code"    # class that is cropped and decoded
DEFAULT_QR_MARGIN   = 0.15         # quiet zone added around the qr box

# ── Validation / machine config ──────────────────────────────────────────────
DEFAULT_XLSX        = "validation_x300.xlsx"   # expected QR sequence
DEFAULT_START_RELAY = 0                   # relay 0 = winding machine start
DEFAULT_START_DELAY = 2.0                 # seconds of reading after START is
                                          # pressed, before the relay goes on
DEFAULT_RESULT_DIR  = "result"            # <result>/<xlsx name>/labels/

ROTATE_MAP = {
    0:   None,
    90:  cv2.ROTATE_90_CLOCKWISE,
    180: cv2.ROTATE_180,
    270: cv2.ROTATE_90_COUNTERCLOCKWISE,
}


# videoflip's equivalent of ROTATE_MAP, so the rotation can be done inside
# GStreamer instead of on the capture loop. cv2.rotate on a full 2592x1944
# BGR frame costs ~3.1ms of the ~21ms frame budget; videoflip does the same
# work on the GStreamer thread and the camera still delivers its full 60fps,
# so those 3.1ms come straight off the loop.
VIDEOFLIP_MAP = {
    0:   None,
    90:  "clockwise",
    180: "rotate-180",
    270: "counterclockwise",
}


def rotate_frame(frame, degrees):
    """Rotate a frame by a fixed angle (0/90/180/270). No-op for 0.

    Only used when the rotation could not be pushed into the pipeline; the
    normal path has videoflip deliver frames already rotated."""
    code = ROTATE_MAP[degrees]
    return frame if code is None else cv2.rotate(frame, code)


GS_CAMERA_NAME = "Global Shutter Camera"
WINDOW_VIEW = "Rolling window"


def list_cameras():
    """List (index, name) for every /dev/videoN device via its v4l2 sysfs name."""
    cams = []
    for d in sorted(os.listdir("/dev")):
        if not d.startswith("video"):
            continue
        sys_path = f"/sys/class/video4linux/{d}/name"
        if not os.path.exists(sys_path):
            continue
        with open(sys_path) as f:
            name = f.read().strip()
        cams.append((int(d.replace("video", "")), name))
    return cams


def find_camera_index(name_substring=GS_CAMERA_NAME, default=DEFAULT_CAM_INDEX):
    """Find the /dev/videoN index whose v4l2 name contains name_substring."""
    cams = list_cameras()
    for index, name in cams:
        if name_substring.lower() in name.lower():
            return index
    print(f"[camera] '{name_substring}' not found among {cams} — "
          f"falling back to index {default}")
    return default


def gstreamer_pipeline(cam_index=DEFAULT_CAM_INDEX, width=DEFAULT_WIDTH,
                        height=DEFAULT_HEIGHT, fps=DEFAULT_FPS, format=DEFAULT_FORMAT,
                        rotate=0):
    """Build a GStreamer pipeline string for v4l2src (MJPG or YUYV).

    `rotate` (0/90/180/270) is applied by videoflip inside the pipeline, so
    frames arrive already rotated and the loop never touches them."""
    QUEUE = "queue leaky=downstream max-size-buffers=1"
    method = VIDEOFLIP_MAP.get(rotate)
    FLIP  = f"videoflip method={method} ! " if method else ""
    SINK  = (f"{FLIP}videoconvert ! video/x-raw, format=BGR ! "
             "appsink drop=true max-buffers=1 sync=false")

    if format.upper() == "MJPG":
        return (
            f"v4l2src device=/dev/video{cam_index} ! "
            f"image/jpeg, width={width}, height={height}, framerate={fps}/1 ! "
            f"{QUEUE} ! jpegdec ! {SINK}"
        )
    else:
        return (
            f"v4l2src device=/dev/video{cam_index} ! "
            f"video/x-raw, width={width}, height={height}, framerate={fps}/1 ! "
            f"{QUEUE} ! {SINK}"
        )


def load_class_names(path):
    if not path:
        return None
    with open(path) as f:
        return [line.strip() for line in f if line.strip()]


class RollingWindow:
    """A sliding window of expected sheet rows, matched in any order.

    Line-crossing validation required the labels to arrive in sequence: a row
    was judged as its column passed a fixed point, and anything the camera
    missed became an out-of-sequence fault. Here the sequence constraint is
    dropped. Every QR the detector finds anywhere in the frame is offered to a
    window of the next `size` sheet rows; if the payload belongs to any cell
    inside that window, the cell is ticked off. Order does not matter, and a
    label can be read on whichever frame it happens to be legible in.

    A row leaves the window once every checked cell has been ticked, and the
    head slides forward one row at a time. A payload that belongs to no cell
    in the window is the fault worth stopping for — it means a label is on the
    web that the sheet does not expect here.
    """

    MATCH, REPEAT, UNKNOWN = "match", "repeat", "unknown"
    UNREAD = "unread"          # a group went by and nothing on it decoded

    def __init__(self, sheet, size=8, check=None, grace=4):
        self.sheet = sheet
        self.size = max(1, size)
        self.check = check
        # Rows kept in the lookup *behind* the head. A label stays in view for
        # many frames and is decoded on every one, so codes keep arriving for
        # rows that already finished. Keeping those rows matchable lets such a
        # read be recognised as a repeat instead of being mistaken for the
        # next block's copy of the same code, or for a code from nowhere.
        self.grace = max(0, grace)
        self.start = None            # sheet row index at the head of the window
        self.seen = {}               # row index -> set of columns ticked off
        self.index = {}              # payload -> [(row index, column)]
        self.done = []               # [(row index, complete?)] as rows retire
        self.decoded_cells = set()   # (row index, column) ever decoded
        self.texts = {}              # (row index, column) -> payload as read
        self.last_hit = None         # (row index, column) most recently ticked
        self.unexpected = []         # recent payloads that belonged nowhere
        self.reads = 0
        self.repeats = 0

    # ── window contents ───────────────────────────────────────────────────
    def rows(self):
        if self.start is None:
            return []
        end = min(self.start + self.size, len(self.sheet.rows))
        return list(range(self.start, end))

    def required(self, row_idx):
        """Columns that must be decoded before this row can retire."""
        n = len(self.sheet.rows[row_idx].texts)
        if self.check is None:
            return {c for c in range(n) if normalize(self.sheet.rows[row_idx].texts[c])}
        return {c for c in self.check if c < n}

    def missing(self, row_idx):
        return self.required(row_idx) - self.seen.get(row_idx, set())

    def _rebuild(self):
        """The payload lookup covers the window plus the grace rows behind it,
        which is what makes 'not expected here' meaningful."""
        self.index = {}
        lo = max(0, self.start - self.grace) if self.start is not None else 0
        for row_idx in list(range(lo, self.start or 0)) + self.rows():
            for col, text in enumerate(self.sheet.rows[row_idx].texts):
                key = normalize(text)
                if key:
                    self.index.setdefault(key, []).append((row_idx, col))

    def anchor(self, row_idx):
        self.start = max(0, row_idx)
        self.seen = {}
        self._rebuild()
        return self.sheet.rows[self.start].number

    # ── matching ──────────────────────────────────────────────────────────
    def offer(self, text):
        """Tick off whatever cell this payload belongs to.
        Returns (status, row_index, column, slot) — slot is how deep into the
        window the hit landed, 0 being the head."""
        hits = self.index.get(normalize(text))
        if not hits:
            return self.UNKNOWN, None, None, None
        row_idx, col = min(hits, key=lambda h: h[0])   # nearest the head
        if row_idx < self.start:
            self.repeats += 1        # a finished row being read again
            return self.REPEAT, row_idx, col, None
        slot = row_idx - self.start
        self.last_hit = (row_idx, col)
        if col in self.seen.get(row_idx, set()):
            self.repeats += 1
            return self.REPEAT, row_idx, col, slot
        self.seen.setdefault(row_idx, set()).add(col)
        self.decoded_cells.add((row_idx, col))
        self.texts[(row_idx, col)] = text
        self.reads += 1
        return self.MATCH, row_idx, col, slot

    # ── sliding ───────────────────────────────────────────────────────────
    def note_unexpected(self, text, belongs):
        """Keep the last few codes that matched nothing, for the view."""
        self.unexpected.append((text, belongs))
        del self.unexpected[:-4]

    def advance(self):
        """Retire finished rows from the head. Returns [(row index, True)]."""
        retired = []
        while self.start is not None and self.start < len(self.sheet.rows):
            if self.missing(self.start):
                break
            retired.append((self.start, True))
            self.done.append((self.start, True))
            self.start += 1
        if retired:
            self._rebuild()
        return retired

    def evict_head(self):
        """Drop the head row even though it is short — the web has moved a
        whole window past it, so it is never going to be completed."""
        if self.start is None or self.start >= len(self.sheet.rows):
            return None
        row_idx = self.start
        self.done.append((row_idx, False))
        self.start += 1
        self._rebuild()
        return row_idx

    @property
    def exhausted(self):
        return self.start is not None and self.start >= len(self.sheet.rows)


def _overlap(a, b):
    """Intersection over union of two boxes — used to recognise the same
    physical label from one frame to the next."""
    ix = max(0.0, min(a[2], b[2]) - max(a[0], b[0]))
    iy = max(0.0, min(a[3], b[3]) - max(a[1], b[1]))
    inter = ix * iy
    if inter <= 0:
        return 0.0
    area_a = max(0.0, a[2] - a[0]) * max(0.0, a[3] - a[1])
    area_b = max(0.0, b[2] - b[0]) * max(0.0, b[3] - b[1])
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


def sheet_period(sheet, limit=400):
    """Smallest number of rows after which the sheet repeats itself, or None.

    A production sheet is often one block of codes duplicated many times over
    (validation_x300.xlsx is 17 rows copied 300 times). The window has to stay
    shorter than that block: if it spans a whole period, every payload appears
    in it twice and the matching cannot tell one pass from the next.
    """
    rows = [tuple(r.texts) for r in sheet.rows]
    n = len(rows)
    for period in range(1, min(limit, n // 2 + 1)):
        if n % period:
            continue
        if all(rows[i] == rows[i % period] for i in range(n)):
            return period
    return None


def parse_check(spec, per_row):
    """Turn "D2,D3" (or "2 3", or "d2;d4") into the 0-based positions to check.
    None means every position."""
    if not spec:
        return None
    wanted = set()
    for part in spec.replace(";", ",").replace(" ", ",").split(","):
        part = part.strip().upper().lstrip("D")
        if not part:
            continue
        if not part.isdigit():
            raise SystemExit(f"--check: '{part}' is not a QR DATA number")
        n = int(part)
        if not 1 <= n <= per_row:
            raise SystemExit(f"--check: QR DATA{n} is out of range — the sheet "
                             f"has {per_row} code columns")
        wanted.add(n - 1)
    if not wanted:
        return None
    return wanted


def class_index(class_names, name, fallback):
    """Resolve a class name to its id; fall back to a fixed index when no
    classes.txt was supplied (or the name isn't in it)."""
    if class_names and name in class_names:
        return class_names.index(name)
    print(f"[qr] class '{name}' not found in --classes, using index {fallback}")
    return fallback


def main():
    ap = argparse.ArgumentParser(
        description="Global Shutter Camera + YOLO26 TensorRT + rolling-window "
                    "QR validation, with re-inspection on any short row.")
    # camera args
    ap.add_argument("--index", type=int, default=None,
                     help="Force a /dev/videoN index (skips auto-detect).")
    ap.add_argument("--width", type=int, default=DEFAULT_WIDTH)
    ap.add_argument("--height", type=int, default=DEFAULT_HEIGHT)
    ap.add_argument("--fps", type=int, default=DEFAULT_FPS)
    ap.add_argument("--format", default=DEFAULT_FORMAT, choices=["MJPG", "YUYV"])
    ap.add_argument("--rotate", type=int, default=DEFAULT_ROTATE, choices=[0, 90, 180, 270],
                     help="Rotate every frame by a fixed angle (clockwise).")
    ap.add_argument("--no-display", action="store_true",
                     help="Just print FPS/detections instead of opening a window (headless).")
    ap.add_argument("--debug", action="store_true",
                     help="print every payload as it reads and where it sits in "
                          "the sheet; without it only the per-row verdicts print.")
    # inference args
    ap.add_argument("--engine", default="best.engine", help="path to .engine file")
    ap.add_argument("--classes", default=None,
                     help="txt file, one class name per line "
                          "(default: classes.txt beside run.py, if present)")
    ap.add_argument("--conf-thres", type=float, default=DEFAULT_CONF_THRES,
                     help="confidence threshold for classes without their own.")
    ap.add_argument("--conf-label", type=float, default=DEFAULT_CONF_LABEL,
                     help="confidence threshold for the label class — this is "
                          "what gates a trigger-line crossing.")
    ap.add_argument("--conf-qr", type=float, default=DEFAULT_CONF_QR,
                     help="confidence threshold for the qr class — this is what "
                          "gates which box gets cropped and decoded.")
    ap.add_argument("--imgsz", type=int, default=DEFAULT_IMGSZ)
    ap.add_argument("--save", default=None, help="optional path to record annotated video")
    # qr decode args
    ap.add_argument("--label-class", default=DEFAULT_LABEL_CLASS,
                     help="class name that triggers a decode when it crosses the line.")
    ap.add_argument("--qr-class", default=DEFAULT_QR_CLASS,
                     help="class name of the QR box that gets cropped and decoded.")
    ap.add_argument("--qr-margin", type=float, default=DEFAULT_QR_MARGIN,
                     help="quiet zone around the qr box, as a fraction of its size.")
    ap.add_argument("--qr-margin-min", type=int, default=8,
                     help="minimum quiet zone in pixels.")
    ap.add_argument("--dump-crops", default=None,
                     help="directory to save qr crops that failed to decode.")
    # label crop args
    ap.add_argument("--result-dir", default=DEFAULT_RESULT_DIR,
                     help="root for saved label crops: "
                          "<result-dir>/<xlsx name>/labels/.")
    ap.add_argument("--no-save-labels", action="store_true",
                     help="don't save a crop of each decoded label.")
    ap.add_argument("--label-format", default="jpg", choices=["jpg", "png"],
                     help="image format for the saved label crops.")
    ap.add_argument("--label-pad", type=float, default=0.0,
                     help="padding around the saved label crop, as a fraction "
                          "of the box size.")
    # validation args
    ap.add_argument("--xlsx", default=DEFAULT_XLSX,
                     help="xlsx holding the expected QR DATA1..N sequence.")
    ap.add_argument("--sheet", default=None,
                     help="worksheet name inside --xlsx (default: the first one).")
    ap.add_argument("--labels-per-row", type=int, default=None,
                     help="labels per crossing (default: the number of QR DATA "
                          "columns found in the sheet).")
    ap.add_argument("--check", default=None,
                     help="which label positions to validate, top to bottom, "
                          "e.g. 'D2,D3' or '2,3'. The rest are neither decoded "
                          "nor held against the row. Default: all of them.")
    ap.add_argument("--no-stop-on-fail", action="store_true",
                     help="keep the machine running when a row fails "
                          "validation (default: stop it).")
    ap.add_argument("--no-result-log", action="store_true",
                     help="don't write the per-row CSV of verdicts.")
    ap.add_argument("--window-size", type=int, default=8,
                     help="how many sheet rows the rolling window holds. "
                          "Bigger tolerates more out-of-order arrival and more "
                          "missed rows; smaller catches a stray label sooner.")
    ap.add_argument("--window-grace", type=int, default=4,
                     help="rows kept matchable behind the window, so a label "
                          "still in view after its row finished is recognised "
                          "as a re-read instead of an unexpected code.")
    ap.add_argument("--out-xlsx", default=None,
                     help="where to write the sheet annotated with what was "
                          "decoded. Defaults to a fixed path per sheet so it "
                          "can be resumed; the source --xlsx is never touched.")
    ap.add_argument("--xlsx-every", type=int, default=0,
                     help="also rewrite the annotated .xlsx every N rows. "
                          "Costly (openpyxl is pure Python and stalls the "
                          "capture loop), and not needed for durability — "
                          "progress is journalled continuously either way. "
                          "0 = write the xlsx at exit only.")
    ap.add_argument("--max-decodes", type=int, default=0,
                     help="cap how many labels are decoded per frame. 0 = no "
                          "cap. Labels already read are skipped regardless.")
    ap.add_argument("--no-resume", action="store_true",
                     help="ignore the annotated sheet from a previous run and "
                          "start the record empty, anchoring wherever the "
                          "sheet first matches. By default a previous run is "
                          "picked up and the window continues past it.")
    # relay args
    ap.add_argument("--no-relay", action="store_true",
                     help="don't touch the relay board (vision only).")
    ap.add_argument("--relay-port", default=None,
                     help="serial port of the relay board (default: auto-detect).")
    ap.add_argument("--start-delay", type=float, default=DEFAULT_START_DELAY,
                     help="seconds to spend reading the labels already in "
                          "front of the camera after START is pressed, before "
                          "the relay is switched on. This is what validates "
                          "the coil at the position it is actually in — and "
                          "what re-checks a row that was held open. 0 = "
                          "energise immediately.")
    ap.add_argument("--start-relay", type=int, default=DEFAULT_START_RELAY,
                     help="relay that starts the winding machine.")
    ap.add_argument("--relay-verbose", action="store_true",
                     help="print every modbus frame sent to the relay board.")
    args = ap.parse_args()

    # ── the expected sheet, and the window over it ───────────────────────
    sheet = ValidationSheet(args.xlsx, args.sheet)
    per_row = args.labels_per_row or sheet.per_row
    checked = parse_check(args.check, per_row)
    if checked is None:
        print(f"[validate] checking all {per_row} positions")
    else:
        on = ", ".join(f"QR DATA{i + 1}" for i in sorted(checked))
        off = ", ".join(f"QR DATA{i + 1}" for i in range(per_row)
                        if i not in checked)
        print(f"[validate] checking {on}" + (f" (ignoring {off})" if off else ""))

    size, grace = args.window_size, args.window_grace
    period = sheet_period(sheet)
    if period:
        # A window that spans the sheet's repeat would hold the same payload
        # twice over, and a code could tick off either copy.
        room = period - grace - 1
        if size > room:
            print(f"[window] the sheet repeats every {period} rows, so a "
                  f"window of {size} would see each code twice — "
                  f"using {max(1, room)} instead")
            size = max(1, room)
    window = RollingWindow(sheet, size=size, check=checked, grace=grace)
    print(f"[window] rolling window of {window.size} sheet rows "
          f"(+{window.grace} kept behind for re-reads); every QR in frame "
          f"is matched on its own, order does not matter")
    print(f"[window] a row that comes up short stops the line and is held "
          f"open — wind it back past the camera, then press START")

    machine_running = False
    # When START was pressed, or None when it was not. Between that moment and
    # --start-delay seconds later the camera is reading but the relay is still
    # off: the coil is standing still in front of the lens, which is the one
    # chance to validate the position it is actually in before it moves.
    starting_at = [None]
    start_reason = [""]

    def validating():
        """Codes are only read when the machine is running, or about to be."""
        return machine_running or starting_at[0] is not None

    def start_machine(reason="operator"):
        """Begin the start sequence. This does NOT energise the relay.

        The relay goes on in _finish_start, once the labels standing in front
        of the camera have had --start-delay seconds to read. Starting the web
        first would drag them out of frame before they could be checked.
        """
        if machine_running or starting_at[0] is not None:
            return
        starting_at[0] = time.time()
        start_reason[0] = reason
        if panel is not None:
            panel.note = f"reading labels… ({args.start_delay:.0f}s)"
        held = (f", re-checking row {recheck['row']}"
                if recheck["row"] is not None else "")
        print(f"\n[relay] START pressed ({reason}) — reading the labels in "
              f"frame for {args.start_delay:.1f}s before the relay goes on"
              f"{held}")

    def _finish_start():
        """The read-in has run its course: settle up and let the web go."""
        nonlocal machine_running
        starting_at[0] = None

        # Now, not on the button press: a row held for re-inspection has just
        # had its second look, so it either filled in during the read-in — in
        # which case the retire path has already cleared it — or it is a
        # confirmed defect.
        if recheck["row"] is not None:
            _adjudicate_recheck()

        relay.on(args.start_relay)
        machine_running = True
        if panel is not None:
            panel.note = None
        print(f"[relay] winding machine STARTED ({start_reason[0]}) "
              f"— relay {args.start_relay} ON")

    def stop_machine(reason="operator"):
        nonlocal machine_running
        aborted = starting_at[0] is not None
        starting_at[0] = None
        if not machine_running and not aborted:
            return
        relay.off(args.start_relay)
        machine_running = False
        if panel is not None:
            panel.note = reason
        what = "start ABORTED" if aborted and not machine_running else "STOPPED"
        print(f"\n[relay] winding machine {what} ({reason}) "
              f"— relay {args.start_relay} OFF")

    run_name = os.path.splitext(os.path.basename(args.xlsx))[0]
    saver = None
    if not args.no_save_labels:
        saver = LabelSaver(root=args.result_dir, name=run_name,
                           ext=args.label_format, pad=args.label_pad)

    results = None
    if not args.no_result_log:
        results = ResultLog(root=args.result_dir, name=run_name,
                            columns=per_row)

    relay = RelayController(port=args.relay_port, enabled=not args.no_relay,
                            verbose=args.relay_verbose)
    panel = None

    cam_index = args.index if args.index is not None else find_camera_index()
    print(f"[camera] using /dev/video{cam_index}")

    # Preferred path: videoflip rotates inside the pipeline, off the loop. If
    # that pipeline won't open (no videoflip element, say), fall back to
    # rotating each frame in the loop as before.
    pipeline = gstreamer_pipeline(cam_index, args.width, args.height, args.fps,
                                  args.format, rotate=args.rotate)
    print(f"[camera] pipeline: {pipeline}")
    cap = cv2.VideoCapture(pipeline, cv2.CAP_GSTREAMER)
    rotate_in_loop = 0
    if not cap.isOpened() and args.rotate:
        print("[camera] videoflip pipeline would not open — rotating in the loop")
        pipeline = gstreamer_pipeline(cam_index, args.width, args.height,
                                      args.fps, args.format)
        cap = cv2.VideoCapture(pipeline, cv2.CAP_GSTREAMER)
        rotate_in_loop = args.rotate
    if not cap.isOpened():
        raise RuntimeError("Failed to open camera via GStreamer pipeline")

    classes_path = args.classes
    if classes_path is None:
        beside = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                              "classes.txt")
        if os.path.exists(beside):
            classes_path = beside
            print(f"[model] using {beside}")
    class_names = load_class_names(classes_path)
    model = YOLO26TRT(args.engine, input_size=(args.imgsz, args.imgsz))
    print(f"[model] loaded {args.engine}")

    # 90/270 rotation swaps the effective width/height for sizing the window/writer.
    disp_w, disp_h = (args.height, args.width) if args.rotate in (90, 270) else (args.width, args.height)

    # Per-class thresholds: the label one gates line crossings, the qr one
    # gates which box gets cropped and decoded, and --conf-thres is what every
    # other class is held to.
    conf_per_class = {}
    label_cls = class_index(class_names, args.label_class, 0)
    qr_cls = class_index(class_names, args.qr_class, 1)
    if args.conf_label is not None:
        conf_per_class[label_cls] = args.conf_label
    if args.conf_qr is not None:
        conf_per_class[qr_cls] = args.conf_qr
    print(f"[model] conf thresholds: default={args.conf_thres}"
          + "".join(f"  {args.label_class if c == label_cls else args.qr_class}={v}"
                    for c, v in conf_per_class.items()))

    xlsx_path = args.out_xlsx or os.path.join(
        args.result_dir, run_name, f"checked_{run_name}.xlsx")
    journal_path = os.path.join(args.result_dir, run_name, "progress.csv")
    journal = [None]         # append-only record; the durable source of truth
    xlsx_lock = threading.Lock()
    xlsx_busy = [False]
    # Where the previous run got to. A repeating sheet holds each payload many
    # times over, so without this the first code the camera sees anchors on
    # its copy in the very first block and the whole run starts again.
    resume_hint = [None]
    handled = [[]]           # boxes whose code was accepted on the last frame
    verified = {}          # excel row number -> (per-column marks, status)

    # ── re-inspection state (row mode) ───────────────────────────────────
    # Nothing is tolerated: a group that will not validate stops the line and
    # is held open, so the operator can wind the coil back and show the same
    # labels to the camera again. Reading clean the second time clears it;
    # failing again makes it a confirmed defect.
    recheck = {"row": None, "attempt": 0}   # sheet row awaiting re-inspection
    defects = set()                         # rows confirmed bad on re-check

    def _open_journal():
        """A one-line-per-row append log, flushed as it goes.

        The .xlsx is the deliverable, but writing it is far too slow to do
        during a run — openpyxl is pure Python, so it holds the GIL and stalls
        the capture loop. This carries the same information at a cost of
        microseconds, and the workbook is rebuilt from it at the end.
        """
        os.makedirs(os.path.dirname(journal_path) or ".", exist_ok=True)
        fresh = not os.path.exists(journal_path)
        journal[0] = open(journal_path, "a", buffering=1)
        if fresh:
            per_row = len(sheet.rows[0].texts)
            journal[0].write("sheet_row,status," +
                             ",".join(f"d{i+1}" for i in range(per_row)) + "\n")

    def _load_journal():
        """Rows verified by earlier runs. Last entry for a row wins."""
        if args.no_resume or not os.path.exists(journal_path):
            return
        import csv as _csv
        with open(journal_path, newline="") as fh:
            for rec in _csv.reader(fh):
                if len(rec) < 3 or rec[0] == "sheet_row":
                    continue
                try:
                    number = int(rec[0])
                except ValueError:
                    continue
                verified[number] = (rec[2:], rec[1])

    def _load_previous():
        """Pick up what an earlier run already verified, so the record adds up
        across sessions instead of starting blank each time."""
        if args.no_resume or not os.path.exists(xlsx_path):
            return
        import openpyxl
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            rows = list(wb.worksheets[0].iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            # A file left over from a kill during an older, non-atomic save.
            # Starting fresh beats refusing to run.
            spoiled = xlsx_path + ".damaged"
            try:
                os.replace(xlsx_path, spoiled)
                where = f" — moved to {spoiled}"
            except OSError:
                where = ""
            print(f"[window] could not read {xlsx_path} ({exc}){where}; "
                  f"starting the record fresh")
            return
        if not rows:
            return
        header = [str(c) if c is not None else "" for c in rows[0]]
        try:
            first = header.index("READ D1")
        except ValueError:
            return
        per_row = len(sheet.rows[0].texts)
        for n, values in enumerate(rows[1:], start=2):
            status = values[first + per_row] if first + per_row < len(values) else None
            if not status:
                continue
            marks = [values[first + i] if first + i < len(values) else None
                     for i in range(per_row)]
            verified[n] = (marks, status)
        pass

    def record_row(row_idx, complete, status=None):
        """Fold one retired row into the record kept for the sheet."""
        per_row = len(sheet.rows[0].texts)
        required = window.required(row_idx)
        seen = window.seen.get(row_idx, set())
        marks = []
        for col in range(per_row):
            if col not in required:
                marks.append("not checked")
            elif col in seen:
                marks.append("OK")
            else:
                marks.append("NOT DECODED")
        status = status or ("OK" if complete and not (required - seen)
                            else "INCOMPLETE")
        number = sheet.rows[row_idx].number
        # A row verified on an earlier pass is not un-verified by a later one
        # that happened to read it badly — but a confirmed defect always wins,
        # because those labels were shown to the camera twice on purpose.
        if (verified.get(number, (None, None))[1] == "OK"
                and status not in ("OK", "DEFECT")):
            return
        verified[number] = (marks, status)
        if journal[0] is not None:
            journal[0].write(f"{number},{status}," + ",".join(marks) + "\n")

    def write_annotated_xlsx(background=False):
        """Copy the source sheet and add, per row, which codes were decoded.

        The source file is the reference for what was printed and is never
        touched — this is a second file recording what the camera could
        actually read of it. Written to a fixed path so --resume can find it.
        """
        if not verified:
            return
        if background:
            with xlsx_lock:
                if xlsx_busy[0]:
                    return          # one already in flight; the next will cover it
                xlsx_busy[0] = True
            snapshot = dict(verified)
            threading.Thread(target=_do_write, args=(snapshot, True),
                             daemon=True).start()
            return
        _do_write(dict(verified), False)

    def _do_write(snapshot, background):
        try:
            import openpyxl
            os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
            wb = openpyxl.load_workbook(args.xlsx)
            ws = wb[args.sheet] if args.sheet else wb.worksheets[0]
            first = ws.max_column + 1
            per_row = len(sheet.rows[0].texts)
            for i in range(per_row):
                ws.cell(row=1, column=first + i, value=f"READ D{i + 1}")
            ws.cell(row=1, column=first + per_row, value="ROW STATUS")
            ws.cell(row=1, column=first + per_row + 1, value="CHECKED AT")

            stamp = time.strftime("%Y-%m-%d %H:%M:%S")
            for number, (marks, status) in snapshot.items():
                for col, mark in enumerate(marks):
                    ws.cell(row=number, column=first + col, value=mark)
                ws.cell(row=number, column=first + per_row, value=status)
                ws.cell(row=number, column=first + per_row + 1, value=stamp)
            # Written to one side and moved into place, because os.replace is
            # atomic: a kill during the save leaves the previous good file
            # intact rather than a truncated one that cannot be reopened.
            tmp = xlsx_path + ".tmp"
            wb.save(tmp)
            wb.close()
            os.replace(tmp, xlsx_path)
            if not background:
                ok = sum(1 for _, st in snapshot.values() if st == "OK")
                print(f"[window] wrote {xlsx_path} — {ok} rows OK, "
                      f"{len(snapshot) - ok} incomplete")
        except Exception as exc:
            print(f"[window] could not write the annotated sheet: {exc}")
        finally:
            if background:
                with xlsx_lock:
                    xlsx_busy[0] = False

    # The panel costs ~2.9ms of the ~21ms frame budget to draw — it is ~50
    # putText calls — but its content only moves when the window does. Cache
    # it against the window state it reads, so a frame that changed nothing
    # reuses the last image instead of redrawing it.
    view_cache = [None, None]        # [state key, rendered image]

    def _window_view_key():
        return (window.start, window.reads, len(window.done), window.last_hit,
                len(window.unexpected))

    def render_window_view():
        """The rolling window as its own panel: what the sheet expects against
        what has actually been read, plus the state of every open row.

        Redrawn only when the window state behind it has changed."""
        key = _window_view_key()
        if view_cache[0] == key and view_cache[1] is not None:
            return view_cache[1]
        img = _draw_window_view()
        view_cache[0], view_cache[1] = key, img
        return img

    def _draw_window_view():
        import numpy as np
        font = cv2.FONT_HERSHEY_SIMPLEX
        W = 1120
        open_rows = window.rows()[:8] if window.start is not None else []
        H = 150 + 4 * 30 + 40 + max(len(open_rows), 1) * 28 + 90
        img = np.full((H, W, 3), (32, 32, 32), np.uint8)

        BG_OK, BG_BAD, DIM = (80, 220, 80), (70, 70, 235), (130, 130, 130)
        HEAD, WHITE, AMBER = (200, 200, 200), (240, 240, 240), (60, 200, 235)

        def put(text, x, y, colour=WHITE, scale=0.55, thick=1):
            cv2.putText(img, text, (x, y), font, scale, colour, thick, cv2.LINE_AA)

        def tail(text, n=30):
            if not text:
                return ""
            text = str(text)
            return text if len(text) <= n else ".." + text[-(n - 2):]

        if window.start is None:
            put("ROLLING WINDOW", 24, 42, WHITE, 0.9, 2)
            put("waiting for a code the sheet recognises", 24, 80, DIM, 0.65)
            return img

        passed = sum(1 for _, ok in window.done if ok)
        short = len(window.done) - passed
        put("ROLLING WINDOW", 24, 42, WHITE, 0.9, 2)
        put(f"{passed} pass   {short} short   {window.reads} codes read   "
            f"window {window.size} rows",
            24, 72, BG_BAD if short else BG_OK, 0.6)

        # ── the row most recently touched, expected against decoded ───────
        y = 110
        if window.last_hit is not None:
            row_idx, _ = window.last_hit
            row_no = sheet.rows[row_idx].number
            required = window.required(row_idx)
            seen = window.seen.get(row_idx, set())
            put(f"LAST MATCHED  ->  sheet row {row_no}", 24, y, HEAD, 0.65)
            y += 26
            for label, x in (("POS", 24), ("EXPECTED (xlsx)", 96),
                             ("DECODED (camera)", 450), ("RESULT", 810)):
                put(label, x, y, HEAD, 0.5)
            cv2.line(img, (16, y + 10), (W - 16, y + 10), (70, 70, 70), 1)
            for col, expected in enumerate(sheet.rows[row_idx].texts):
                y2 = y + 10 + (col + 1) * 30
                if col not in required:
                    got, res, colour = "(not checked)", "-", DIM
                elif col in seen:
                    got = tail(window.texts.get((row_idx, col), expected))
                    res, colour = "OK", BG_OK
                else:
                    got, res, colour = "(waiting)", "..", AMBER
                put(f"D{col + 1}", 24, y2, HEAD, 0.55)
                put(tail(expected) or "-", 96, y2, DIM, 0.5)
                put(got, 450, y2, colour, 0.5)
                put(res, 810, y2, colour, 0.5)
            y = y + 10 + 5 * 30
        else:
            y += 20

        # ── every open row, and what each still needs ─────────────────────
        y += 14
        put("OPEN ROWS", 24, y, HEAD, 0.6)
        y += 8
        for i, row_idx in enumerate(open_rows):
            y2 = y + (i + 1) * 28
            row_no = sheet.rows[row_idx].number
            required = window.required(row_idx)
            seen = window.seen.get(row_idx, set())
            complete = required and required <= seen
            colour = BG_OK if complete else (AMBER if i == 0 else WHITE)
            put(("> " if i == 0 else "  ") + f"row {row_no:<6}", 24, y2, colour, 0.55)
            x = 190
            for col in range(len(sheet.rows[row_idx].texts)):
                if col not in required:
                    mark, c = "-", DIM
                elif col in seen:
                    mark, c = "OK", BG_OK
                else:
                    mark, c = "..", AMBER
                put(f"D{col + 1}:{mark}", x, y2, c, 0.55)
                x += 92
            still = sorted(required - seen)
            if still:
                put("needs " + ", ".join(f"D{c + 1}" for c in still),
                    x + 20, y2, AMBER, 0.5)
        y += (len(open_rows) + 1) * 28

        # ── anything that matched nothing at all ──────────────────────────
        if window.unexpected:
            y += 24
            put("UNEXPECTED", 24, y, BG_BAD, 0.6)
            for i, (text, belongs) in enumerate(window.unexpected[-2:]):
                put(f"{tail(text, 44)}   -> {belongs}", 190, y + i * 24,
                    BG_BAD, 0.5)
        return img

    # ── what an earlier run already got through ──────────────────────────
    _load_previous()      # an .xlsx from an older run, if there is one
    _load_journal()       # then the journal, which wins where they differ
    _open_journal()
    if verified:
        last = max(verified)
        by_number = {r.number: i for i, r in enumerate(sheet.rows)}
        resume_hint[0] = by_number.get(last)
        done = sum(1 for _, st in verified.values() if st == "OK")
        print(f"[window] resuming from row {last} — {done} rows "
              f"already verified ({len(verified)} recorded)")

    def draw_window(frame, compact=False):
        """One status line on the video, since the detail now lives in its own
        window. `compact` skips the per-row breakdown."""
        font = cv2.FONT_HERSHEY_SIMPLEX
        x, y = 20, 120
        if window.start is None:
            cv2.putText(frame, "WINDOW: waiting for a known code", (x, y),
                        font, 0.8, (0, 255, 255), 2, cv2.LINE_AA)
            return frame

        passed = sum(1 for _, ok in window.done if ok)
        failed = len(window.done) - passed
        head = sheet.rows[window.start].number if not window.exhausted else "-"
        cv2.putText(frame, f"WINDOW  {passed} pass / {failed} short"
                           f"   reads {window.reads}   head row {head}",
                    (x, y), font, 0.8,
                    (0, 0, 255) if failed else (0, 255, 0), 2, cv2.LINE_AA)
        if compact:
            return frame

        for i, row_idx in enumerate(window.rows()[:8]):
            row_no = sheet.rows[row_idx].number
            required = window.required(row_idx)
            seen = window.seen.get(row_idx, set())
            marks = []
            for col in range(len(sheet.rows[row_idx].texts)):
                if col not in required:
                    marks.append("-")
                elif col in seen:
                    marks.append("OK")
                else:
                    marks.append("..")
            head = ">" if i == 0 else " "
            colour = ((0, 255, 0) if required and required <= seen
                      else (255, 255, 255) if i else (0, 255, 255))
            cv2.putText(frame, f"{head} row {row_no:<5} "
                               + "  ".join(f"D{c+1}:{m}" for c, m in enumerate(marks)),
                        (x, y + 32 + i * 30), font, 0.6, colour, 2, cv2.LINE_AA)
        return frame

    def scan_frame(frame, dets):
        """Decode every label in the frame and offer each payload to the
        window. No trigger line: a label is read on whichever frame it happens
        to be legible in, and the window decides whether it belongs here.
        """
        if label_cls is None:
            return

        qr_dets = [d for d in dets if int(d[5]) == qr_cls]
        settled = list(handled[0])       # labels dealt with on the last frame
        handled[0] = []
        budget = args.max_decodes or None

        for det in dets:
            if int(det[5]) != label_cls:
                continue

            # A label sits in view for many frames and would otherwise be
            # decoded on every one of them. Once its code has been accepted
            # there is nothing more to learn from it, so carry the box forward
            # by overlap and skip it — this is most of the per-frame cost.
            box = det[:4]
            if any(_overlap(box, prev) > 0.45 for prev in settled):
                handled[0].append(tuple(float(v) for v in box))
                continue

            if budget is not None:
                if budget <= 0:
                    continue         # rest of the labels wait for next frame
                budget -= 1

            qr = pick_qr_for_label(det[:4], qr_dets)
            text, box = decode_qr(frame, det[:4], margin=0.0, min_px=0)
            if not text and qr is not None:
                text, box = decode_qr(frame, qr[:4], args.qr_margin,
                                      args.qr_margin_min)
            if not text:
                continue

            # The first payload the sheet recognises decides where the window
            # sits; until then there is nothing to hold anything against.
            if window.start is None:
                hit = sheet.find(text, near=resume_hint[0])
                if hit is None:
                    continue
                anchored = window.anchor(hit[0])
                note = ("" if resume_hint[0] is None else
                        f", carrying on from row "
                        f"{sheet.rows[resume_hint[0]].number}")
                print(f"[window] anchored on sheet row {anchored}"
                      f"{note} — window covers {window.size} rows from there")

            status, row_idx, col, slot = window.offer(text)

            if status == RollingWindow.UNKNOWN:
                # A real code that belongs to no row inside the window: either
                # the wrong label is on the web, or the window has been left
                # far behind. Either way it is not something to tolerate.
                where = sheet.find(text)
                belongs = (f"sheet row {sheet.rows[where[0]].number} "
                           f"QR DATA{where[1] + 1}" if where else
                           "nothing in the sheet")
                head = sheet.rows[window.start].number if not window.exhausted else "?"
                print(f"\n[window] UNEXPECTED code: {text}")
                print(f"[window]   belongs to {belongs}; window starts at row {head}")
                window.note_unexpected(text, belongs)
                if saver is not None:
                    saver.save(frame, det[:4], text)
                stop_machine(f"unexpected code ({belongs})")
                # Carried forward like an accepted label so the same offending
                # code is reported once while it is in view, not every frame.
                handled[0].append(tuple(float(v) for v in det[:4]))
                continue

            if status == RollingWindow.REPEAT:
                handled[0].append(tuple(float(v) for v in det[:4]))
                continue

            if status == RollingWindow.MATCH:
                handled[0].append(tuple(float(v) for v in det[:4]))
                row_no = sheet.rows[row_idx].number
                if args.debug:
                    print(f"[window] row {row_no} QR DATA{col + 1} ok "
                          f"(slot {slot})")
                if saver is not None:
                    saver.save(frame, det[:4], text)

                # A hit in the last slot means the web has run a full window
                # past the head, so the head is not going to fill in on its
                # own. Rather than write it off, hold the line and keep the
                # row open: the operator winds the coil back, the same labels
                # come past again, and the row either completes or is judged
                # a real defect.
                if slot >= window.size - 1 and recheck["row"] is None:
                    head_idx = window.start
                    if head_idx is not None and window.missing(head_idx):
                        _hold_head_for_recheck(head_idx)

                for done_idx, _ in window.advance():
                    if recheck["row"] == sheet.rows[done_idx].number:
                        print(f"\n[recheck] row {sheet.rows[done_idx].number} "
                              f"PASSED on re-inspection — every code read, "
                              f"press START to carry on")
                        recheck["row"] = None
                        recheck["attempt"] = 0
                    retire_row(done_idx, complete=True)

    def _hold_head_for_recheck(row_idx):
        """The head row came up short. Stop, and keep it open.

        Nothing is written off on the first look. The row stays at the head
        of the window, still matchable, so that when the operator winds the
        coil back its codes tick off exactly as they would have the first
        time. Reading clean clears it; pressing START while it is still short
        is what turns it into a confirmed defect.
        """
        row_no = sheet.rows[row_idx].number
        missing = sorted(window.missing(row_idx))
        cols = ", ".join(f"QR DATA{c + 1}" for c in missing)
        recheck["row"] = row_no
        recheck["attempt"] = 1
        print(f"\n[recheck] row {row_no} did not validate: never read {cols}")
        for col in missing:
            print(f"[recheck]   QR DATA{col + 1}: expected "
                  f"'{sheet.rows[row_idx].texts[col]}'")
        print(f"[recheck] STOPPED — wind the coil back so row {row_no} passes "
              f"the camera again; its codes will be picked up automatically.")
        print(f"[recheck] then press START: if it read clean the run carries "
              f"on, if not the row is recorded as a label defect.")
        if args.no_stop_on_fail:
            # Nothing is going to come and look at this row, so writing it off
            # is the only way the window can keep moving.
            recheck["row"] = None
            recheck["attempt"] = 0
            stale = window.evict_head()
            if stale is not None:
                retire_row(stale, complete=False)
            return
        stop_machine(f"row {row_no} short of {cols} — held for re-check")

    def _adjudicate_recheck():
        """Called when the operator presses START with a row held open.

        By now the coil has been wound back and the row has had its second
        look. If it filled in, the retire path above has already cleared it
        and there is nothing here to do. If it is still short, the labels have
        been shown to the camera twice and would not read either time — that
        is a defect, not a vision hiccup.
        """
        row_no = recheck["row"]
        if row_no is None:
            return
        row_idx = next((i for i in window.rows()
                        if sheet.rows[i].number == row_no), None)
        if row_idx is None:                  # already retired some other way
            recheck["row"] = None
            return

        missing = sorted(window.missing(row_idx))
        recheck["row"] = None
        recheck["attempt"] = 0
        if not missing:
            return                           # filled in; the retire path has it

        print(f"\n[recheck] row {row_no} is STILL short after re-inspection")
        print(f"[defect]  LABEL HAS ISSUE — sheet row {row_no}")
        for col in missing:
            print(f"[defect]    QR DATA{col + 1}: expected "
                  f"'{sheet.rows[row_idx].texts[col]}'")
            print(f"[defect]    {' ' * 9}  read     nothing")
        defects.add(row_no)
        record_row(row_idx, complete=False, status="DEFECT")
        if results is not None:
            results.write(_window_result(row_idx, False))
        while (window.start is not None and window.start <= row_idx
               and not window.exhausted):
            window.evict_head()
        head = (sheet.rows[window.start].number
                if window.start is not None and not window.exhausted
                else "end of sheet")
        print(f"[defect]  accepted — window moves on to row {head}")

    def retire_row(row_idx, complete):
        """A sheet row has left the window. Record it and log the verdict.

        A row that came up short normally never reaches here — it is held open
        by _hold_head_for_recheck and settled by _adjudicate_recheck when the
        operator presses START. The short path below is only reached under
        --no-stop-on-fail, the one setting that lets a row be written off
        without a human looking at it.
        """
        row_no = sheet.rows[row_idx].number
        missing = sorted(window.missing(row_idx))
        record_row(row_idx, complete)
        if args.xlsx_every and len(verified) % args.xlsx_every == 0:
            write_annotated_xlsx(background=True)   # off unless asked for
        if complete and not missing:
            print(f"[window] PASS row {row_no}: all "
                  f"{len(window.required(row_idx))} codes read")
            if results is not None:
                results.write(_window_result(row_idx, True))
            return

        cols = ", ".join(f"QR DATA{c + 1}" for c in missing)
        print(f"[window] FAIL row {row_no}: never read {cols}")
        if results is not None:
            results.write(_window_result(row_idx, False))

    class _WindowResult:
        """Shaped like a BatchResult so the existing CSV writer can take it."""
        def __init__(self, row, entries, ok):
            self.row, self.entries, self.ok = row, entries, ok
        def summary(self):
            bad = [e.column for e in self.entries if e.status not in ("OK", "SKIPPED")]
            return f"row {self.row} missing {', '.join(bad)}" if bad else f"row {self.row}"

    class _WindowEntry:
        __slots__ = ("pos", "text", "expected", "status")
        def __init__(self, pos, text, expected, status):
            self.pos, self.text, self.expected, self.status = pos, text, expected, status
        @property
        def column(self):
            return f"QR DATA{self.pos + 1}"

    def _window_result(row_idx, ok):
        row = sheet.rows[row_idx]
        required = window.required(row_idx)
        seen = window.seen.get(row_idx, set())
        entries = []
        for col, expected in enumerate(row.texts):
            if col not in required:
                status = "SKIPPED"
            elif col in seen:
                status = "OK"
            else:
                status = "NO-READ"
            entries.append(_WindowEntry(col, expected if col in seen else None,
                                        expected, status))
        return _WindowResult(row.number, entries, ok)

    writer = None
    if args.save:
        writer = cv2.VideoWriter(args.save, cv2.VideoWriter_fourcc(*"mp4v"),
                                  args.fps, (disp_w, disp_h))

    if not args.no_display:
        # WINDOW_NORMAL makes the window resizable; without it, imshow opens
        # at the frame's native resolution which overflows most screens. We
        # set an initial on-screen size, capped to DISPLAY_MAX_*, while the
        # capture/inference itself still runs at full resolution.
        win_name = "Global Shutter Camera - YOLO26 TRT"
        cv2.namedWindow(win_name, cv2.WINDOW_NORMAL)
        scale = min(DISPLAY_MAX_W / disp_w, DISPLAY_MAX_H / disp_h, 1.0)
        cv2.resizeWindow(win_name, int(disp_w * scale), int(disp_h * scale))

        # The window is capped to DISPLAY_MAX_* anyway, so pushing the full
        # 5MP frame through imshow just makes the GUI thread scale it down
        # every frame. Do it here instead, with a cheap interpolation.
        def display_frame(frame):
            if scale >= 1.0:
                return frame
            return cv2.resize(frame, None, fx=scale, fy=scale,
                              interpolation=cv2.INTER_NEAREST)

        view = render_window_view()
        cv2.namedWindow(WINDOW_VIEW, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(WINDOW_VIEW, view.shape[1], view.shape[0])
        print(f"[window] comparison panel in the '{WINDOW_VIEW}' window")

        panel = ControlPanel(disp_w, disp_h,
                             on_start=lambda: start_machine("start button"),
                             on_stop=lambda: stop_machine("stop button"))
        # display_frame shrinks the frame by `scale` before imshow, and GTK
        # reports clicks in the coordinates of the image it was handed — so
        # scale them back up before hit-testing the buttons, which are laid
        # out against the full-resolution frame.
        def on_mouse(event, x, y, flags, param):
            panel.on_mouse(event, int(x / scale), int(y / scale), flags, param)

        cv2.setMouseCallback(win_name, on_mouse)
        print("[ui] START/STOP buttons in the window (keys: s = start, x = stop)")

    try:
        # Deliberately not started here. Nothing is decoded and nothing is
        # held against the sheet until the operator presses START, which is
        # what makes the read-in below mean something: it validates the coil
        # at the position it is actually standing in.
        print("[ui] idle — press START (or 's') to validate and run")

        prev_t = time.time()
        fps = 0.0
        while True:
            ok, frame = cap.read()
            if not ok:
                print("[camera] frame grab failed, retrying...")
                continue

            if rotate_in_loop:
                frame = rotate_frame(frame, rotate_in_loop)

            # ── inference ────────────────────────────────────────────────
            inp, ratio, pad = preprocess(frame, model.input_size)
            raw = model.infer(inp)
            dets = postprocess(raw, ratio, pad, frame.shape, args.conf_thres,
                               conf_per_class)

            # ── decode and validate ──────────────────────────────────────
            # Only while the machine is running, or in the read-in after START
            # was pressed. Stopped, the detector still draws boxes but nothing
            # is decoded or held against the sheet — the operator is handling
            # the coil, and whatever drifts past the lens is not a verdict.
            if validating():
                scan_frame(frame, dets)
                if (starting_at[0] is not None
                        and time.time() - starting_at[0] >= args.start_delay):
                    _finish_start()

            frame = draw_detections(frame, dets, class_names)
            draw_window(frame, compact=True)
            if starting_at[0] is not None:
                left = max(args.start_delay - (time.time() - starting_at[0]), 0)
                cv2.putText(frame, f"READING LABELS… {left:.1f}s",
                            (20, 156), cv2.FONT_HERSHEY_SIMPLEX, 0.9,
                            (0, 200, 255), 2, cv2.LINE_AA)
            elif not machine_running:
                cv2.putText(frame, "IDLE — press START to validate and run",
                            (20, 156), cv2.FONT_HERSHEY_SIMPLEX, 0.8,
                            (150, 150, 150), 2, cv2.LINE_AA)
            if panel is not None:
                panel.draw(frame, machine_running)

            # Simple running FPS: time between consecutive frames, smoothed
            # with an exponential moving average so the readout doesn't
            # jitter frame-to-frame.
            now = time.time()
            dt = now - prev_t
            prev_t = now
            if dt > 0:
                inst_fps = 1.0 / dt
                fps = inst_fps if fps == 0.0 else (0.9 * fps + 0.1 * inst_fps)

            if args.no_display:
                passed = sum(1 for _, ok in window.done if ok)
                held = f"  HELD row {recheck['row']}" if recheck["row"] else ""
                if starting_at[0] is not None:
                    left = args.start_delay - (time.time() - starting_at[0])
                    held += f"  READ-IN {max(left, 0):.1f}s"
                elif not machine_running:
                    held += "  IDLE"
                print(f"[camera] frame {frame.shape}  fps={fps:.1f}  "
                      f"dets={len(dets)}  pass={passed} "
                      f"short={len(window.done) - passed}  "
                      f"reads={window.reads}{held}   ", end="\r")
            else:
                cv2.putText(frame, f"FPS: {fps:.1f}  dets: {len(dets)}", (20, 40),
                            cv2.FONT_HERSHEY_SIMPLEX, 1.0, (0, 255, 0), 2, cv2.LINE_AA)
                cv2.imshow(win_name, display_frame(frame))
                cv2.imshow(WINDOW_VIEW, render_window_view())

            if writer:
                writer.write(frame)

            if not args.no_display:
                key = cv2.waitKey(1) & 0xFF
                if key == ord("q"):
                    break
                if panel is not None:
                    panel.on_key(key)
    finally:
        stop_machine("shutting down")
        relay.close()
        cap.release()
        if writer:
            writer.release()
        if not args.no_display:
            cv2.destroyAllWindows()
        if journal[0] is not None:
            journal[0].close()
        if verified:
            for _ in range(60):          # let any background save finish first
                with xlsx_lock:
                    if not xlsx_busy[0]:
                        break
                time.sleep(0.1)
            write_annotated_xlsx()
        if results is not None:
            print(f"[results] {results.rows} rows written to {results.path}")
            results.close()
        if saver is not None:
            print(f"[crops] saved {saver.count} label crops to {saver.dir}/")
        if defects:
            print(f"[defect] {len(defects)} row(s) failed twice and were "
                  f"recorded as LABEL ISSUE: "
                  + ", ".join(str(r) for r in sorted(defects)))
        passed = sum(1 for _, ok in window.done if ok)
        print(f"[validate] done: {passed} rows passed, "
              f"{len(window.done) - passed} short, "
              f"{window.reads} codes read")


if __name__ == "__main__":
    main()
