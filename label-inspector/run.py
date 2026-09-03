#!/usr/bin/env python3
"""
Global Shutter Camera + YOLO26 TensorRT + rolling-window QR validation.

A single-purpose build of run.py: rolling-window mode only. Every QR the
detector finds anywhere in the frame is decoded once and offered, on its own,
to a window of the next few sheet rows. Because every payload in the sheet is
unique a code identifies itself — which row and which QR DATA column it
belongs to — so the order codes happen to decode in never matters.

Nothing is tolerated. A row that reaches the end of the window still missing
a code does not get written off: the line stops and the row is held open at
the head of the window.

Stopping on a fault puts the app into REWIND. The relay is off, but the camera
keeps decoding, so the operator can wind the coil backwards by hand and watch
the labels that caused the stop come past again on screen. What is wrong is
drawn on the frame — the held row's positions ticking from NOT READ YET to
READ as they come, or the offending label outlined in red — so the console
text and the coil in front of the operator are talking about the same thing.

The fault clears itself, and the machine restarts itself, the moment the
reason for it is gone:

    a held row reads every code       ->  cleared, relay back on, run carries
                                          on where it left off
    an unexpected code goes out of    ->  cleared after --rewind-clear seconds
    frame and stays out                   (once per code: if the same one
                                          stops the line twice it waits for a
                                          human)

The model finds three things on the web: the label, the qr_code printed on it
and the logo printed on it. All three are used. The code is what is cropped --
with a quiet zone round it -- and handed to the reader; the logo is never
decoded, it only has to be there. A label the model finds without one of its
parts stops the line, before anything is read, because a label with no code
cannot be validated at all and would otherwise pass for a row that merely came
up short. It takes --part-looks clear looks at one label, every one of them
missing the same part, while its neighbours have it -- a dropped box is not a
missing logo. --no-part-check turns the whole thing off.

A label that will not read at all moves no part of this machine -- it is not
a match, not a repeat, not an unexpected code, and it neither fills a cell nor
empties one. A roll printed without codes, or a lens that has been knocked,
would therefore wind through as a quiet, faultless run. --no-read-secs is the
watchdog for that: labels in frame and not one code out of any of them stops
the line, and one label reading starts it again.

One fault does not clear itself, because nothing on the coil can clear it: if
no code read has been in the sheet at all, the roll on the machine is not the
one the sheet describes -- a new roll went on and the sheet was not changed
with it. The line stops before the relay is ever energised, the console says
WRONG SHEET, and LOAD SHEET stays live so the right one can be loaded.

Pressing START while the fault is still live is the operator overruling it:

    row still short                   ->  LABEL HAS ISSUE, recorded as a
                                          defect, and the window moves past it
    unexpected code still there       ->  accepted, and that payload will not
                                          stop the line again this run
    wrong sheet for the roll          ->  that code is accepted; the next one
                                          that is in no row stops it again

Nothing is read while the machine is idle. Pressing START does not energise
the relay straight away: for --start-delay seconds the camera reads the labels
standing in front of it while the coil is still, and only then does the relay
go on. That read-in is what validates the position the coil is actually in,
and it is also the last look a held row gets — starting the web first would
drag those labels out of frame before they could be checked.

What this build does NOT do, compared to run.py: no trigger-line mode, no
label grouping, and so no check of *where* a label sat. Four correct codes in
the wrong four positions tick four correct cells and the row passes. Use
run.py --mode row if position matters.

The window itself is the operator console: START and STOP, the sheet and the
folders down the right, and buttons to load a sheet, reopen one that was
loaded before, or point the label crops somewhere else. Those three are only
live with the machine idle, because each opens a new record and closes the
old one. The crop folder and the recently loaded sheets are remembered
between runs (utils/settings.py); the record itself — the checked .xlsx, the
progress journal and the run log — always stays inside the project. The
counters, the window's check status and the decoder tally are diagnostics and
only appear with --debug, or when 'd' is pressed.

No sheet is loaded at start-up, not even the one loaded last: which roll is
on the machine is something only the operator knows, and a roll checked
against somebody else's paperwork is worse than one not checked at all. START
is dead until a sheet has been chosen.

The winder's AUTO/MANUAL selector is mirrored above START, and the one relay
this app drives follows it.

    MANUAL   the contact closes and stays closed. The winder runs on its own
             controls and this app stands out of the way -- it still watches
             and still reads, but it starts and stops nothing, so START is
             dead.
    AUTO     the contact is the console's. It opens on the way in and closes
             only when START has been pressed and the labels standing in
             front of the camera have been read.

Thrown to MANUAL mid-run, the contact does not open on the way to closing:
it is a maintained start input, and a gap in it is a pulse, which is the one
edge a motor starter is built to act on.

It opens in AUTO, which is the position that leaves the contact open, so
launching the app energises nothing. The choice is never remembered between
runs: MANUAL closes a start contact, and an app that did that on the
strength of something chosen yesterday is the hazard the switch removes.

Under it sits the other selector, CHECK FORWARD / CHECK REVERSE. The printer
lays the sheet down the reel in order, but a reel wound onto a second spool
comes off it the other way round -- the last row printed is the first one
past the camera -- so REVERSE walks the window from the row it anchors on
back toward row one. Only the direction changes: a payload still identifies
its own row and column, which is what made the matching indifferent to
arrival order to begin with. It can only be thrown while the line is idle,
because it decides where the pass goes from here.

Ctrl+Alt+E opens the camera: exposure, gain and brightness, stepped one at
a time, applied as they change and remembered for next time
(utils/camera.py, straight to V4L2). What the model can find is decided by
the light in the frame, and a reel with more gloss or a lamp that has aged
is a thing the operator has to be able to answer at the machine.
Ctrl+Alt+W shows the diagnostics. Chords, not bare letters: this console
gets leaned on and wiped down, and a single keystroke that opens a window
over the live picture is one that will happen by accident.

What belongs to an *installation* rather than to the code -- which camera,
how big the picture is, which relay starts the motor, how long the read-in
runs -- is in config.json beside the application, written out with every
value in it on first run (utils/config.py). The command line still wins over
it. That file is what makes this packageable with PyInstaller: nothing that
changes from one machine to the next is compiled in, and the paths it
resolves know the difference between where the app was unpacked and where
the app actually lives.

That console is Qt (utils/qt_ui.py): the chrome is widgets and the overlay is
painted at display resolution, which is both quicker than burning it into a
5MP frame sixty times a second and far easier to read. Qt owns the main
thread, so the capture loop runs on a worker and posts a snapshot of drawing
instructions across; the console sends back nothing but named commands, which
the loop picks up at the top of its next pass. That way all the machine's
state stays on one thread. --ui opencv selects the older console drawn onto
the video, which is unchanged and still works.

Usage:
    python3 run.py                                  # defaults
    python3 run.py --xlsx validation.xlsx
    python3 run.py --window-size 4                  # shorter rewind
    python3 run.py --no-display                     # headless
    python3 run.py --no-relay                       # vision only
    python3 run.py --debug                          # the diagnostic readout

Relay. One, --start-relay 0, and it starts the winding machine. Nothing else
on the board is touched: a second contact that follows the motor, or a third
that follows a fault, is a thing to wire off the motor itself rather than a
thing for this app to hold an opinion about. Fewer coils under software
control is fewer coils that can be left closed by a crash.

It is closed only with the winder selector in AUTO, and opened the instant it
leaves it.

Sound. The operator's hands are on the coil and the screen is across the
machine, so the console says out loud what happened and what to do about it --
a tone and then the instruction on a fault, each position called as it comes
back in on the rewind, and the restart announced when the fault clears. The
voice is Microsoft's en-IN neural voice through edge-tts, female by default
(--voice-name male for Prabhat). Every phrase is rendered to a .wav the first
time it is used and replayed from disk after, so it costs nothing to say twice
and a machine that has run once keeps its voice with the network unplugged;
espeak is the fallback for one that has not. --no-voice runs it silent.
"""

import argparse
import os
import queue
import signal
import threading
import time
import cv2

from utils.camera import CameraControls
from utils.config import Config, app_dir
from utils.crops import LabelSaver
from utils.prepare import prepare as prepare_sheet
from utils.qr import decode_qr, decode_qr_at, decode_qr_pyzbar, \
    pick_qr_for_label, read_datamatrix
from utils.relay import RelayController
from utils.results import ResultLog
from utils.settings import Settings
from utils.trt_engine import YOLO26TRT
from utils.ui import ControlPanel
from utils.utils import preprocess, postprocess, draw_detections
from utils.voice import Voice
from utils.validation import ValidationSheet, normalize

# ── Everything an installation can be told ───────────────────────────────────
# Read from config.json beside the application, with these defaults behind it
# (utils/config.py holds them, and writes the file out on first run so an
# operator can see and edit every one of them). The command line still wins
# over both, so a shift can be run differently without editing anything.
#
# This is what makes the app packageable: nothing that has to change per
# installation -- which camera, which relay, how big the picture is -- is
# compiled in any more.
CFG = Config()
CAM, DISP, MODEL = CFG["camera"], CFG["display"], CFG["model"]
DECODE, MACHINE, RELAY = CFG["decode"], CFG["machine"], CFG["relay"]

# Where this app lives: the directory the application sits in, which under
# PyInstaller is not the same as the one the code was unpacked to. The
# record it writes belongs beside the application, not in whatever directory
# it happened to be launched from, and not in a temp folder that is deleted
# on exit.
APP_DIR             = app_dir()
# There is deliberately no default sheet. A sheet is the description of one
# roll, and a wrong one is worse than none: it would let the line run and
# validate the coil against somebody else's paperwork. So the app starts with
# nothing loaded and START stays dead until the operator says which sheet
# this roll is.
#
# Two roots, because the two things a run writes belong in different
# places. The record -- the checked .xlsx, progress.csv and the run logs
# -- is the project's own paperwork and stays beside the application. The
# label crops are bulk image data that fills a disk, so they go wherever
# the operator points them, and that choice is remembered between runs.
DEFAULT_RESULT_DIR  = CFG["paths"]["result_dir"] or "result"
DEFAULT_LABEL_DIR   = CFG.get_path(CFG["paths"]["label_dir"]) or \
    os.path.join(APP_DIR, "labels")
GS_CAMERA_NAME      = CAM["name"]
# Ups the voice is warmed up for before a sheet says how many there really
# are. The console can open with no sheet loaded, and rendering a phrase at
# the moment it is needed is the one thing the warm-up exists to avoid.
MAX_UPS             = MACHINE["max_ups"]

ROTATE_MAP = {
    0:   None,
    90:  cv2.ROTATE_90_CLOCKWISE,
    180: cv2.ROTATE_180,
    270: cv2.ROTATE_90_COUNTERCLOCKWISE,
}


# videoflip's equivalent of ROTATE_MAP, so the rotation can be done inside
# GStreamer instead of on the capture loop. cv2.rotate on a full 2592x1944
# BGR frame costs ~3.1ms of the ~21ms frame budget; videoflip does the same
# work on the GStreamer thread and the camera still delivers its full 60fps,
# so those 3.1ms come straight off the loop.
VIDEOFLIP_MAP = {
    0:   None,
    90:  "clockwise",
    180: "rotate-180",
    270: "counterclockwise",
}


def rotate_frame(frame, degrees):
    """Rotate a frame by a fixed angle (0/90/180/270). No-op for 0.

    Only used when the rotation could not be pushed into the pipeline; the
    normal path has videoflip deliver frames already rotated."""
    code = ROTATE_MAP[degrees]
    return frame if code is None else cv2.rotate(frame, code)


WINDOW_VIEW = "Rolling window"


def list_cameras():
    """List (index, name) for every /dev/videoN device via its v4l2 sysfs name."""
    cams = []
    for d in sorted(os.listdir("/dev")):
        if not d.startswith("video"):
            continue
        sys_path = f"/sys/class/video4linux/{d}/name"
        if not os.path.exists(sys_path):
            continue
        with open(sys_path) as f:
            name = f.read().strip()
        cams.append((int(d.replace("video", "")), name))
    return cams


def find_camera_index(name_substring=GS_CAMERA_NAME, default=0):
    """Find the /dev/videoN index whose v4l2 name contains name_substring."""
    cams = list_cameras()
    for index, name in cams:
        if name_substring.lower() in name.lower():
            return index
    print(f"[camera] '{name_substring}' not found among {cams} — "
          f"falling back to index {default}")
    return default


def gstreamer_pipeline(cam_index=0, width=None, height=None, fps=None,
                        format="MJPG", rotate=0):
    """Build a GStreamer pipeline string for v4l2src (MJPG or YUYV).

    `rotate` (0/90/180/270) is applied by videoflip inside the pipeline, so
    frames arrive already rotated and the loop never touches them."""
    QUEUE = "queue leaky=downstream max-size-buffers=1"
    method = VIDEOFLIP_MAP.get(rotate)
    FLIP  = f"videoflip method={method} ! " if method else ""
    SINK  = (f"{FLIP}videoconvert ! video/x-raw, format=BGR ! "
             "appsink drop=true max-buffers=1 sync=false")

    if format.upper() == "MJPG":
        return (
            f"v4l2src device=/dev/video{cam_index} ! "
            f"image/jpeg, width={width}, height={height}, framerate={fps}/1 ! "
            f"{QUEUE} ! jpegdec ! {SINK}"
        )
    else:
        return (
            f"v4l2src device=/dev/video{cam_index} ! "
            f"video/x-raw, width={width}, height={height}, framerate={fps}/1 ! "
            f"{QUEUE} ! {SINK}"
        )


def load_class_names(path):
    if not path:
        return None
    with open(path) as f:
        return [line.strip() for line in f if line.strip()]


class RollingWindow:
    """A sliding window of expected sheet rows, matched in any order.

    Line-crossing validation required the labels to arrive in sequence: a row
    was judged as its column passed a fixed point, and anything the camera
    missed became an out-of-sequence fault. Here the sequence constraint is
    dropped. Every QR the detector finds anywhere in the frame is offered to a
    window of the next `size` sheet rows; if the payload belongs to any cell
    inside that window, the cell is ticked off. Order does not matter, and a
    label can be read on whichever frame it happens to be legible in.

    A row leaves the window once every checked cell has been ticked, and the
    head slides one row at a time. A payload that belongs to no cell in the
    window is the fault worth stopping for — it means a label is on the web
    that the sheet does not expect here.

    Which way it slides is `step`. The printer lays the sheet down the reel
    in order, but a reel wound onto a second spool comes off it the other way
    round: the last row printed is the first one past the camera. So the
    window walks the sheet backwards, from the row it anchors on toward row
    one, and everything that means "further into the window" or "already
    behind us" is measured along that direction rather than up the row
    numbers. Nothing else changes -- a payload still identifies its own row
    and column, which is what made the matching indifferent to order in the
    first place.
    """

    MATCH, REPEAT, UNKNOWN = "match", "repeat", "unknown"
    UNREAD = "unread"          # a group went by and nothing on it decoded
    FORWARD, REVERSE = 1, -1

    def __init__(self, sheet, size=8, check=None, grace=4, step=FORWARD):
        self.sheet = sheet
        self.size = max(1, size)
        self.check = check
        self.step = self.REVERSE if step < 0 else self.FORWARD
        # Rows kept in the lookup *behind* the head. A label stays in view for
        # many frames and is decoded on every one, so codes keep arriving for
        # rows that already finished. Keeping those rows matchable lets such a
        # read be recognised as a repeat instead of being mistaken for the
        # next block's copy of the same code, or for a code from nowhere.
        self.grace = max(0, grace)
        self.start = None            # sheet row index at the head of the window
        self.seen = {}               # row index -> set of columns ticked off
        self.index = {}              # payload -> [(row index, column)]
        self.done = []               # [(row index, complete?)] as rows retire
        self.decoded_cells = set()   # (row index, column) ever decoded
        self.texts = {}              # (row index, column) -> payload as read
        self.last_hit = None         # (row index, column) most recently ticked
        self.unexpected = []         # recent payloads that belonged nowhere
        self.reads = 0
        self.repeats = 0

    # ── window contents ───────────────────────────────────────────────────
    def _in_sheet(self, row_idx):
        return 0 <= row_idx < len(self.sheet.rows)

    def depth(self, row_idx):
        """How far into the window a row sits, along the direction of travel.

        0 is the head, positive is ahead of it and still to come, negative is
        behind it and already retired. Everywhere this file used to subtract
        row numbers it asks this instead, which is the whole of what makes
        the reverse pass work.
        """
        return (row_idx - self.start) * self.step

    def rows(self):
        if self.start is None:
            return []
        out = []
        for i in range(self.size):
            row_idx = self.start + i * self.step
            if not self._in_sheet(row_idx):
                break                       # ran off the end of the sheet
            out.append(row_idx)
        return out

    def required(self, row_idx):
        """Columns that must be decoded before this row can retire."""
        n = len(self.sheet.rows[row_idx].texts)
        if self.check is None:
            return {c for c in range(n) if normalize(self.sheet.rows[row_idx].texts[c])}
        return {c for c in self.check if c < n}

    def missing(self, row_idx):
        return self.required(row_idx) - self.seen.get(row_idx, set())

    def _rebuild(self):
        """The payload lookup covers the window plus the grace rows behind it,
        which is what makes 'not expected here' meaningful. Behind means
        behind along the direction of travel, not down the row numbers."""
        self.index = {}
        behind = []
        if self.start is not None:
            behind = [self.start - i * self.step
                      for i in range(1, self.grace + 1)]
            behind = [r for r in behind if self._in_sheet(r)]
        for row_idx in behind + self.rows():
            for col, text in enumerate(self.sheet.rows[row_idx].texts):
                key = normalize(text)
                if key:
                    self.index.setdefault(key, []).append((row_idx, col))

    def anchor(self, row_idx):
        self.start = min(max(row_idx, 0), len(self.sheet.rows) - 1)
        self.seen = {}
        self._rebuild()
        return self.sheet.rows[self.start].number

    # ── matching ──────────────────────────────────────────────────────────
    def offer(self, text):
        """Tick off whatever cell this payload belongs to.
        Returns (status, row_index, column, slot) — slot is how deep into the
        window the hit landed, 0 being the head."""
        hits = self.index.get(normalize(text))
        if not hits:
            return self.UNKNOWN, None, None, None
        # Nearest the head, measured along the direction of travel: the rows
        # already retired sit at negative depth and so are found first,
        # which is what makes a label still in view read as a re-read.
        row_idx, col = min(hits, key=lambda h: self.depth(h[0]))
        # A datamatrix is printed several times over, and the working copy
        # gives each printing a row of its own, so one payload legitimately
        # holds a cell in each row of its group. The nearest one being
        # ticked already does not make this a re-read: it is the next label
        # along, carrying the same value on purpose. Take the first cell of
        # the group that is still open, so the second and third printings
        # are checked rather than swallowed — and so a printing that never
        # came leaves its row short, which is the whole point of the check.
        # The nearest copy is behind the head once the first printing has
        # retired, and ticked already before it has; both ask the same
        # question, which is whether a row of this group is still open.
        group = self.sheet.rows[row_idx].group
        if group is not None and (self.depth(row_idx) < 0
                                  or col in self.seen.get(row_idx, set())):
            for cand_row, cand_col in sorted(
                    hits, key=lambda h: (self.depth(h[0]), h[1])):
                if (self.depth(cand_row) >= 0
                        and self.sheet.rows[cand_row].group == group
                        and cand_col not in self.seen.get(cand_row, set())):
                    row_idx, col = cand_row, cand_col
                    break
            # Nothing open left in the group: every printing has been seen,
            # so this is the last of them still standing in front of the
            # lens. That falls through to the repeat path below, which is
            # what it is. Payload is all there is to go on here -- three
            # labels printed alike are alike -- so what keeps one label from
            # ticking all three rows is the same carry-forward by overlap
            # that stops any label being decoded twice.
        if self.depth(row_idx) < 0:
            self.repeats += 1        # a finished row being read again
            return self.REPEAT, row_idx, col, None
        slot = self.depth(row_idx)
        self.last_hit = (row_idx, col)
        if col in self.seen.get(row_idx, set()):
            self.repeats += 1
            return self.REPEAT, row_idx, col, slot
        self.seen.setdefault(row_idx, set()).add(col)
        self.decoded_cells.add((row_idx, col))
        self.texts[(row_idx, col)] = text
        self.reads += 1
        return self.MATCH, row_idx, col, slot

    # ── sliding ───────────────────────────────────────────────────────────
    def note_unexpected(self, text, belongs):
        """Keep the last few codes that matched nothing, for the view."""
        self.unexpected.append((text, belongs))
        del self.unexpected[:-4]

    def advance(self):
        """Retire finished rows from the head. Returns [(row index, True)]."""
        retired = []
        while self.start is not None and self._in_sheet(self.start):
            if self.missing(self.start):
                break
            retired.append((self.start, True))
            self.done.append((self.start, True))
            self.start += self.step
        if retired:
            self._rebuild()
        return retired

    def evict_head(self):
        """Drop the head row even though it is short — the web has moved a
        whole window past it, so it is never going to be completed."""
        if self.start is None or not self._in_sheet(self.start):
            return None
        row_idx = self.start
        self.done.append((row_idx, False))
        self.start += self.step
        self._rebuild()
        return row_idx

    @property
    def exhausted(self):
        """Walked off the end of the sheet -- whichever end that is."""
        return self.start is not None and not self._in_sheet(self.start)


def _overlap(a, b):
    """Intersection over union of two boxes — used to recognise the same
    physical label from one frame to the next."""
    ix = max(0.0, min(a[2], b[2]) - max(a[0], b[0]))
    iy = max(0.0, min(a[3], b[3]) - max(a[1], b[1]))
    inter = ix * iy
    if inter <= 0:
        return 0.0
    area_a = max(0.0, a[2] - a[0]) * max(0.0, a[3] - a[1])
    area_b = max(0.0, b[2] - b[0]) * max(0.0, b[3] - b[1])
    union = area_a + area_b - inter
    return inter / union if union > 0 else 0.0


def _clearance(label_box, symbol):
    """Least gap between a symbol and the edges of a label, in pixels.

    Negative when the symbol hangs outside. This is what tells the label a
    code is printed on from a label that has merely swallowed its
    neighbour's: a code on its own label sits inside it with room on every
    side, while the same code seen from the label next door is jammed hard
    against the edge it came over.
    """
    return min(symbol[0] - label_box[0], symbol[1] - label_box[1],
               label_box[2] - symbol[2], label_box[3] - symbol[3])


def _owns_symbol(label_box, symbol, label_dets):
    """Is this symbol printed on this label, rather than on its neighbour?

    Reading a whole label means reading whatever else falls inside its box,
    and the labels on this web stand shoulder to shoulder with the code near
    an edge. So a label with nothing printed on it can come back carrying
    the code off the label beside it -- and a blank label wearing a payload
    is worse than a blank label, because it passes.

    Whichever label holds the symbol furthest from its own edges is the one
    it is printed on. Nothing else in the running leaves it here, which is
    the ordinary case: one label, its own code, no argument.
    """
    if symbol is None:
        return True                # nothing to place; the crop is all we have
    mine = _clearance(label_box, symbol)
    for det in label_dets:
        other = tuple(float(v) for v in det[:4])
        if other == tuple(float(v) for v in label_box):
            continue
        if _clearance(other, symbol) > mine:
            return False
    return True


def web_motion(boxes, previous):
    """How far the web moved between two frames, in pixels: (dx, dy).

    Every label in the picture is on the same web and moves together, so the
    question is one number, not one per label -- and the median of what the
    labels individually say is the way to get it without a stray box or a
    label entering the frame pulling the answer about.

    Matched by nearest centre rather than by overlap, because at the speed
    this matters at the boxes no longer overlap -- which is exactly the
    case the whole thing exists for. A match further than half a label is
    refused: on a web of identical labels a pitch of travel looks precisely
    like no travel at all, and guessing wrong there is worse than saying
    nothing.
    """
    if not boxes or not previous:
        return 0.0, 0.0
    reach = max(min(b[2] - b[0] for b in boxes) * 0.5, 1.0)
    dxs, dys = [], []
    for x1, y1, x2, y2 in boxes:
        cx, cy = (x1 + x2) * 0.5, (y1 + y2) * 0.5
        best = None
        for px1, py1, px2, py2 in previous:
            dx = cx - (px1 + px2) * 0.5
            dy = cy - (py1 + py2) * 0.5
            far = abs(dx) + abs(dy)
            if best is None or far < best[0]:
                best = (far, dx, dy)
        if best is not None and best[0] <= reach:
            dxs.append(best[1])
            dys.append(best[2])
    if not dxs:
        return 0.0, 0.0
    dxs.sort()
    dys.sort()
    mid = len(dxs) // 2
    return dxs[mid], dys[mid]


def sheet_period(sheet, limit=400):
    """Smallest number of rows after which the sheet repeats itself, or None.

    A production sheet is often one block of codes duplicated many times over
    (validation_x300.xlsx is 17 rows copied 300 times). The window has to stay
    shorter than that block: if it spans a whole period, every payload appears
    in it twice and the matching cannot tell one pass from the next.
    """
    rows = [tuple(r.texts) for r in sheet.rows]
    n = len(rows)
    for period in range(1, min(limit, n // 2 + 1)):
        if n % period:
            continue
        if all(rows[i] == rows[i % period] for i in range(n)):
            return period
    return None


def parse_check(spec, per_row):
    """Turn "UP2,UP3" into the 0-based ups to check. None means all of them.

    "D2,D3" and a bare "2 3" are taken too: the ups were called D1..DN
    before, and a command line written then should still run.
    """
    if not spec:
        return None
    wanted = set()
    for part in spec.replace(";", ",").replace(" ", ",").split(","):
        part = part.strip().upper()
        part = part[2:] if part.startswith("UP") else part.lstrip("D")
        if not part:
            continue
        if not part.isdigit():
            raise SystemExit(f"--check: '{part}' is not an up number")
        n = int(part)
        if not 1 <= n <= per_row:
            raise SystemExit(f"--check: {up(n - 1)} is out of range — the sheet "
                             f"has {per_row} code columns")
        wanted.add(n - 1)
    if not wanted:
        return None
    return wanted


def up(col):
    """What a position across the web is called on the machine.

    The sheet's columns are headed QR DATA1..N and the code has always used
    D1..DN internally, but on the floor they are the ups -- UP1 is the first
    label across the web. One word for one thing, so the screen, the console
    and the spoken prompts all say what the operator says.
    """
    return f"UP{col + 1}"


def class_index(class_names, name, fallback):
    """Resolve a class name to its id; fall back to a fixed index when no
    classes.txt was supplied (or the name isn't in it)."""
    if class_names and name in class_names:
        return class_names.index(name)
    print(f"[qr] class '{name}' not found in --classes, using index {fallback}")
    return fallback


def main():
    ap = argparse.ArgumentParser(
        description="Global Shutter Camera + YOLO26 TensorRT + rolling-window "
                    "QR validation, with re-inspection on any short row.")
    # camera args
    ap.add_argument("--index", type=int, default=CAM["index"],
                     help="Force a /dev/videoN index (skips auto-detect).")
    ap.add_argument("--width", type=int, default=CAM["width"])
    ap.add_argument("--height", type=int, default=CAM["height"])
    ap.add_argument("--fps", type=int, default=CAM["fps"])
    ap.add_argument("--format", default=CAM["format"], choices=["MJPG", "YUYV"])
    ap.add_argument("--rotate", type=int, default=CAM["rotate"], choices=[0, 90, 180, 270],
                     help="Rotate every frame by a fixed angle (clockwise).")
    ap.add_argument("--ui", default="qt", choices=["qt", "opencv"],
                     help="which console. 'qt' is a real window: the chrome "
                          "is widgets rather than pixels burnt into the "
                          "frame, which is both quicker and far easier to "
                          "read. 'opencv' is the older console drawn onto "
                          "the video, kept as a fallback.")
    ap.add_argument("--fullscreen", action="store_true",
                     help="open borderless, filling the whole screen, with "
                          "no title bar. Without it the window opens "
                          "maximised, which fills the screen just the same "
                          "but keeps the title bar. F11 switches between "
                          "them, and Escape leaves full screen.")
    ap.add_argument("--unlock-window", action="store_true",
                     help="let the window close on a single click of its X, "
                          "the way an ordinary application does. By default "
                          "it will not close while the machine is running or "
                          "a row is held, and asks first even when idle -- "
                          "the close button sits a few pixels from nothing "
                          "in particular, and a stray click on it would stop "
                          "the line.")
    ap.add_argument("--no-display", action="store_true",
                     help="Just print FPS/detections instead of opening a window (headless).")
    ap.add_argument("--debug", action="store_true",
                     help="print every payload as it reads and where it sits in "
                          "the sheet; without it only the per-row verdicts print.")
    # inference args
    ap.add_argument("--engine", default=None,
                     help=f"path to the .engine file "
                          f"(default: {MODEL['engine']} beside the app). It is "
                          f"expected to detect three classes: the label, the "
                          f"qr_code inside it and the logo inside it.")
    ap.add_argument("--classes", default=None,
                     help="txt file, one class name per line "
                          "(default: classes.txt beside run.py, if present)")
    ap.add_argument("--conf-thres", type=float, default=MODEL["conf"],
                     help="confidence threshold for classes without their own.")
    ap.add_argument("--conf-label", type=float, default=MODEL["conf_label"],
                     help="confidence threshold for the label class — this is "
                          "what gates a trigger-line crossing.")
    ap.add_argument("--conf-qr", type=float, default=MODEL["conf_qr"],
                     help="confidence threshold for the qr class — this is what "
                          "gates which box gets cropped and decoded.")
    ap.add_argument("--imgsz", type=int, default=MODEL["imgsz"])
    ap.add_argument("--save", default=None, help="optional path to record annotated video")
    # qr decode args
    ap.add_argument("--label-class", default=MODEL["label_class"],
                     help="class name that triggers a decode when it crosses the line.")
    ap.add_argument("--qr-class", default=MODEL["qr_class"],
                     help="class name of the QR box that gets cropped and decoded.")
    ap.add_argument("--logo-class", default=MODEL["logo_class"],
                     help="class name of the logo printed on each label. It "
                          "is never decoded — what matters is that it is "
                          "there.")
    ap.add_argument("--part-looks", type=int, default=MACHINE["part_looks"],
                     help="how many clear looks at one label before deciding "
                          "a part is not printed on it. A clear look is a "
                          "frame in which that label was whole and well "
                          "inside the picture. The part has to be missing "
                          "from every one of them: found even once, the "
                          "label is settled and never looked at again. A "
                          "detector drops a box now and then, so no smaller "
                          "amount of evidence is worth stopping a line for.")
    ap.add_argument("--no-part-check", action="store_true",
                     help="do not stop when a label is missing its QR or its "
                          "logo. Off by default: a label with no code cannot "
                          "be validated at all, and one with no logo is a "
                          "misprint whatever its code says.")
    ap.add_argument("--qr-margin", type=float, default=DECODE["qr_margin"],
                     help="quiet zone around the qr box, as a fraction of its size.")
    ap.add_argument("--qr-margin-min", type=int,
                     default=DECODE["qr_margin_min"],
                     help="minimum quiet zone in pixels.")
    ap.add_argument("--zbar-fallback", type=int,
                     default=DECODE["zbar_fallback"],
                     help="how many labels per frame may fall back to zbar "
                          "after zxing has failed on them. zxing and zbar do "
                          "not fail on the same codes, and the gap is not "
                          "marginal: over 79 crops off this line that zxing "
                          "could not read at all, zbar read 79, with no false "
                          "decodes. At ~2.2ms a call it is cheap enough that "
                          "the cap is a safety net rather than a real "
                          "constraint. 0 = off.")
    ap.add_argument("--dump-crops", default=None,
                     help="directory to save what the camera saw whenever a "
                          "label was detected but would not decode: the label "
                          "crop, the tighter qr crop, and the whole frame at "
                          "the moment a row is held. Files touching the frame "
                          "edge are named -CLIPPED. This is the tool for "
                          "answering why a row came up short.")
    # label crop args
    ap.add_argument("--result-dir", default=DEFAULT_RESULT_DIR,
                     help="root for the record this run writes: "
                          "<result-dir>/<xlsx name>/ holds the checked .xlsx, "
                          "progress.csv and the run logs. Kept inside the "
                          "project, and a relative path is taken as relative "
                          "to it rather than to the working directory.")
    ap.add_argument("--label-dir", default=None,
                     help="root for the saved label crops, and nothing else: "
                          "<label-dir>/<xlsx name>/. This is the one the "
                          "operator sets from the console, and the last "
                          "choice is remembered between runs. "
                          f"Default: {DEFAULT_LABEL_DIR}.")
    ap.add_argument("--no-save-labels", action="store_true",
                     help="don't save a crop of each decoded label.")
    ap.add_argument("--label-format", default="jpg", choices=["jpg", "png"],
                     help="image format for the saved label crops.")
    ap.add_argument("--label-pad", type=float, default=0.0,
                     help="fixed padding around the saved label crop, as a "
                          "fraction of the box size, on every side. Only for "
                          "a run that wants one: left to itself the crop "
                          "takes half the gutter it can actually see either "
                          "side of the label, measured off the labels next to "
                          "it, which is what keeps a code the detection box "
                          "clipped from being cut out of the picture. Setting "
                          "this replaces that.")
    ap.add_argument("--label-pad-px", type=int, default=0,
                     help="least padding around a saved crop, in pixels, "
                          "whatever --label-pad works out to. Small boxes "
                          "need the margin most and get the least from a "
                          "fraction. Setting either one replaces the measured "
                          "margin described above.")
    # validation args
    ap.add_argument("--xlsx", default=None,
                     help="xlsx holding the expected QR DATA1..N sequence. "
                          "There is no default: without one the console opens "
                          "with no sheet and START is dead until LOAD SHEET "
                          "or OPEN RECENT SHEET has been used, because a roll "
                          "checked against the wrong paperwork is worse than "
                          "one not checked at all.")
    ap.add_argument("--sheet", default=None,
                     help="worksheet name inside --xlsx (default: the first one).")
    ap.add_argument("--reverse", action="store_true",
                     default=MACHINE["check_reverse"],
                     help="check the sheet from its last row toward its "
                          "first. A reel wound onto a second spool comes off "
                          "it the other way round, so the last row printed is "
                          "the first one past the camera and the window has "
                          "to walk backwards to follow it. The console's "
                          "CHECK FORWARD / CHECK REVERSE button is the same "
                          "switch, and can be thrown while the line is idle.")
    ap.add_argument("--dm-repeats", type=int, default=MACHINE["dm_repeats"],
                     help="how many times each DATA MATRIX value is printed "
                          "down the web. Every value in a DATA MATRIX<n> "
                          "column is printed on labels of its own, this many "
                          "in a row, after the QR labels of the row it is "
                          "listed against — so on load the sheet is copied "
                          "and each of those becomes this many rows of its "
                          "own. The copy is what the machine checks against; "
                          "the operator's file is never written to.")
    ap.add_argument("--labels-per-row", type=int, default=None,
                     help="ups across the web (default: the number of QR DATA "
                          "columns found in the sheet).")
    ap.add_argument("--check", default=None,
                     help="which ups to validate, across the web, e.g. "
                          "'UP2,UP3' or '2,3' (D2,D3 still works). The rest "
                          "are neither decoded nor held against the row. "
                          "Default: all of them.")
    ap.add_argument("--no-stop-on-fail", action="store_true",
                     help="keep the machine running when a row fails "
                          "validation (default: stop it).")
    ap.add_argument("--no-result-log", action="store_true",
                     help="don't write the per-row CSV of verdicts.")
    ap.add_argument("--window-size", type=int,
                     default=MACHINE["window_size"],
                     help="how many sheet rows the rolling window holds. "
                          "Bigger tolerates more out-of-order arrival and more "
                          "missed rows; smaller catches a stray label sooner. "
                          "This is how far a code may be recognised from, not "
                          "how long a short row is tolerated — that is "
                          "--hold-after.")
    ap.add_argument("--hold-after", type=int,
                     default=MACHINE["hold_after"],
                     help="how far past an incomplete row the coil may move "
                          "before the line stops and that row is held open "
                          "for a re-check, in sheet rows. 1 stops at the very "
                          "next row; raise it if labels from several rows are "
                          "in view at once, so a row's last code has time to "
                          "arrive. Capped at --window-size - 1.")
    ap.add_argument("--window-grace", type=int, default=None,
                     help="rows kept matchable behind the window, so a label "
                          "still in view after its row finished is recognised "
                          "as a re-read instead of an unexpected code. "
                          "Defaults to --window-size, which is the least that "
                          "is safe: a held row healing on a rewind can retire "
                          "every row behind it at once, and the labels that "
                          "just retired are still standing in front of the "
                          "lens.")
    ap.add_argument("--out-xlsx", default=None,
                     help="where to write the sheet annotated with what was "
                          "decoded. Defaults to a fixed path per sheet so it "
                          "can be resumed; the source --xlsx is never touched.")
    ap.add_argument("--xlsx-every", type=int, default=0,
                     help="also rewrite the annotated .xlsx every N rows. "
                          "Costly (openpyxl is pure Python and stalls the "
                          "capture loop), and not needed for durability — "
                          "progress is journalled continuously either way. "
                          "0 = write the xlsx at exit only.")
    ap.add_argument("--max-decodes", type=int, default=0,
                     help="cap how many labels are decoded per frame. 0 = no "
                          "cap. Labels already read are skipped regardless.")
    ap.add_argument("--no-resume", action="store_true",
                     help="ignore the annotated sheet from a previous run and "
                          "start the record empty, anchoring wherever the "
                          "sheet first matches. By default a previous run is "
                          "picked up and the window continues past it.")
    # relay args
    ap.add_argument("--no-relay", action="store_true",
                     help="don't touch the relay board (vision only).")
    ap.add_argument("--relay-port", default=RELAY["port"],
                     help="serial port of the relay board (default: auto-detect).")
    ap.add_argument("--start-delay", type=float, default=MACHINE["start_delay"],
                     help="seconds to spend reading the labels already in "
                          "front of the camera after START is pressed, before "
                          "the relay is switched on. This is what validates "
                          "the coil at the position it is actually in — and "
                          "what re-checks a row that was held open. 0 = "
                          "energise immediately.")
    ap.add_argument("--start-relay", type=int, default=RELAY["start"],
                     help="the one relay this app drives. In AUTO it is "
                          "the console's, closed for a validated run and "
                          "open otherwise; in MANUAL it is closed and "
                          "stays closed, and the winder runs on its own "
                          "controls.")
    ap.add_argument("--relay-verbose", action="store_true",
                     help="print every modbus frame sent to the relay board.")
    ap.add_argument("--no-voice", action="store_true",
                     help="do not speak. By default the console says what "
                          "happened and what to do about it, because the "
                          "operator is at the coil rather than at the screen.")
    ap.add_argument("--voice-engine", default=CFG["voice"]["engine"],
                     choices=["auto", "edge", "espeak"],
                     help="how the prompts are spoken. 'edge' is Microsoft's "
                          "en-IN neural voice, which sounds like a person; "
                          "'espeak' is the offline formant synthesiser, which "
                          "does not. 'auto' uses edge when it can be reached "
                          "or has been cached, and espeak when it cannot.")
    ap.add_argument("--voice-name", default=CFG["voice"]["name"],
                     help="which voice: female (Neerja), male (Prabhat), "
                          "expressive, or any edge-tts voice name in full.")
    ap.add_argument("--voice-rate", type=int, default=CFG["voice"]["rate"],
                     help="speaking rate as a percentage off normal, e.g. 15 "
                          "for a little quicker, -10 for slower.")
    ap.add_argument("--no-tone", action="store_true",
                     help="do not sound the alert tone before a fault is "
                          "announced.")
    ap.add_argument("--no-auto-restart", action="store_true",
                     help="after a fault, wait for START even once the fault "
                          "has cleared on the rewind. By default the machine "
                          "restarts itself the moment the reason it stopped "
                          "is gone.")
    ap.add_argument("--no-read-secs", type=float,
                     default=MACHINE["no_read_secs"],
                     help="stop the line if labels have been in front of the "
                          "camera for this many seconds and not one of them "
                          "has read. Everything this app decides is driven by "
                          "a payload, so a label that decodes to nothing "
                          "exercises nothing: a roll printed without codes, a "
                          "lens that has been knocked or a light that has "
                          "failed would otherwise wind through as a quiet, "
                          "faultless run. 0 turns the watchdog off.")
    ap.add_argument("--rewind-clear", type=float,
                     default=MACHINE["rewind_clear"],
                     help="seconds an unexpected code must stay out of frame "
                          "during a rewind before the fault counts as "
                          "cleared and the machine restarts itself.")
    args = ap.parse_args()
    if args.window_grace is None:
        args.window_grace = args.window_size

    # ── what the console remembered from last time ───────────────────────
    # The command line always wins; the settings file only fills in what was
    # left out, which is the usual case because the operator starts this from
    # a desktop icon and sets everything from the screen.
    prefs = Settings()
    args.result_dir = os.path.join(APP_DIR, args.result_dir)
    if args.label_dir is None:
        args.label_dir = prefs.label_dir or DEFAULT_LABEL_DIR
    else:
        prefs.remember_label_dir(args.label_dir)
    # The sheet that was loaded last is offered, not reopened. Which roll is
    # on the machine is something only the operator knows, and a sheet that
    # loads itself is one nobody chose: the console lists it under OPEN
    # RECENT SHEET and waits to be told.
    if args.xlsx is None and prefs.recent:
        print(f"[settings] {len(prefs.recent)} sheet(s) loaded before — the "
              f"most recent is {prefs.sheet}")

    # ── the expected sheet, and the window over it ───────────────────────
    # Loaded through a function rather than inline, because the operator can
    # load a different sheet from the console without restarting the app.
    # All None until one is chosen: nothing is read, nothing is recorded and
    # START does nothing while there is no sheet to check the roll against.
    sheet = window = per_row = checked = None
    work_xlsx = [None]           # the expanded copy the window runs against
    # How many rows the coil may move past a row that has not read everything
    # before the line is stopped for it. Kept here rather than read straight
    # off args because the window can be shrunk by a repeating sheet, and this
    # has to stay inside it.
    hold_after = [1]
    # Which way the window walks the sheet. Set from the console's CHECK
    # FORWARD / CHECK REVERSE button, and only while the line is idle: it
    # decides where the pass goes from here, so throwing it mid-run would
    # send the window off in the opposite direction from the coil.
    check_step = [RollingWindow.REVERSE if args.reverse
                  else RollingWindow.FORWARD]

    def checking_reverse():
        return check_step[0] == RollingWindow.REVERSE

    def _open_sheet():
        """Read args.xlsx and build the rolling window over it.

        What the window actually runs against is the working copy: the same
        sheet with every DATA MATRIX value expanded into rows of its own, one
        per printing, because those are physical rows of labels on the coil
        that the operator's sheet has nowhere to put. utils/prepare.py writes
        it, beside the record for this sheet, and never touches the original.
        """
        nonlocal sheet, per_row, checked, window
        run_dir = os.path.join(
            args.result_dir,
            os.path.splitext(os.path.basename(args.xlsx))[0])
        try:
            work_xlsx[0] = prepare_sheet(args.xlsx, run_dir, sheet=args.sheet,
                                         repeats=args.dm_repeats)
        except Exception as exc:
            # Never fatal: this runs on the capture thread, from a button
            # press, and a sheet that cannot be expanded must not take the
            # console down with it. Running against the sheet unexpanded is
            # the safe direction — the datamatrix labels then belong to no
            # row, which stops the line rather than passing them.
            print(f"\n[prepare] could not expand {args.xlsx}: {exc}")
            print(f"[prepare]   running against it unexpanded — if it has "
                  f"datamatrix rows the line will stop at them as unexpected "
                  f"codes")
            work_xlsx[0] = args.xlsx
        sheet = ValidationSheet(work_xlsx[0], args.sheet)
        # The reader only looks for a datamatrix on a roll that has one: it
        # is another format for zxing to try on every crop that will not
        # read, and most rolls carry none.
        if read_datamatrix(sheet.has_dm):
            print(f"[qr] the reader is {'now' if sheet.has_dm else 'no longer'}"
                  f" looking for a datamatrix as well as a QR")
        per_row = args.labels_per_row or sheet.per_row
        checked = parse_check(args.check, per_row)
        if checked is None:
            print(f"[validate] checking all {per_row} positions")
        else:
            on = ", ".join(up(i) for i in sorted(checked))
            off = ", ".join(up(i) for i in range(per_row)
                            if i not in checked)
            print(f"[validate] checking {on}"
                  + (f" (ignoring {off})" if off else ""))

        size, grace = args.window_size, args.window_grace
        period = sheet_period(sheet)
        if period:
            # A window that spans the sheet's repeat would hold the same
            # payload twice over, and a code could tick off either copy.
            room = period - grace - 1
            if size > room:
                print(f"[window] the sheet repeats every {period} rows, so a "
                      f"window of {size} would see each code twice — "
                      f"using {max(1, room)} instead")
                size = max(1, room)
        window = RollingWindow(sheet, size=size, check=checked,
                               grace=grace, step=check_step[0])
        hold_after[0] = max(1, min(args.hold_after, window.size - 1))
        # Only once it has parsed: the recent list is for sheets that worked.
        prefs.remember_sheet(args.xlsx)
        print(f"[window] rolling window of {window.size} sheet rows "
              f"(+{window.grace} kept behind for re-reads), walking the "
              f"sheet "
              + ("BACKWARDS, last row first — the reel is wound over"
                 if checking_reverse() else "forwards, first row first"))

    if args.xlsx:
        try:
            _open_sheet()
        except Exception as exc:
            # A sheet named on the command line that will not open -- moved,
            # deleted, on a stick that is not in, not a workbook at all. That
            # is no longer a reason to die: opening with no sheet is a state
            # this console has, and the operator can load the right one from
            # the buttons without being sent back to a terminal.
            print(f"\n[window] cannot use {args.xlsx}: {exc}")
            print(f"[window]   starting with no sheet — LOAD SHEET, or OPEN "
                  f"RECENT SHEET, for the one this roll needs")
            args.xlsx = None
            sheet = window = None
    if args.xlsx:
        print(f"[window] a row that comes up short stops the line and is held "
              f"open — wind the coil back and the screen shows what is still "
              f"missing; fill it and the machine starts itself")
        print(f"[window] the line stops as soon as the coil has moved "
              f"{hold_after[0]} row(s) past a row that did not read everything")
        if args.no_auto_restart:
            print(f"[window] --no-auto-restart: a cleared fault still waits "
                  f"for START")
    else:
        print(f"[window] no sheet loaded — LOAD SHEET for this roll's xlsx, "
              f"or OPEN RECENT SHEET for one that has been run before. "
              f"Nothing is read and START does nothing until then.")

    def loaded():
        """Is there a sheet to check the roll against?"""
        return window is not None

    machine_running = False
    # The winder's mode. In AUTO the start relay is the console's, closed
    # when a run has been validated into life and open otherwise; in MANUAL
    # it is closed and stays closed, and the machine runs on its own
    # controls with this app out of the way.
    #
    # None until somebody says. Not False: MANUAL means a closed contact on
    # a maintained start input, and this app is not going to make that
    # decision on its own at launch -- it may be launching from a desktop
    # shortcut, or on boot, with nobody yet standing at the machine. So it
    # starts with the relay open, shows MANUAL because that is what an open
    # console with no run in it amounts to, and applies whichever mode the
    # operator picks first, even if they pick the one already showing.
    #
    # Never remembered between runs, either. An app that came back up and
    # closed a start contact because of something chosen yesterday is the
    # hazard the switch exists to remove.
    winder_auto = [True]

    def winder_is_auto():
        return bool(winder_auto[0])
    # When START was pressed, or None when it was not. Between that moment and
    # --start-delay seconds later the camera is reading but the relay is still
    # off: the coil is standing still in front of the lens, which is the one
    # chance to validate the position it is actually in before it moves.
    starting_at = [None]
    start_reason = [""]

    def validating():
        """Codes are only read when the machine is running, or about to be."""
        return machine_running or starting_at[0] is not None

    def start_machine(reason="operator"):
        """Begin the start sequence. This does NOT energise the relay.

        The relay goes on in _finish_start, once the labels standing in front
        of the camera have had --start-delay seconds to read. Starting the web
        first would drag them out of frame before they could be checked.
        """
        if machine_running or starting_at[0] is not None:
            return
        if not loaded():
            # Nothing to check the roll against, so there is nothing this
            # machine could do but wind a coil through unvalidated — which is
            # the one thing it exists to prevent.
            print("[ui] no sheet loaded — LOAD SHEET (or OPEN RECENT SHEET) "
                  "before starting")
            _note[0] = "Load the sheet for this roll before starting"
            voice.say("Load a sheet first.", key="no-sheet")
            return
        if not winder_is_auto():
            # The winder is on hand control. Its motor is the operator's, and
            # nothing this app does may take it back off them.
            print("[ui] the winder is in MANUAL — it is running on its "
                  "own controls. Switch to AUTO to hand it to the "
                  "console.")
            _note[0] = "Winder is in MANUAL — switch to AUTO to run "\
                       "from here"
            voice.say("Put the winder in auto.", key="manual")
            return
        starting_at[0] = time.time()
        start_reason[0] = reason
        # The read-in is its own chance to read something, so the watchdog
        # starts counting again from here rather than from whenever the last
        # payload happened to come in.
        last_read[0] = time.time()

        # Pressing START while the fault is still live is the operator saying
        # they have looked at it and want the line to run regardless. A short
        # row settles itself in _finish_start once the read-in has had its
        # chance at it. An unexpected code has to be forgiven by name, or it
        # would stop the line again on the very next frame it decodes on.
        if fault["kind"] == "incomplete":
            # No payload to forgive by name, so it is forgiven by the clock:
            # long enough for the web to carry that label out of frame.
            waive_until[0] = time.time() + args.start_delay + args.rewind_clear
            print(f"[label] operator accepted the incomplete label — it will "
                  f"not stop the line again while it winds out of frame")
            voice.say("Incomplete label accepted.", key="accepted-label")
        if fault["kind"] in ("unexpected", "mismatch") and fault["text"]:
            forgiven.add(normalize(fault["text"]))
            print(f"[rewind] operator accepted the unexpected code "
                  f"{_tail(fault['text'], 30)} — it will not stop the line "
                  f"again this run")
            voice.say("Wrong label accepted.", key="accepted")
        _note[0] = f"Reading the labels ({args.start_delay:.0f}s)"
        held = (f", re-checking row {recheck['row']}"
                if recheck["row"] is not None else "")
        print(f"\n[relay] START pressed ({reason}) — reading the labels in "
              f"frame for {args.start_delay:.1f}s before the relay goes on"
              f"{held}")
        voice.say("Reading the labels. Stand clear.", key="reading")

    def _finish_start():
        """The read-in has run its course: settle up and let the web go."""
        nonlocal machine_running
        starting_at[0] = None

        # Now, not on the button press: a row held for re-inspection has just
        # had its second look, so it either filled in during the read-in — in
        # which case the retire path has already cleared it — or it is a
        # confirmed defect.
        if recheck["row"] is not None:
            _adjudicate_recheck()

        # The fault survived the button press so that the overlay kept
        # showing what was missing all through the read-in — the last few
        # seconds in which it could still fill in. Now it is settled either
        # way, so it goes.
        if fault["kind"] is not None:
            fault.update(_NO_FAULT)

        relay.on(args.start_relay)
        machine_running = True
        _note[0] = None
        print(f"[relay] winding machine STARTED ({start_reason[0]}) "
              f"— relay {args.start_relay} ON")
        voice.say("Machine running.", key="running")

    def stop_machine(reason="operator", open_relay=True):
        """Stop the line. `open_relay` is only ever False for the handover to
        MANUAL, where the contact is not opening at all -- it is staying
        closed for the operator, and taking it off and putting it straight
        back would be a pulse on a start input."""
        nonlocal machine_running
        aborted = starting_at[0] is not None
        starting_at[0] = None
        if not machine_running and not aborted:
            return
        if open_relay:
            relay.off(args.start_relay)
        machine_running = False
        _note[0] = reason
        what = "start ABORTED" if aborted and not machine_running else "STOPPED"
        print(f"\n[relay] winding machine {what} ({reason}) — relay "
              f"{args.start_relay} "
              + ("OFF" if open_relay else "stays ON for hand control"))
        # A fault announces itself, in its own words, at the point it is
        # raised; this is only for the ordinary stops.
        if fault["kind"] is None:
            voice.say("Machine stopped.", key="stopped")

    def set_direction(reverse):
        """CHECK FORWARD / CHECK REVERSE, from the console.

        Which end of the sheet the pass is working toward. Changing it
        restarts the pass rather than turning the window round where it
        stands: the rows already ticked off were ticked in the other
        direction, and the window has to find the coil again from whatever
        code comes past next. What the run has recorded is untouched -- the
        journal is the record of which rows were checked, and they were.
        """
        nonlocal window
        reverse = bool(reverse)
        if reverse == checking_reverse():
            return
        if not _configurable():
            print("[ui] stop the machine before changing the check direction")
            return
        check_step[0] = RollingWindow.REVERSE if reverse else \
            RollingWindow.FORWARD
        which = ("REVERSE — last row of the sheet first, for a reel wound "
                 "onto a second spool" if reverse else
                 "FORWARD — first row of the sheet first, as printed")
        print(f"\n[window] checking {which}")
        if window is not None:
            fresh = RollingWindow(sheet, size=window.size, check=checked,
                                  grace=window.grace, step=check_step[0])
            # The pass starts again; the run's tally does not.
            fresh.done, fresh.reads = window.done, window.reads
            fresh.repeats = window.repeats
            window = fresh
            print(f"[window] the window will re-anchor on the next code it "
                  f"recognises")
        _note[0] = f"Checking {'in reverse' if reverse else 'forward'}"

    def set_winder(auto):
        """AUTO or MANUAL, from the console's toggle. The relay follows it.

        MANUAL hands the winder back: the contact closes and stays closed,
        so the machine runs on its own controls with this app standing out
        of the way. It is not a stop -- the coil may well keep turning --
        but it is the end of anything the console was doing, so a run
        part-way through its read-in is abandoned and START goes dead.

        AUTO takes it back, and takes it back open: in AUTO the contact is
        the app's to close, and it closes it when the labels standing in
        front of the camera have been read and not before.

        The contact is never taken off and put straight back on the way
        into MANUAL. It is a maintained start input, and a gap in it is a
        pulse -- which is the one edge a motor starter is built to act on.
        """
        auto = bool(auto)
        if auto == winder_auto[0]:
            return
        winder_auto[0] = auto
        if auto:
            print(f"\n[winder] AUTO — relay {args.start_relay} OFF; the line "
                  f"is the console's to start now")
            relay.off(args.start_relay)
            _note[0] = None
            voice.say("Winder in auto.", key="winder-auto")
            return
        print(f"\n[winder] MANUAL — relay {args.start_relay} ON; the winder "
              f"runs on its own controls")
        stop_machine("winder switched to manual", open_relay=False)
        relay.on(args.start_relay)
        _note[0] = "Winder in MANUAL — the machine is on hand control"
        voice.say("Winder in manual.", key="winder-manual")

    def scanning():
        """Codes are read when running, when starting — and while a fault is
        live. That last one is the rewind: the relay is off and the operator
        is winding the coil backwards by hand, but the camera keeps decoding,
        because the whole point is to see the offending labels come past
        again and prove the fault has gone.
        """
        return validating() or fault["kind"] is not None

    def _raise_fault(kind, **kw):
        """Record why the line stopped, which puts it into rewind mode."""
        fault.update(_NO_FAULT)
        fault.update(kind=kind, since=time.time(), **kw)

    def _clear_fault(why, restart=True):
        """The reason for the stop is gone. Let the machine go by itself.

        Auto-restart runs the normal start sequence, read-in and all, so the
        labels standing in front of the lens are validated in place before
        the relay comes back on — exactly as if the operator had pressed the
        button.
        """
        if fault["kind"] is None:
            return
        was = fault["kind"]
        fault.update(_NO_FAULT)
        print(f"\n[rewind] fault cleared — {why}")
        if not restart:
            return
        if args.no_auto_restart:
            print("[rewind] --no-auto-restart: press START to carry on")
            voice.say("Fault cleared. Press start.", key="cleared-manual")
            return
        voice.say("Row complete. Starting again."
                  if was == "short" else
                  "Wrong label is clear. Starting again.",
                  key="cleared", urgent=True)
        start_machine("re-check passed")

    globals()["_press_start"] = lambda: start_machine("start button")  # SEAM

    # All named after the sheet, and all reopened when the operator loads a
    # different one, so they are set up in _start_record further down — after
    # the loaders that fill them in have been defined.
    run_name = None
    saver = None
    results = None

    relay = RelayController(port=args.relay_port, enabled=not args.no_relay,
                            verbose=args.relay_verbose)
    # The whole vocabulary, rendered once in the background at start-up so
    # nothing has to be synthesised at the moment it is needed. Only the two
    # lines carrying a row number are left out -- there is a row number for
    # every row in the sheet -- and those are the `text` half of an alert,
    # spoken after a `lead` that is always ready.
    SPOKEN = ["Reading the labels. Stand clear.",
              "Machine running.",
              "Machine stopped.",
              "Stopped. Rotate the coil back.",
              "Stopped. Wrong label on the coil.",
              "Rotate the coil back and take it off.",
              "Row complete. Starting again.",
              "Wrong label is clear. Starting again.",
              "The same wrong label again. Press start to accept it, or take "
              "it off the coil.",
              "Fault cleared. Press start.",
              "Wrong label accepted.",
              "Load a sheet first.",
              "Put the winder in auto.",
              "Winder in auto.",
              "Winder in manual.",
              "Label defect."] + \
             [f"Up {i} found." for i in range(1, (per_row or MAX_UPS) + 1)]

    voice = Voice(enabled=not args.no_voice, engine=args.voice_engine,
                  name=args.voice_name, rate=args.voice_rate,
                  tone=not args.no_tone, warm=SPOKEN)

    # A previous run that was killed rather than closed can leave the motor
    # energised, so start from a state this file knows rather than inheriting
    # a coil that is on for no reason.
    relay.off(args.start_relay)
    panel = None

    cam_index = args.index if args.index is not None else find_camera_index()
    print(f"[camera] using /dev/video{cam_index}")

    # Preferred path: videoflip rotates inside the pipeline, off the loop. If
    # that pipeline won't open (no videoflip element, say), fall back to
    # rotating each frame in the loop as before.
    pipeline = gstreamer_pipeline(cam_index, args.width, args.height, args.fps,
                                  args.format, rotate=args.rotate)
    print(f"[camera] pipeline: {pipeline}")
    cap = cv2.VideoCapture(pipeline, cv2.CAP_GSTREAMER)
    rotate_in_loop = 0
    if not cap.isOpened() and args.rotate:
        print("[camera] videoflip pipeline would not open — rotating in the loop")
        pipeline = gstreamer_pipeline(cam_index, args.width, args.height,
                                      args.fps, args.format)
        cap = cv2.VideoCapture(pipeline, cv2.CAP_GSTREAMER)
        rotate_in_loop = args.rotate
    if not cap.isOpened():
        raise RuntimeError("Failed to open camera via GStreamer pipeline")

    # Exposure, gain and brightness, on the same device the pipeline is
    # streaming from. Whatever the operator set last time goes back on now:
    # these are a property of this camera in this light, not of a session,
    # and having to set them again every morning -- in a separate tool --
    # is how a reel gets run under a setting nobody chose.
    camera = CameraControls(CAM["device"] or f"/dev/video{cam_index}",
                            limits=CAM.get("limits"))
    remembered = prefs.camera
    if remembered:
        applied = camera.apply(remembered)
        if applied:
            print(f"[camera] restored "
                  + ", ".join(f"{k} {v}" for k, v in sorted(applied.items())))
    if camera.available:
        print(f"[camera] press 's' for the exposure, gain and brightness "
              f"sliders — now at "
              + ", ".join(f"{k} {v}"
                          for k, v in sorted(camera.snapshot().items())))

    # Held rather than read back every frame: the console needs these to draw
    # the sliders, and asking the device sixty times a second for three
    # numbers that only change when somebody moves a slider is work for
    # nothing.
    cam_values = [camera.snapshot()]

    def _set_camera(arg):
        """One slider moved. Set it, and remember where it was left."""
        try:
            name, value = arg
        except (TypeError, ValueError):
            return
        if not camera.set(name, value):
            print(f"[camera] {name} would not take {value}")
            return
        cam_values[0] = camera.snapshot()
        # Written on every move rather than on close: the operator adjusts
        # these and then walks back to the machine, and a console that is
        # switched off at the wall must not lose them.
        prefs.remember_camera(cam_values[0])

    classes_path = args.classes
    if classes_path is None:
        beside = CFG.asset(MODEL["classes"])
        if beside and os.path.exists(beside):
            classes_path = beside
            print(f"[model] using {beside}")
    class_names = load_class_names(classes_path)
    if args.engine is None:
        args.engine = CFG.asset(MODEL["engine"])
    model = YOLO26TRT(args.engine, input_size=(args.imgsz, args.imgsz))
    print(f"[model] loaded {args.engine}")

    # 90/270 rotation swaps the effective width/height for sizing the window/writer.
    disp_w, disp_h = (args.height, args.width) if args.rotate in (90, 270) else (args.width, args.height)

    # Per-class thresholds: the label one gates line crossings, the qr one
    # gates which box gets cropped and decoded, and --conf-thres is what every
    # other class is held to.
    conf_per_class = {}
    label_cls = class_index(class_names, args.label_class, 0)
    qr_cls = class_index(class_names, args.qr_class, 1)
    logo_cls = class_index(class_names, args.logo_class, 2)
    named = {label_cls: args.label_class, qr_cls: args.qr_class,
             logo_cls: args.logo_class}
    if args.conf_label is not None:
        conf_per_class[label_cls] = args.conf_label
    if args.conf_qr is not None:
        conf_per_class[qr_cls] = args.conf_qr
    print(f"[model] conf thresholds: default={args.conf_thres}"
          + "".join(f"  {named.get(c, c)}={v}"
                    for c, v in conf_per_class.items()))
    print(f"[model] classes: {args.label_class} (the label), "
          f"{args.qr_class} (decoded), {args.logo_class} "
          f"({'checked for presence' if not args.no_part_check else 'ignored'})")

    xlsx_path = journal_path = None      # set by _start_record
    journal = [None]         # append-only record; the durable source of truth
    xlsx_lock = threading.Lock()
    xlsx_busy = [False]
    # Where the previous run got to. A repeating sheet holds each payload many
    # times over, so without this the first code the camera sees anchors on
    # its copy in the very first block and the whole run starts again.
    resume_hint = [None]
    handled = [[]]           # boxes whose code was accepted on the last frame
    last_frame = [None]      # most recent frame, for the diagnostic dump
    recent = []              # [(payload, verdict)] most recent decodes
    ever_read = set()        # every payload decoded this run, normalised
    verified = {}          # excel row number -> (per-column marks, status)

    # ── re-inspection state (row mode) ───────────────────────────────────
    # Nothing is tolerated: a group that will not validate stops the line and
    # is held open, so the operator can wind the coil back and show the same
    # labels to the camera again. Reading clean the second time clears it;
    # failing again makes it a confirmed defect.
    recheck = {"row": None, "attempt": 0}   # sheet row awaiting re-inspection
    defects = set()                         # rows confirmed bad on re-check

    # The live fault, if any. Setting this is what puts the app into rewind
    # mode: the relay is off, but the camera keeps decoding, so the operator
    # can wind the coil backwards and watch the very labels that caused the
    # stop come past again. What the fault records is what the overlay draws
    # and what has to go away before the machine is allowed to start itself.
    #
    #   kind   'short'      the head row never read every code it needed
    #          'unexpected' a code came off the web that belongs nowhere here
    #   row    sheet row number held open (short)
    #   text   the offending payload (unexpected)
    #   seen   when that payload was last decoded, so its absence can be timed
    #   strikes payloads that have already been forgiven once, so the same
    #          bad label cannot bounce the line between stop and start on its
    #          own — the second time it stops, a human has to press START.
    _NO_FAULT = {"kind": None, "row": None, "row_idx": None, "text": None,
                 "belongs": None, "seen": 0.0, "since": 0.0, "warned": False,
                 "in_frame": False, "box": None}
    fault = dict(_NO_FAULT)
    forgiven = set()      # unexpected payloads the operator has waved through
    bounced = set()       # unexpected payloads already given one free auto-restart

    def _open_journal():
        """A one-line-per-row append log, flushed as it goes.

        The .xlsx is the deliverable, but writing it is far too slow to do
        during a run — openpyxl is pure Python, so it holds the GIL and stalls
        the capture loop. This carries the same information at a cost of
        microseconds, and the workbook is rebuilt from it at the end.
        """
        os.makedirs(os.path.dirname(journal_path) or ".", exist_ok=True)
        fresh = not os.path.exists(journal_path)
        journal[0] = open(journal_path, "a", buffering=1)
        if fresh:
            per_row = len(sheet.rows[0].texts)
            journal[0].write("sheet_row,status," +
                             ",".join(f"d{i+1}" for i in range(per_row)) + "\n")

    def _load_journal():
        """Rows verified by earlier runs. Last entry for a row wins."""
        if args.no_resume or not os.path.exists(journal_path):
            return
        import csv as _csv
        with open(journal_path, newline="") as fh:
            for rec in _csv.reader(fh):
                if len(rec) < 3 or rec[0] == "sheet_row":
                    continue
                try:
                    number = int(rec[0])
                except ValueError:
                    continue
                verified[number] = (rec[2:], rec[1])

    def _load_previous():
        """Pick up what an earlier run already verified, so the record adds up
        across sessions instead of starting blank each time."""
        if args.no_resume or not os.path.exists(xlsx_path):
            return
        import openpyxl
        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            rows = list(wb.worksheets[0].iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            # A file left over from a kill during an older, non-atomic save.
            # Starting fresh beats refusing to run.
            spoiled = xlsx_path + ".damaged"
            try:
                os.replace(xlsx_path, spoiled)
                where = f" — moved to {spoiled}"
            except OSError:
                where = ""
            print(f"[window] could not read {xlsx_path} ({exc}){where}; "
                  f"starting the record fresh")
            return
        if not rows:
            return
        header = [str(c) if c is not None else "" for c in rows[0]]
        try:
            first = header.index("READ D1")
        except ValueError:
            return
        per_row = len(sheet.rows[0].texts)
        for n, values in enumerate(rows[1:], start=2):
            status = values[first + per_row] if first + per_row < len(values) else None
            if not status:
                continue
            marks = [values[first + i] if first + i < len(values) else None
                     for i in range(per_row)]
            verified[n] = (marks, status)
        pass

    def record_row(row_idx, complete, status=None):
        """Fold one retired row into the record kept for the sheet."""
        per_row = len(sheet.rows[0].texts)
        required = window.required(row_idx)
        seen = window.seen.get(row_idx, set())
        marks = []
        for col in range(per_row):
            if col not in required:
                marks.append("not checked")
            elif col in seen:
                marks.append("OK")
            else:
                marks.append("NOT DECODED")
        status = status or ("OK" if complete and not (required - seen)
                            else "INCOMPLETE")
        number = sheet.rows[row_idx].number
        # A row verified on an earlier pass is not un-verified by a later one
        # that happened to read it badly — but a confirmed defect always wins,
        # because those labels were shown to the camera twice on purpose.
        if (verified.get(number, (None, None))[1] == "OK"
                and status not in ("OK", "DEFECT")):
            return
        verified[number] = (marks, status)
        if journal[0] is not None:
            journal[0].write(f"{number},{status}," + ",".join(marks) + "\n")

    def write_annotated_xlsx(background=False):
        """Copy the source sheet and add, per row, which codes were decoded.

        The source file is the reference for what was printed and is never
        touched — this is a second file recording what the camera could
        actually read of it. Written to a fixed path so --resume can find it.
        """
        if not verified:
            return
        if background:
            with xlsx_lock:
                if xlsx_busy[0]:
                    return          # one already in flight; the next will cover it
                xlsx_busy[0] = True
            snapshot = dict(verified)
            threading.Thread(target=_do_write, args=(snapshot, True),
                             daemon=True).start()
            return
        _do_write(dict(verified), False)

    def _do_write(snapshot, background):
        try:
            import openpyxl
            os.makedirs(os.path.dirname(xlsx_path) or ".", exist_ok=True)
            # The working copy, not the operator's file: the marks are per
            # row and the copy is the one whose rows the window counted, with
            # the datamatrix printings expanded into rows of their own.
            wb = openpyxl.load_workbook(work_xlsx[0] or args.xlsx)
            ws = wb[args.sheet] if args.sheet else wb.worksheets[0]
            first = ws.max_column + 1
            per_row = len(sheet.rows[0].texts)
            for i in range(per_row):
                ws.cell(row=1, column=first + i, value=f"READ D{i + 1}")
            ws.cell(row=1, column=first + per_row, value="ROW STATUS")
            ws.cell(row=1, column=first + per_row + 1, value="CHECKED AT")

            stamp = time.strftime("%Y-%m-%d %H:%M:%S")
            for number, (marks, status) in snapshot.items():
                for col, mark in enumerate(marks):
                    ws.cell(row=number, column=first + col, value=mark)
                ws.cell(row=number, column=first + per_row, value=status)
                ws.cell(row=number, column=first + per_row + 1, value=stamp)
            # Written to one side and moved into place, because os.replace is
            # atomic: a kill during the save leaves the previous good file
            # intact rather than a truncated one that cannot be reopened.
            tmp = xlsx_path + ".tmp"
            wb.save(tmp)
            wb.close()
            os.replace(tmp, xlsx_path)
            if not background:
                ok = sum(1 for _, st in snapshot.values() if st == "OK")
                print(f"[window] wrote {xlsx_path} — {ok} rows OK, "
                      f"{len(snapshot) - ok} incomplete")
        except Exception as exc:
            print(f"[window] could not write the annotated sheet: {exc}")
        finally:
            if background:
                with xlsx_lock:
                    xlsx_busy[0] = False

    # The panel costs ~2.9ms of the ~21ms frame budget to draw — it is ~50
    # putText calls — but its content only moves when the window does. Cache
    # it against the window state it reads, so a frame that changed nothing
    # reuses the last image instead of redrawing it.
    view_cache = [None, None]        # [state key, rendered image]

    def _window_view_key():
        if not loaded():
            return None
        return (window.start, window.reads, len(window.done), window.last_hit,
                len(window.unexpected))

    def render_window_view():
        """The rolling window as its own panel: what the sheet expects against
        what has actually been read, plus the state of every open row.

        Redrawn only when the window state behind it has changed."""
        key = _window_view_key()
        if view_cache[0] == key and view_cache[1] is not None:
            return view_cache[1]
        img = _no_sheet_view() if not loaded() else _draw_window_view()
        view_cache[0], view_cache[1] = key, img
        return img

    def _no_sheet_view():
        """The diagnostics panel before a sheet has been chosen."""
        import numpy as np
        img = np.full((120, 1120, 3), (32, 32, 32), np.uint8)
        cv2.putText(img, "NO SHEET LOADED", (24, 50),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.9, (60, 200, 235), 2,
                    cv2.LINE_AA)
        cv2.putText(img, "LOAD SHEET, or OPEN RECENT SHEET for one that has "
                         "been run before", (24, 90),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.55, (200, 200, 200), 1,
                    cv2.LINE_AA)
        return img

    def _draw_window_view():
        import numpy as np
        font = cv2.FONT_HERSHEY_SIMPLEX
        W = 1120
        open_rows = window.rows()[:8] if window.start is not None else []
        H = 150 + 4 * 30 + 40 + max(len(open_rows), 1) * 28 + 90
        img = np.full((H, W, 3), (32, 32, 32), np.uint8)

        BG_OK, BG_BAD, DIM = (80, 220, 80), (70, 70, 235), (130, 130, 130)
        HEAD, WHITE, AMBER = (200, 200, 200), (240, 240, 240), (60, 200, 235)

        def put(text, x, y, colour=WHITE, scale=0.55, thick=1):
            cv2.putText(img, text, (x, y), font, scale, colour, thick, cv2.LINE_AA)

        def tail(text, n=30):
            if not text:
                return ""
            text = str(text)
            return text if len(text) <= n else ".." + text[-(n - 2):]

        if window.start is None:
            put("ROLLING WINDOW", 24, 42, WHITE, 0.9, 2)
            put("waiting for a code the sheet recognises", 24, 80, DIM, 0.65)
            return img

        passed = sum(1 for _, ok in window.done if ok)
        short = len(window.done) - passed
        put("ROLLING WINDOW", 24, 42, WHITE, 0.9, 2)
        put(f"{passed} pass   {short} short   {window.reads} codes read   "
            f"window {window.size} rows",
            24, 72, BG_BAD if short else BG_OK, 0.6)

        # ── the row most recently touched, expected against decoded ───────
        y = 110
        if window.last_hit is not None:
            row_idx, _ = window.last_hit
            row_no = sheet.rows[row_idx].number
            required = window.required(row_idx)
            seen = window.seen.get(row_idx, set())
            put(f"LAST MATCHED  ->  sheet row {row_no}", 24, y, HEAD, 0.65)
            y += 26
            for label, x in (("POS", 24), ("EXPECTED (xlsx)", 96),
                             ("DECODED (camera)", 450), ("RESULT", 810)):
                put(label, x, y, HEAD, 0.5)
            cv2.line(img, (16, y + 10), (W - 16, y + 10), (70, 70, 70), 1)
            for col, expected in enumerate(sheet.rows[row_idx].texts):
                y2 = y + 10 + (col + 1) * 30
                if col not in required:
                    got, res, colour = "(not checked)", "-", DIM
                elif col in seen:
                    got = tail(window.texts.get((row_idx, col), expected))
                    res, colour = "OK", BG_OK
                else:
                    got, res, colour = "(waiting)", "..", AMBER
                put(up(col), 24, y2, HEAD, 0.55)
                put(tail(expected) or "-", 96, y2, DIM, 0.5)
                put(got, 450, y2, colour, 0.5)
                put(res, 810, y2, colour, 0.5)
            y = y + 10 + 5 * 30
        else:
            y += 20

        # ── every open row, and what each still needs ─────────────────────
        y += 14
        put("OPEN ROWS", 24, y, HEAD, 0.6)
        y += 8
        for i, row_idx in enumerate(open_rows):
            y2 = y + (i + 1) * 28
            row_no = sheet.rows[row_idx].number
            required = window.required(row_idx)
            seen = window.seen.get(row_idx, set())
            complete = required and required <= seen
            colour = BG_OK if complete else (AMBER if i == 0 else WHITE)
            put(("> " if i == 0 else "  ") + f"row {row_no:<6}", 24, y2, colour, 0.55)
            x = 190
            for col in range(len(sheet.rows[row_idx].texts)):
                if col not in required:
                    mark, c = "-", DIM
                elif col in seen:
                    mark, c = "OK", BG_OK
                else:
                    mark, c = "..", AMBER
                put(f"{up(col)}:{mark}", x, y2, c, 0.55)
                x += 92
            still = sorted(required - seen)
            if still:
                put("needs " + ", ".join(up(c) for c in still),
                    x + 20, y2, AMBER, 0.5)
        y += (len(open_rows) + 1) * 28

        # ── anything that matched nothing at all ──────────────────────────
        if window.unexpected:
            y += 24
            put("UNEXPECTED", 24, y, BG_BAD, 0.6)
            for i, (text, belongs) in enumerate(window.unexpected[-2:]):
                put(f"{tail(text, 44)}   -> {belongs}", 190, y + i * 24,
                    BG_BAD, 0.5)
        return img

    # ── the record that goes with the sheet ──────────────────────────────
    # Everything a run writes is named after the .xlsx it was checking, so the
    # sheet is what defines a record. It is written to two roots: the
    # paperwork — the checked .xlsx, progress.csv and the run log — into the
    # project under --result-dir, and the label crops on their own into
    # --label-dir/<xlsx name>/, which is the folder the operator sets and
    # which holds nothing but images. Changing either the sheet or that folder
    # closes one record and opens another — _bind_run is the only place that
    # happens.

    def _record_shape():
        """What a record written now would be numbered against.

        Everything a run records — the journal, the annotated sheet — is
        keyed by spreadsheet row number, and expanding a datamatrix into
        rows of its own moves every row after it down. A record written
        before the expansion therefore has its marks against rows that are
        now something else, so it is not a record of this sheet at all.
        """
        return {"rows": len(sheet.rows), "dm_rows": sheet.dm_rows,
                "per_row": per_row}

    def _stamp_path():
        return os.path.join(os.path.dirname(journal_path) or ".",
                            "record.json")

    def _aside(path, stamp):
        """Move a file out of the way, keeping it. Returns its new name."""
        if not os.path.exists(path):
            return None
        stem, ext = os.path.splitext(path)
        aside = f"{stem}.before-{stamp}{ext}"
        try:
            os.replace(path, aside)
            return os.path.basename(aside)
        except OSError as exc:
            print(f"[window] could not put {path} aside ({exc})")
            return None

    def _retire_stale_record():
        """Bring forward a record whose row numbers no longer mean anything.

        Only ever triggered by a sheet that has a datamatrix in it: without
        one nothing is expanded, the numbering is what it always was, and
        the record from any earlier run carries straight on untouched.

        A run before the expansion counted rows the way the operator's own
        sheet does, and the working copy still says, per row, which of those
        it came from — so the old journal is not lost, it is renumbered.
        A shift's work stands, against the rows it was really about.
        """
        import csv as _csv
        import json
        if not os.path.exists(journal_path):
            return
        want = _record_shape()
        try:
            with open(_stamp_path()) as fh:
                have = json.load(fh)
        except (OSError, ValueError):
            # No stamp at all: written before this app expanded anything. Its
            # numbering is only wrong if this sheet is one that gets expanded.
            have = None if sheet.has_dm else want
        if have == want:
            return

        stamp = time.strftime("%Y%m%d-%H%M%S")
        # The annotated .xlsx is rebuilt from the journal at the end of every
        # run, so it only has to be got out of the way; the journal is the
        # part worth carrying over.
        was_xlsx = _aside(xlsx_path, stamp)
        was_journal = _aside(journal_path, stamp)
        kept = [n for n in (was_xlsx, was_journal) if n]
        old = (os.path.join(os.path.dirname(journal_path) or ".", was_journal)
               if was_journal else None)
        print(f"\n[window] the record here was written against a different "
              f"shape of this sheet")
        print(f"[window]   then: "
              f"{have or 'the sheet as the operator numbers it'}")
        print(f"[window]   now:  {want}")

        # Only a record written against the operator's own numbering can be
        # brought forward by source row: one written against some other
        # expansion (a different --dm-repeats) is numbered by rows this copy
        # does not have, and guessing at it would be worse than starting over.
        by_source = {r.source: r.number for r in sheet.rows
                     if not r.is_dm and r.source is not None}
        if have is not None or not by_source or old is None:
            print(f"[window]   kept as {', '.join(kept)}; this run starts a "
                  f"fresh record")
            return

        moved = dropped = 0
        try:
            os.makedirs(os.path.dirname(journal_path) or ".", exist_ok=True)
            with open(old, newline="") as src, \
                    open(journal_path, "w", newline="") as dst:
                out = _csv.writer(dst)
                for rec in _csv.reader(src):
                    if len(rec) < 2:
                        continue
                    if rec[0] == "sheet_row":
                        out.writerow(rec)          # the header, as it was
                        continue
                    try:
                        number = by_source[int(rec[0])]
                    except (ValueError, KeyError):
                        dropped += 1
                        continue
                    out.writerow([number] + rec[1:])
                    moved += 1
        except OSError as exc:
            print(f"[window]   could not carry it over ({exc}) — kept as "
                  f"{', '.join(kept)} and starting a fresh record")
            return
        print(f"[window]   {moved} row(s) carried over onto the new numbering"
              + (f", {dropped} skipped (no such row in this sheet)"
                 if dropped else "")
              + f"; the originals are kept as {', '.join(kept)}")

    def _stamp_record():
        import json
        try:
            with open(_stamp_path(), "w") as fh:
                json.dump(_record_shape(), fh, indent=2)
        except OSError as exc:
            print(f"[window] could not write {_stamp_path()} ({exc})")

    def _start_record():
        """Open the books for the sheet and output folder now in `args`."""
        nonlocal run_name, saver, results, xlsx_path, journal_path
        run_name = os.path.splitext(os.path.basename(args.xlsx))[0]
        xlsx_path = args.out_xlsx or os.path.join(
            args.result_dir, run_name, f"checked_{run_name}.xlsx")
        journal_path = os.path.join(args.result_dir, run_name, "progress.csv")

        saver = None
        if not args.no_save_labels:
            saver = LabelSaver(root=args.label_dir, name=run_name,
                               subdir=None, ext=args.label_format,
                               pad=args.label_pad, min_pad=args.label_pad_px)
        results = None
        if not args.no_result_log:
            results = ResultLog(root=args.result_dir, name=run_name,
                                columns=per_row)

        # What an earlier run already got through.
        _retire_stale_record()
        _load_previous()  # an .xlsx from an older run, if there is one
        _load_journal()   # then the journal, which wins where they differ
        _open_journal()
        _stamp_record()
        if verified:
            # Where the last run got to is the far end of what it
            # checked, and which end that is depends on which way it
            # was walking the sheet.
            last = min(verified) if checking_reverse() else max(verified)
            by_number = {r.number: i for i, r in enumerate(sheet.rows)}
            resume_hint[0] = by_number.get(last)
            done = sum(1 for _, st in verified.values() if st == "OK")
            print(f"[window] resuming from row {last} — {done} rows "
                  f"already verified ({len(verified)} recorded)")

    def _close_record():
        """Shut the current record's books, so nothing is left half written
        when the run is repointed at another sheet or folder."""
        nonlocal saver, results
        if journal[0] is not None:
            journal[0].close()
            journal[0] = None
        if verified:
            for _ in range(60):      # let any background save finish first
                with xlsx_lock:
                    if not xlsx_busy[0]:
                        break
                time.sleep(0.1)
            write_annotated_xlsx()
        if results is not None:
            print(f"[results] {results.rows} rows written to {results.path}")
            results.close()
            results = None
        if saver is not None:
            print(f"[crops] saved {saver.count} label crops to {saver.dir}/")
            saver = None

    def _bind_run(new_xlsx=None, new_label_dir=None):
        """Point the app at a different sheet and/or label folder.

        Only ever called with the machine idle. Everything a run accumulates
        belongs to the sheet it was checking, so it is all closed off and
        started again rather than carried across: a window anchored on the old
        sheet's rows means nothing against a new one, and codes read from the
        old coil must not tick off cells in the new record.
        """
        nonlocal defects
        _close_record()
        if new_xlsx:
            args.xlsx = new_xlsx
            args.out_xlsx = None      # it was named after the old sheet
        if new_label_dir:
            args.label_dir = new_label_dir
            prefs.remember_label_dir(new_label_dir)

        verified.clear()
        resume_hint[0] = None
        recent.clear()
        ever_read.clear()
        handled[0] = []
        marks[0] = []
        defects = set()
        forgiven.clear()
        bounced.clear()
        credited.clear()
        pending_crops.clear()
        rewound[0] = 0
        recheck["row"] = None
        recheck["attempt"] = 0
        fault.update(_NO_FAULT)
        zb_reads[0] = 0

        if not args.xlsx:
            # The folder was repointed with no sheet loaded. There is nothing
            # to open books for yet; the sheet will do it when it is chosen.
            print(f"[run] crops will go to {args.label_dir}/ — still waiting "
                  f"for a sheet")
            return
        _open_sheet()
        _start_record()
        print(f"[run] now checking {args.xlsx} — record in "
              f"{args.result_dir}/, crops in {args.label_dir}/")

    if loaded():
        _start_record()

    def _band_top():
        """First row of picture the console chrome does not own."""
        return panel.header_h if panel is not None else 0

    def _band_bottom():
        """Last row of picture the console chrome does not own."""
        return panel.footer_h if panel is not None else 0

    def _band_right(width):
        """Last column of picture the console chrome does not own."""
        return panel.content_right if panel is not None else width

    def _tk():
        """How much bigger than nominal every caption has to be drawn.

        The frame is decoded at 1944x2592 and shown at about a third of that,
        so a caption drawn at scale 1 arrives on screen a third the size it
        looks like here. Everything written on the picture multiplies by this
        so it is sized for the screen the operator is actually reading.
        """
        return panel.text if panel is not None else 1.0

    def draw_window(frame, compact=False):
        """One status line on the video, since the detail now lives in its own
        window. `compact` skips the per-row breakdown."""
        font = cv2.FONT_HERSHEY_SIMPLEX
        x, y = 20, _band_top() + int(56 * _tk())
        if not loaded():
            cv2.putText(frame, "NO SHEET LOADED", (x, y),
                        font, 0.8, (0, 255, 255), 2, cv2.LINE_AA)
            return frame
        if window.start is None:
            cv2.putText(frame, "WINDOW: waiting for a known code", (x, y),
                        font, 0.8, (0, 255, 255), 2, cv2.LINE_AA)
            return frame

        passed = sum(1 for _, ok in window.done if ok)
        failed = len(window.done) - passed
        head = sheet.rows[window.start].number if not window.exhausted else "-"
        cv2.putText(frame, f"WINDOW  {passed} pass / {failed} short"
                           f"   reads {window.reads}   head row {head}",
                    (x, y), font, 0.62 * _tk(),
                    (0, 0, 255) if failed else (0, 255, 0), 2, cv2.LINE_AA)
        if compact:
            return frame

        for i, row_idx in enumerate(window.rows()[:8]):
            row_no = sheet.rows[row_idx].number
            required = window.required(row_idx)
            seen = window.seen.get(row_idx, set())
            marks = []
            for col in range(len(sheet.rows[row_idx].texts)):
                if col not in required:
                    marks.append("-")
                elif col in seen:
                    marks.append("OK")
                else:
                    marks.append("..")
            head = ">" if i == 0 else " "
            colour = ((0, 255, 0) if required and required <= seen
                      else (255, 255, 255) if i else (0, 255, 255))
            cv2.putText(frame, f"{head} row {row_no:<5} "
                               + "  ".join(f"{up(c)}:{m}" for c, m in enumerate(marks)),
                        (x, y + 32 + i * 30), font, 0.6, colour, 2, cv2.LINE_AA)
        return frame

    dump_n = [0]
    DUMP_CAP = 500

    def _dump_miss(frame, label_box, qr_box):
        """A label was detected but would not decode. Save what the camera saw.

        Both crops go down: the label crop that was tried first, and the
        tighter qr_code crop that was tried as a fallback. Between them they
        answer the only question that matters when a row comes up short — was
        the code clipped, soft, or simply not there.
        """
        if not args.dump_crops or dump_n[0] >= DUMP_CAP:
            return
        os.makedirs(args.dump_crops, exist_ok=True)
        stamp = time.strftime("%H%M%S") + f"-{int((time.time() % 1) * 1000):03d}"
        h, w = frame.shape[:2]
        for tag, b in (("label", label_box), ("qr", qr_box)):
            if b is None:
                continue
            x1, y1, x2, y2 = (max(int(b[0]), 0), max(int(b[1]), 0),
                              min(int(b[2]), w), min(int(b[3]), h))
            if x2 - x1 < 4 or y2 - y1 < 4:
                continue
            edge = ""
            if x1 <= 2 or y1 <= 2 or x2 >= w - 2 or y2 >= h - 2:
                edge = "-CLIPPED"
            cv2.imwrite(os.path.join(args.dump_crops,
                                     f"miss-{stamp}-{tag}{edge}.jpg"),
                        frame[y1:y2, x1:x2])
        dump_n[0] += 1
        if dump_n[0] == DUMP_CAP:
            print(f"\n[dump] {DUMP_CAP} failed crops saved to "
                  f"{args.dump_crops}/ — not saving any more")

    def _dump_frame(tag):
        """The whole frame, for the moment a row is held. Context the crops
        cannot give: where the row was, and what else was in view."""
        if not args.dump_crops or last_frame[0] is None:
            return
        os.makedirs(args.dump_crops, exist_ok=True)
        path = os.path.join(args.dump_crops,
                            f"frame-{time.strftime('%H%M%S')}-{tag}.jpg")
        cv2.imwrite(path, last_frame[0])
        print(f"[dump] frame at the moment of the hold: {path}")

    def draw_decoders(frame, legend=False):
        """Repaint each label's box in the colour of the decoder that read it.

        draw_detections has already outlined every detection in one colour;
        this goes over the labels afterwards so the box itself carries the
        verdict — green read by zxing, amber rescued by zbar, red read by
        nothing. A zbar box is the interesting one: it marks a code the
        primary decoder cannot see at all.

        The boxes and the payload under each one are what the operator is
        looking at, so they are always drawn. `legend` adds the per-frame
        tally of which decoder read what, which is a diagnostic and belongs
        with the rest of the debug readout.
        """
        font = cv2.FONT_HERSHEY_SIMPLEX
        for box, kind in parts[0]:
            x1, y1, x2, y2 = (int(v) for v in box)
            cv2.rectangle(frame, (x1, y1), (x2, y2), PART_COLOUR[kind], 2)
        for box, who, text in marks[0]:
            x1, y1, x2, y2 = (int(v) for v in box)
            bad = _bad_label(box)
            colour = (DECODER_COLOUR["fail"] if bad
                      else DECODER_COLOUR.get(who, DECODER_COLOUR["fail"]))
            cv2.rectangle(frame, (x1, y1), (x2, y2), colour, 4)

            # The payload under the box, on a filled strip in the decoder's
            # colour. Only the tail is shown: every code on this reel shares
            # the same prefix and differs only at the end, and the shorter it
            # is the larger it can be set without swamping the label.
            caption = (_bad_caption() if bad else
                       _tail(text, 14) if text else "NO READ")
            cs = 0.55 * _tk()
            # A caption a little wider than its label is readable; one three
            # times as wide just covers the neighbours, so it gives up size.
            room = int((x2 - x1) * 1.6)
            (tw, th), _ = cv2.getTextSize(caption, font, cs, 2)
            if tw > room:
                cs *= room / float(tw)
                (tw, th), _ = cv2.getTextSize(caption, font, cs, 2)
            ty = y2 + th + 12
            if ty > frame.shape[0] - _band_bottom() - 4:
                ty = y1 + th + 12            # off the bottom: put it inside
            cv2.rectangle(frame, (x1, ty - th - 9), (x1 + tw + 14, ty + 8),
                          colour, -1)
            cv2.putText(frame, caption, (x1 + 7, ty), font, cs,
                        (0, 0, 0), 2, cv2.LINE_AA)

        if not legend:
            return frame

        counts = {}
        for _, who, _t in marks[0]:
            counts[who] = counts.get(who, 0) + 1
        tk = _tk()
        ls = 0.55 * tk
        box = int(26 * tk)
        x, y = 20, _band_top() + int(100 * tk)
        for who, text in (("zxing", "zxing"), ("zbar", "zbar"),
                          ("fail", "no read")):
            swatch = DECODER_COLOUR[who]
            cv2.rectangle(frame, (x, y - box), (x + box, y), swatch, -1)
            caption = f"{text} {counts.get(who, 0)}"
            cv2.putText(frame, caption, (x + box + int(10 * tk), y - 4), font,
                        ls, swatch, 2, cv2.LINE_AA)
            x += box + int(28 * tk) + cv2.getTextSize(caption, font, ls, 2)[0][0]
        if zb_reads[0]:
            cv2.putText(frame, f"zbar rescues this run: {zb_reads[0]}",
                        (20, y + int(34 * tk)), font, ls,
                        DECODER_COLOUR["zbar"], 2, cv2.LINE_AA)
        return frame

    # ── the rewind overlay ────────────────────────────────────────────────
    # A stop is only half the job. The operator now winds the coil backwards
    # by hand, and while they do the camera keeps decoding — so the screen can
    # show, live, which physical label the fault is about and whether winding
    # back has fixed it yet.
    FAULT_RED = (60, 60, 235)
    FAULT_GREEN = (80, 220, 80)
    FAULT_WANT = (255, 200, 0)       # a position of the held row not read yet
    FAULT_OVER = (60, 160, 255)      # a row the coil has been wound past
    FAULT_TEXT = (240, 240, 240)

    def _rewind_guidance():
        """How far, and which way, the coil still has to go.

        The operator cannot see the sheet, so left to themselves they wind
        back until the line starts again -- which is how the coil ends up
        wound past the row being re-checked and the camera starts reading
        labels signed off long ago. This is the marker that stops that.

        Everything it says has to hold on a sheet that repeats every few
        rows, which rules out reading a row number off a label. Two things
        do hold: whether one of the held row's own payloads is in frame, and
        whether the labels in frame are ones this run has already signed off.
        Between them they give the direction, and on a sheet that does not
        repeat there is a distance to go with it.
        """
        target = fault["row_idx"]
        if target is None:
            return None
        here = [normalize(t) for _b, _w, t in marks[0] if t]
        if not here:
            return (">> nothing is decoding - wind slowly", FAULT_TEXT)

        # Exact, whatever the sheet does: these are the row's own codes.
        wanted = {normalize(t) for t in sheet.rows[target].texts if t}
        if wanted & set(here):
            return (f">> ROW {fault['row']} IS IN FRAME - wind slowly",
                    FAULT_GREEN)

        # Where the labels in frame were checked, relative to the row being
        # re-checked. A label this run has never credited is one the web has
        # not reached yet, which puts it ahead like the rest.
        rows = [credited[t] for t in here if t in credited]
        # A label this run has never credited is one the web has not reached
        # yet, so it lies ahead -- but by how far is not known, and that is
        # what decides whether a distance can honestly be quoted.
        unknown = len(here) - len(rows)
        # How far each of them sits from the held row, counted the way the
        # pass runs: negative is already gone past, positive is still to
        # come. On a reel being checked backwards the row numbers run the
        # other way, and comparing them raw would send the operator winding
        # in exactly the wrong direction.
        step = window.step
        gaps = [(r - target) * step for r in rows]
        behind = [d for d in gaps if d < 0]
        ahead = [d for d in gaps if d > 0]

        if behind and not ahead and not unknown:
            return (f">> WOUND TOO FAR - wind forward {-max(behind)} "
                    f"row(s), until row {fault['row']} shows", FAULT_OVER)
        if not behind:
            gap = (f" - {min(ahead)} more row(s)"
                   if ahead and not unknown else "")
            return (f">> KEEP WINDING BACK{gap}, until row {fault['row']} "
                    f"shows", FAULT_WANT)
        return (f">> ROW {fault['row']} IS CLOSE - wind slowly", FAULT_WANT)

    def _fault_headline():
        """One line saying what is wrong, for the headless status line."""
        if fault["kind"] == "short":
            idx = fault["row_idx"]
            miss = sorted(window.missing(idx)) if idx is not None else []
            return (f"row {fault['row']} needs "
                    + (", ".join(up(c) for c in miss) if miss
                       else "nothing"))
        if fault["kind"] == "unexpected":
            return f"unexpected {_tail(fault['text'], 20)}"
        if fault["kind"] == "unread":
            return "nothing is reading"
        if fault["kind"] == "incomplete":
            return f"label with no {(fault['belongs'] or '').lower()}"
        if fault["kind"] == "mismatch":
            return "wrong sheet for this roll"
        return ""

    def _bad_caption():
        """What the strip under the offending label reads.

        Two different faults put a red box on a label, and they call for two
        different things: one label wound out of the roll, or the whole sheet
        changed.
        """
        if fault["kind"] == "mismatch":
            return "NOT IN SHEET"
        if fault["kind"] == "incomplete":
            return f"NO {fault['belongs'] or 'PART'}"
        return "WRONG LABEL"

    def _bad_label(box):
        """Is this the physical label that stopped the line?

        Matched by where it is rather than by what it decodes. A label with a
        wrong one stuck over the printed one carries two codes, and the
        reader returns whichever it happens to find first -- so keyed off the
        payload the culprit flickers in and out of being the culprit, and on
        the frames the printed code wins it goes green like all the rest.
        Its position is steady even when its reading is not.
        """
        bad = fault.get("box")
        return bad is not None and _overlap(box, bad) > 0.25

    def _fault_tag(box, text):
        """What this label is, in terms of the live fault.

        Returns (caption, colour) for a label worth pointing at during the
        rewind, or None for one that has nothing to do with the fault. This
        is the answer to 'where is it' — the offending label carries its own
        label on screen instead of the operator having to work out which of
        the codes on the coil the console was talking about.
        """
        if fault["kind"] is None:
            return None
        key = normalize(text or "")
        if not key and not _bad_label(box):
            return None
        if fault["kind"] == "incomplete":
            if _bad_label(box):
                return (f"NO {fault['belongs'] or 'PART'} - WIND THIS OUT",
                        FAULT_RED)
            return None
        if fault["kind"] == "mismatch":
            if key == normalize(fault["text"] or "") or _bad_label(box):
                return "NOT IN THIS SHEET", FAULT_RED
            return None
        if fault["kind"] == "unexpected":
            if key == normalize(fault["text"] or "") or _bad_label(box):
                return "WRONG LABEL - WIND THIS OUT", FAULT_RED
            return None
        idx = fault["row_idx"]
        if idx is None:
            return None
        for col, want in enumerate(sheet.rows[idx].texts):
            if normalize(want) == key:
                got = col in window.seen.get(idx, set())
                return (f"ROW {fault['row']} {up(col)} "
                        + ("READ" if got else "NEEDED"),
                        FAULT_GREEN if got else FAULT_WANT)

        # Not the held row: outlined, but not captioned. The number it used
        # to carry was picked from three hundred identical copies of the same
        # payload and meant nothing. Which side of the target it falls on
        # does mean something, and that is what the colour says -- orange for
        # a row already checked and so behind, cyan for one still to come.
        where = credited.get(key)
        return "", (FAULT_OVER if where is not None and where < idx
                    else FAULT_WANT)

    def _fault_report():
        """The headline and the detail lines for the live fault.

        What the operator is told while winding the coil back, worked out
        once here and rendered by whichever console is running -- OpenCV
        burns it into the frame, Qt paints it as text. Hershey has no glyph
        for anything outside ASCII, so every string stays plain ASCII: a dash
        or an ellipsis comes out of it as "???".
        """
        banner = ""
        lines = []                    # (text, colour, relative size)
        if fault["kind"] == "short":
            idx = fault["row_idx"]
            row = sheet.rows[idx]
            req = sorted(window.required(idx))
            seen = window.seen.get(idx, set())
            missing = [c for c in req if c not in seen]
            banner = (f"REWIND: ROW {fault['row']} NEEDS "
                      + ", ".join(up(c) for c in missing))
            if row.is_dm:
                # These are not the QR labels. Say so before the operator
                # goes looking down the coil for a code that is not there:
                # they are the datamatrix labels, and this is the Nth of the
                # few that carry the same value one after another.
                banner = (f"REWIND: DATA MATRIX {row.printing or '?'} OF "
                          f"{args.dm_repeats} NEEDS "
                          + ", ".join(up(c) for c in missing))
                lines.append((f"row {fault['row']} is a DATA MATRIX row - "
                              f"printing {row.printing or '?'} of "
                              f"{args.dm_repeats}", FAULT_WANT, 0.8))
                lines.append(("these labels carry a datamatrix, not a QR",
                              FAULT_TEXT, 0.75))
            for col in req:
                got = col in seen
                want = row.texts[col]
                lines.append((f"{up(col):<5}{_tail(want, 20):<22}"
                              f"{'READ' if got else 'NOT READ YET'}",
                              FAULT_GREEN if got else FAULT_RED, 0.8))
            guide = _rewind_guidance()
            if guide is not None:
                lines.append((guide[0], guide[1], 0.85))
            lines.append(("rotate the coil back until every position reads "
                          "- then it starts itself", FAULT_TEXT, 0.7))
            lines.append((f"or press START to record row {fault['row']} as a "
                          f"label defect", FAULT_TEXT, 0.7))
        elif fault["kind"] == "incomplete":
            banner = f"LABEL WITH NO {fault['belongs'] or 'PART'}"
            lines.append(("the model found the label but not its "
                          f"{(fault['belongs'] or '').lower()}", FAULT_RED, 0.8))
            lines.append(("it is outlined in red on the picture",
                          FAULT_RED, 0.8))
            lines.append(("wind the coil back and take that label out",
                          FAULT_WANT, 0.85))
            lines.append(("it starts itself once the label is out of frame",
                          FAULT_TEXT, 0.7))
            lines.append(("or press START to let this one past",
                          FAULT_TEXT, 0.7))
        elif fault["kind"] == "unread":
            banner = "NOTHING IS READING"
            lines.append((f"{fault['row']} label(s) in frame, no code out of "
                          f"any of them", FAULT_RED, 0.8))
            lines.append(("these labels may carry no QR at all - check one",
                          FAULT_TEXT, 0.75))
            lines.append(("then check the lens, the focus and the light",
                          FAULT_TEXT, 0.75))
            lines.append(("it starts itself as soon as one label reads",
                          FAULT_WANT, 0.85))
        elif fault["kind"] == "mismatch":
            # No rewind for this one: there is nothing on the coil to wind
            # back to. What has to change is the sheet, or the roll.
            banner = "WRONG SHEET FOR THIS ROLL"
            lines.append((f"read     {_tail(fault['text'] or '', 26)}",
                          FAULT_RED, 0.8))
            lines.append((f"sheet    {os.path.basename(args.xlsx)}",
                          FAULT_TEXT, 0.8))
            lines.append(("that code is in no row of this sheet, and",
                          FAULT_TEXT, 0.75))
            lines.append(("nothing read so far has matched it either",
                          FAULT_TEXT, 0.75))
            lines.append(("load the sheet for this roll - LOAD SHEET",
                          FAULT_WANT, 0.85))
            lines.append(("or press START to run this label past anyway",
                          FAULT_TEXT, 0.7))
        elif fault["kind"] is not None:
            gone = time.time() - fault["seen"]
            banner = f"REWIND: UNEXPECTED CODE {_tail(fault['text'] or '', 16)}"
            lines.append((f"read     {_tail(fault['text'] or '', 26)}",
                          FAULT_RED, 0.8))
            lines.append((f"belongs  {fault['belongs']}", FAULT_TEXT, 0.8))
            if fault["in_frame"]:
                lines.append(("that label is IN FRAME - outlined in red",
                              FAULT_RED, 0.8))
            elif fault["warned"]:
                lines.append(("stopped the line twice - press START to accept "
                              "it", FAULT_TEXT, 0.8))
            else:
                left = max(args.rewind_clear - gone, 0.0)
                lines.append((f"out of frame - restarting in {left:.1f}s",
                              FAULT_GREEN, 0.8))
        return banner, lines

    def draw_fault(frame):
        """Everything the operator needs while winding the coil back."""
        if fault["kind"] is None:
            return frame

        font = cv2.FONT_HERSHEY_SIMPLEX
        h, w = frame.shape[:2]
        # A red border round the picture: a screen stopped on a fault must not
        # read as an idle one at a glance from across the machine. It frames
        # the picture only, so the console chrome stays legible.
        cv2.rectangle(frame, (0, _band_top()),
                      (_band_right(w), h - _band_bottom()), FAULT_RED, 14)

        banner, lines = _fault_report()

        # The headline goes in a bar right under the header, where nothing
        # else is drawn and it cannot be missed from a few feet away.
        tk = _tk()
        top = _band_top() + int(14 * tk)
        right = _band_right(w)
        bs = 1.0 * tk
        (bw, bh), _ = cv2.getTextSize(banner, font, bs, 3)
        if bw > right - 24:                  # a long row number, a long code
            bs *= (right - 24) / float(bw)
            (bw, bh), _ = cv2.getTextSize(banner, font, bs, 3)
        bar = bh + int(46 * tk)
        cv2.rectangle(frame, (0, top), (right, top + bar), FAULT_RED, -1)
        cv2.putText(frame, banner, (max((right - bw) // 2, 12),
                                    top + (bar + bh) // 2),
                    font, bs, (0, 0, 0), 3, cv2.LINE_AA)

        # The detail stacks up from just above the status bar, clear of the
        # chrome and of the payload strips that hang under each box.
        y = h - _band_bottom() - int(20 * tk)
        for text, colour, scale in reversed(lines):
            ts = scale * tk
            (tw, th), _ = cv2.getTextSize(text, font, ts, 2)
            if tw > right - 60:
                ts *= (right - 60) / float(tw)
                (tw, th), _ = cv2.getTextSize(text, font, ts, 2)
            cv2.rectangle(frame, (16, y - th - 12), (16 + tw + 24, y + 12),
                          (0, 0, 0), -1)
            cv2.putText(frame, text, (28, y), font, ts, colour, 2,
                        cv2.LINE_AA)
            y -= th + int(20 * tk)

        # And on the labels themselves, so the console text and the coil in
        # front of the operator are talking about the same thing.
        for box, _who, text in marks[0]:
            tag = _fault_tag(box, text)
            if tag is None:
                continue
            caption, colour = tag
            x1, y1, x2, y2 = (int(v) for v in box)
            cv2.rectangle(frame, (x1 - 7, y1 - 7), (x2 + 7, y2 + 7), colour, 7)
            if not caption:
                continue
            # Inside the top of the box, not above it: above would land on
            # the payload strip draw_decoders hangs under the box before it.
            cs = 0.62 * _tk()
            room = int((x2 - x1) * 1.6)
            (tw, th), _ = cv2.getTextSize(caption, font, cs, 2)
            if tw > room:
                cs *= room / float(tw)
                (tw, th), _ = cv2.getTextSize(caption, font, cs, 2)
            cv2.rectangle(frame, (x1, y1), (x1 + tw + 18, y1 + th + 18),
                          colour, -1)
            cv2.putText(frame, caption, (x1 + 9, y1 + th + 6), font, cs,
                        (0, 0, 0), 2, cv2.LINE_AA)
        return frame

    def scan_frame(frame, dets):
        """Decode every label in the frame and offer each payload to the
        window. No trigger line: a label is read on whichever frame it happens
        to be legible in, and the window decides whether it belongs here.
        """
        if label_cls is None:
            return

        qr_dets = [d for d in dets if int(d[5]) == qr_cls]
        logo_dets = [d for d in dets if int(d[5]) == logo_cls]
        label_dets = [d for d in dets if int(d[5]) == label_cls]

        # Crops held back from an earlier frame, in case this one has
        # the label whole.
        _flush_crops(frame, label_dets)

        # How fast the coil is running, in the only units that matter here:
        # pixels of this picture, per frame. It is what the saved crop
        # reaches back by, so the margin follows the winder up and down
        # instead of being a number somebody guessed once.
        here = [tuple(float(v) for v in d[:4]) for d in label_dets]
        motion[0] = web_motion(here, last_labels[0])
        last_labels[0] = here
        parts[0] = ([(tuple(float(v) for v in d[:4]), "qr") for d in qr_dets]
                    + [(tuple(float(v) for v in d[:4]), "logo")
                       for d in logo_dets])
        # Before the decoding, and on every label in view rather than only on
        # the ones being decoded this frame: a label that read fine on the
        # frame it came in on is skipped by the carry-forward below, and its
        # logo would never be looked at.
        # marks[0] is still the last frame's verdicts here — it is reset
        # below — and a label that gave up a payload on it has a code on it
        # whatever the detector managed to box this time.
        _check_parts(frame, label_dets, qr_dets, logo_dets, marks[0])
        zb_budget = [args.zbar_fallback]     # spent on this frame only
        settled = list(handled[0])       # labels dealt with on the last frame
        was = list(marks[0])             # last frame's verdicts, to carry over
        handled[0] = []
        marks[0] = []

        claimed = set()

        def carry(box):
            """What this label read, and which decoder got it, last time.

            The best overlap, and each of last frame's labels claimed only
            once. First-match-wins let two boxes inherit the same reading,
            which put one payload under two different labels on screen.
            """
            best, score = None, 0.45
            for i, (prev, _who, _text) in enumerate(was):
                if i in claimed:
                    continue
                overlap = _overlap(box, prev)
                if overlap > score:
                    best, score = i, overlap
            if best is None:
                return "zxing", None
            claimed.add(best)
            return was[best][1], was[best][2]
        budget = args.max_decodes or None

        for det in dets:
            if int(det[5]) != label_cls:
                continue

            # A label sits in view for many frames and would otherwise be
            # decoded on every one of them. Once its code has been accepted
            # there is nothing more to learn from it, so carry the box forward
            # by overlap and skip it — this is most of the per-frame cost.
            box = det[:4]
            if any(_overlap(box, prev) > 0.45 for prev in settled):
                handled[0].append(tuple(float(v) for v in box))
                prev_who, prev_text = carry(box)
                marks[0].append((tuple(float(v) for v in box),
                                 prev_who, prev_text))
                continue

            if budget is not None:
                if budget <= 0:
                    continue         # rest of the labels wait for next frame
                budget -= 1

            # The model finds the code itself, so that is what goes to the
            # reader: the qr box plus a quiet zone (--qr-margin), which is
            # what a QR needs around it to be read at all. The whole label is
            # the fallback for a label whose qr box was missed -- it holds
            # the same symbol, just with the artwork and the print around it.
            qr = pick_qr_for_label(det[:4], qr_dets)
            who = "zxing"
            text, box = (decode_qr(frame, qr[:4], args.qr_margin,
                                   args.qr_margin_min)
                         if qr is not None else (None, det[:4]))
            if not text:
                # The whole label goes to the reader, which means so does
                # anything else inside that box. On a label with nothing
                # printed on it, what comes back can be the code printed on
                # the label next door -- so the symbol has to be where the
                # label is, not merely somewhere in the crop.
                text, box, where = decode_qr_at(frame, det[:4], margin=0.0,
                                                min_px=0)
                if text and not _owns_symbol(det[:4], where, label_dets):
                    text = None
            if not text and zb_budget[0] > 0 and not _at_edge(det[:4], frame):
                # zxing has exhausted every pass it has on a crop that is
                # otherwise sound, so hand this one label to a decoder that
                # fails on different codes. A crop running off the frame edge
                # is still skipped: part of the symbol was never on the
                # sensor, and no library recovers that.
                zb_budget[0] -= 1
                text, box = decode_qr_pyzbar(
                    frame, qr[:4] if qr is not None else det[:4],
                    args.qr_margin if qr is not None else 0.0,
                    args.qr_margin_min if qr is not None else 0)
                if text:
                    zb_reads[0] += 1
                    who = "zbar"
            if not text:
                marks[0].append((tuple(float(v) for v in det[:4]),
                                 "fail", None))
                _dump_miss(frame, det[:4], qr)
                continue
            marks[0].append((tuple(float(v) for v in det[:4]), who, text))

            # The first payload the sheet recognises decides where the window
            # sits; until then there is nothing to hold anything against.
            if window.start is None:
                hit = sheet.find(text, near=resume_hint[0])
                if hit is None:
                    # This code is nowhere in the sheet, and nothing has
                    # anchored the window yet -- so not one label read so far
                    # belongs to this sheet. That is a coil the sheet does not
                    # describe: a new roll went on and the sheet was not
                    # changed with it, or the wrong sheet was loaded.
                    #
                    # It has to be caught here rather than by the unexpected
                    # path below, which only exists once the window has
                    # anchored. Until then every foreign code fell through
                    # this `continue` and the machine wound the whole roll
                    # through, reading and validating nothing.
                    key = normalize(text)
                    if fault["kind"] is not None or key in forgiven:
                        handled[0].append(tuple(float(v) for v in det[:4]))
                        continue
                    print(f"\n[window] WRONG COIL FOR THIS SHEET: {text}")
                    print(f"[window]   this code is in no row of "
                          f"{os.path.basename(args.xlsx)}, and no code read "
                          f"so far is either")
                    print(f"[window]   load the sheet that goes with this "
                          f"roll, or put the right roll on")
                    if saver is not None:
                        _save_crop(frame, det[:4], text, label_dets)
                    _raise_fault("mismatch", text=text,
                                 belongs="nothing in this sheet",
                                 seen=time.time(),
                                 box=tuple(float(v) for v in det[:4]))
                    voice.alert("Load the sheet for this roll.",
                                lead="Stopped. These labels are not in the "
                                     "sheet.", key="mismatch")
                    stop_machine("labels do not match the sheet")
                    continue
                anchored = window.anchor(hit[0])
                note = ("" if resume_hint[0] is None else
                        f", carrying on from row "
                        f"{sheet.rows[resume_hint[0]].number}")
                print(f"[window] anchored on sheet row {anchored}"
                      f"{note} — window covers {window.size} rows from there")

            status, row_idx, col, slot = window.offer(text)

            # Keep what was decoded and where it landed, so a row that comes
            # up short can be explained rather than just reported.
            ever_read.add(normalize(text))
            where = (f"row {sheet.rows[row_idx].number} {up(col)}"
                     if row_idx is not None else "no row in the window")
            recent.append((text, f"{status} -> {where}"))
            del recent[:-40]

            if status == RollingWindow.UNKNOWN:
                key = normalize(text)
                if key in forgiven:
                    # The operator pressed START with this code on screen, so
                    # it has already been judged by a human. Let it pass.
                    handled[0].append(tuple(float(v) for v in det[:4]))
                    continue

                if fault["kind"] == "unexpected" and \
                        normalize(fault["text"]) == key:
                    # The label that stopped the line is still in front of the
                    # camera during the rewind. Say nothing, but keep the
                    # clock alive: the fault only clears once this code has
                    # been out of frame for --rewind-clear seconds, and that
                    # is only measurable if it keeps being decoded while it is
                    # there — so it is deliberately NOT carried forward.
                    fault["seen"] = time.time()
                    fault["in_frame"] = True
                    fault["box"] = tuple(float(v) for v in det[:4])
                    continue

                if key in credited:
                    # This exact payload was matched to a cell and signed off
                    # earlier in this run, and it has now fallen out of the
                    # window's memory behind the head. That is the coil
                    # having been wound back further than it needed to be --
                    # not a wrong label. The row it belongs to was checked
                    # and recorded, so let it go past.
                    #
                    # It has to be `credited` and not merely `ever_read`: a
                    # genuinely wrong label decodes too, and would otherwise
                    # excuse itself on the second frame it was seen.
                    rewound[0] += 1
                    if rewound[0] == 1 or rewound[0] % 200 == 0:
                        print(f"\n[rewind] {_tail(text, 30)} was already "
                              f"read and signed off earlier in this run - "
                              f"the coil has been wound back past it. Not a "
                              f"fault ({rewound[0]} so far).")
                    handled[0].append(tuple(float(v) for v in det[:4]))
                    continue

                if fault["kind"] is not None:
                    # Already stopped and being wound back. Winding the coil
                    # in reverse walks rows that finished long ago past the
                    # lens again, and once they fall outside the grace span
                    # they read as codes from nowhere. That is the rewind
                    # working, not a second fault: one fault at a time.
                    handled[0].append(tuple(float(v) for v in det[:4]))
                    continue

                # A real code that belongs to no row inside the window: either
                # the wrong label is on the web, or the window has been left
                # far behind. Either way it is not something to tolerate.
                where = sheet.find(text)
                belongs = (f"sheet row {sheet.rows[where[0]].number} "
                           f"{up(where[1])}" if where else
                           "nothing in the sheet")
                head = sheet.rows[window.start].number if not window.exhausted else "?"
                print(f"\n[window] UNEXPECTED code: {text}")
                print(f"[window]   belongs to {belongs}; window starts at row {head}")
                window.note_unexpected(text, belongs)
                if saver is not None:
                    _save_crop(frame, det[:4], text, label_dets)
                _raise_fault("unexpected", text=text, belongs=belongs,
                             seen=time.time(),
                             box=tuple(float(v) for v in det[:4]))
                voice.alert("Rotate the coil back and take it off.",
                            lead="Stopped. Wrong label on the coil.",
                            key=f"bad-{key}")
                print(f"[rewind] rotate the coil back — this label stays "
                      f"outlined in red until it is out of frame, then the "
                      f"machine starts itself")
                stop_machine(f"unexpected code ({belongs})")
                continue

            if status == RollingWindow.REPEAT:
                handled[0].append(tuple(float(v) for v in det[:4]))
                continue

            if status == RollingWindow.MATCH:
                handled[0].append(tuple(float(v) for v in det[:4]))
                credited[normalize(text)] = row_idx
                row_no = sheet.rows[row_idx].number

                # Mid-rewind, call each position as it comes in. The operator
                # is watching the coil, not the screen, and this is how they
                # know winding back is working before it finishes.
                if fault["kind"] == "short" and row_idx == fault["row_idx"]:
                    voice.say(f"Up {col + 1} found.",
                              key=f"found-{row_no}-{col}")
                if args.debug:
                    print(f"[window] row {row_no} {up(col)} ok "
                          f"(slot {slot})")
                if saver is not None:
                    _save_crop(frame, det[:4], text, label_dets)

                # A code this far past the head means the coil has moved
                # on and the head row is not going to fill in on its own --
                # what is in front of the camera now is a later row. Rather
                # than write the head off, hold the line and keep the row
                # open: the operator winds the coil back, the same labels
                # come past again, and the row either completes or is judged
                # a real defect.
                #
                # This used to wait for the whole window (8 rows by default),
                # which meant the machine wound on through several more rows
                # before stopping for the one that failed. --hold-after is
                # how many rows of grace it gets, and it is small because the
                # point of stopping is to stop at the fault.
                if (slot >= hold_after[0] and recheck["row"] is None
                        and fault["kind"] is None):
                    head_idx = window.start
                    if head_idx is not None and window.missing(head_idx):
                        _hold_head_for_recheck(head_idx)

                for done_idx, _ in window.advance():
                    passed = sheet.rows[done_idx].number
                    if recheck["row"] == passed:
                        print(f"\n[recheck] row {passed} PASSED on "
                              f"re-inspection — every code read")
                        recheck["row"] = None
                        recheck["attempt"] = 0
                        retire_row(done_idx, complete=True)
                        _clear_fault(f"row {passed} read clean on the rewind")
                        continue
                    retire_row(done_idx, complete=True)

    # payload -> the sheet row it was matched to. A set would say only "this
    # has been checked", which cannot tell the rows standing ahead of a held
    # row at the moment it stops from the rows behind it after an over-wind;
    # both are checked. The row index says which side of the target it is on,
    # and it is a recorded fact rather than a lookup, so it survives a sheet
    # that repeats every few rows.
    credited = {}
    rewound = [0]            # times an already-signed-off code came past again
    # Crops waiting for a frame worth taking them from. A label is decoded on
    # whichever frame it first reads on, and that can easily be one where it
    # is half out of the picture -- the code can be wholly inside the sensor
    # while the rest of the label is not. The box is right; the label is
    # simply not all there. Saving from that frame is where a crop with the
    # code sliced down one edge comes from, and why the same label comes out
    # 244px wide one pass and 301 the next.
    #
    # So the two are separated: read it when it can be read, photograph it
    # when it is whole. A label stands in front of the camera for many
    # frames, so the wait is nearly always a frame or two.
    pending_crops = []       # [{"box", "text", "seen"}]
    CROP_PATIENCE = 120      # frames to wait for a clean look before giving up
    # The web's travel between the last two frames, and the label boxes that
    # measured it. Kept here rather than worked out per crop: every label in
    # the picture is on the same web, so it is one measurement a frame.
    last_labels = [[]]
    motion = [(0.0, 0.0)]
    # A label is three things: the label, the code on it and the logo on it.
    # The model finds all three, so a label with a part missing is visible as
    # such before anything is decoded -- and it has to be, because a label
    # with no code cannot be validated at all and would otherwise pass as a
    # row that simply came up short.
    waive_until = [0.0]      # the operator overruled: don't stop again yet
    # One entry per label being watched, and what has been seen on it:
    #
    #   looks   frames in which that label was whole and well inside the
    #           picture -- the only frames its printing can be judged from
    #   qr      how many of those frames the code was found on
    #   logo    how many of those frames the mark was found on
    #
    # A part found even once is a part that is printed: detectors drop boxes,
    # labels do not grow logos. So what stops the line is a part that is
    # never found at all, over enough looks that the detector cannot be the
    # explanation -- not a part missing from any number of frames in a row.
    tracks = []
    PARTS = ("QR", "LOGO")   # every label carries both
    EDGE_MARGIN = 12         # px of picture edge a label must stand clear of
    PART_OF_FULL = 0.85      # of the median label height, to count as whole

    def _whole_label(box, frame, label_dets):
        """Is all of this label inside the picture?

        Touching the edge is the obvious half of it. The other half is a box
        that came back short of what its neighbours measure: a label on its
        way out of frame stops at the sensor boundary, and the boundary is
        not always the very last pixel. Measured against the labels standing
        beside it, so it needs no notion of how big a label ought to be.
        """
        if _at_edge(box, frame, EDGE_MARGIN):
            return False
        if len(label_dets) < 3:
            return True                  # nothing to measure against
        widths = sorted(float(d[2]) - float(d[0]) for d in label_dets)
        heights = sorted(float(d[3]) - float(d[1]) for d in label_dets)
        full_w = widths[len(widths) // 2] * PART_OF_FULL
        full_h = heights[len(heights) // 2] * PART_OF_FULL
        return (box[2] - box[0]) >= full_w and (box[3] - box[1]) >= full_h

    def _save_crop(frame, box, text, label_dets):
        """Write the crop, or hold it until the label is all in the picture."""
        if saver is None:
            return
        box = tuple(float(v) for v in box[:4])
        if _whole_label(box, frame, label_dets):
            saver.save(frame, box, text, label_dets, motion[0])
            return
        pending_crops.append({"box": box, "text": text, "seen": 0})

    def _flush_crops(frame, label_dets):
        """Take the crops that have been waiting, off this frame if it will
        do. One pass per frame, and a label that never comes whole into the
        picture is let go rather than saved badly."""
        if saver is None or not pending_crops:
            return
        keep = []
        for item in pending_crops:
            best, score = None, 0.45
            for det in label_dets:
                other = tuple(float(v) for v in det[:4])
                lap = _overlap(item["box"], other)
                if lap > score:
                    best, score = other, lap
            item["seen"] += 1
            if best is None:
                continue                 # gone from the picture; let it go
            if _whole_label(best, frame, label_dets):
                saver.save(frame, best, item["text"], label_dets, motion[0])
                continue
            item["box"] = best           # follow it while it is still here
            if item["seen"] < CROP_PATIENCE:
                keep.append(item)
        pending_crops[:] = keep

    def _clear_look(box, frame, full):
        """Is this a frame the label's printing can be judged from?

        Two ways it is not: the label is crossing the edge of the picture, or
        its box came back shorter than the others' because part of it is
        outside. Either way the parts that are 'missing' are missing from the
        camera, not from the label.
        """
        return (not _at_edge(box, frame, EDGE_MARGIN)
                and (full <= 0 or box[3] - box[1] >= full))

    def _check_parts(frame, label_dets, qr_dets, logo_dets, read=()):
        """Stop for a label that is genuinely short of its code or its mark.

        Four things stand between a dropped box and a stopped line, because
        a false stop on a good roll costs as much as a missed bad label:

          * a label that read is a label with a code on it, whether or not
            the detector found a box to call one. That is not a tolerance,
            it is better evidence than the box: the symbol was not just
            seen, it was decoded. It matters most on the datamatrix rows,
            which the detector was never trained on and often does not box
            at all, and which would otherwise stop the line every time;
          * only clear looks count -- the label whole and well inside the
            picture, so nothing is judged on a label the camera has only
            half of;
          * a part found even once is a part that is printed, and that label
            is never questioned again. Detectors drop boxes; labels do not
            grow logos;
          * a part missing from half the labels in the picture at once is the
            detector, the focus or the light, not the labels. It counts
            against nobody.

        What is left -- one label, over --part-looks clear looks, short of
        something every one of its neighbours has -- is a label with a part
        missing.
        """
        if (args.no_part_check or not validating()
                or fault["kind"] is not None
                or time.time() < waive_until[0]):
            tracks[:] = []       # nothing carries across a stop
            return

        # What a whole label looks like on this frame, from the labels
        # themselves: a clipped one comes back shorter than its neighbours.
        heights = sorted(float(d[3]) - float(d[1]) for d in label_dets)
        full = (heights[len(heights) // 2] * PART_OF_FULL
                if len(heights) >= 3 else 0.0)

        decoded = [b for b, _who, text in read if text]

        looks = []
        for det in label_dets:
            box = tuple(float(v) for v in det[:4])
            if not _clear_look(box, frame, full):
                continue
            missing = set()
            if (pick_qr_for_label(box, qr_dets) is None
                    and not any(_overlap(box, prev) > 0.45
                                for prev in decoded)):
                missing.add("QR")
            if pick_qr_for_label(box, logo_dets) is None:
                missing.add("LOGO")
            looks.append((box, missing))
        if not looks:
            tracks[:] = []
            return

        # Half the labels short of the same part is a detector having a bad
        # frame. Not evidence against any one of them, either way.
        systemic = set()
        if len(looks) >= 3:
            systemic = {part for part in PARTS
                        if sum(1 for _b, m in looks if part in m) * 2
                        >= len(looks)}

        live, claimed = [], set()
        for box, missing in looks:
            # Which of the watched labels is this one? Best overlap, each
            # claimed once, as everywhere else in this file.
            best, score = None, 0.10
            for i, t in enumerate(tracks):
                if i in claimed:
                    continue
                overlap = _overlap(box, t["box"])
                if overlap > score:
                    best, score = i, overlap
            if best is None:
                t = {"box": box}
                for part in PARTS:
                    t[part] = [0, 0]      # [clear looks, times found]
            else:
                claimed.add(best)
                t = tracks[best]
                t["box"] = box
            live.append(t)

            short = None
            for part in PARTS:
                if part in systemic:
                    continue              # this frame says nothing about it
                t[part][0] += 1
                if part not in missing:
                    t[part][1] += 1
                elif (t[part][1] == 0
                      and t[part][0] >= max(1, args.part_looks)):
                    short = part
            if short is None:
                continue

            print(f"\n[label] INCOMPLETE LABEL: no {short.lower()} on it")
            print(f"[label]   {t[short][0]} clear looks at this label and the "
                  f"{short.lower()} was not found on any of them"
                  + "".join(f", {p.lower()} {t[p][1]}/{t[p][0]}"
                            for p in PARTS if p != short))
            # No crop for this one. The folder is the record of what was
            # read, one file per code, named after the code -- and a label
            # with nothing on it read nothing, so the file would be a
            # picture of a blank under a name that means nothing and cannot
            # be searched for. What it was and why it stopped the line is
            # on the screen, in the console, and in the run log.
            _raise_fault("incomplete", belongs=short, seen=time.time(),
                         box=box)
            voice.alert(f"A label has no {short.lower()}.",
                        lead="Stopped. Incomplete label.",
                        key=f"incomplete-{short}")
            print(f"[rewind] wind the coil back and take that label out — it "
                  f"stays outlined in red until it is out of frame, then the "
                  f"machine starts itself")
            stop_machine(f"label with no {short.lower()}")
            tracks[:] = []
            return
        tracks[:] = live

    # The web is only ever judged through payloads. A label that decodes to
    # nothing is not a match, not a repeat and not an unexpected code; it
    # fills no cell and empties none, so it moves no part of the machine. A
    # whole roll of them therefore looks exactly like a clean idle web, and
    # the line would wind the lot through. This is the watchdog for that.
    last_read = [0.0]        # when a label with a payload was last in frame

    def _reading():
        """Is anything in front of the camera actually giving us a code?

        Carried-forward labels count: a label read once and skipped on the
        frames after is still a label that reads. Without that, a still coil
        during the read-in would look like a blind camera.
        """
        return any(text for _b, _who, text in marks[0])

    def _check_reading():
        """Stop if labels keep coming past and none of them can be read."""
        if args.no_read_secs <= 0 or not validating():
            return
        if fault["kind"] is not None:
            return
        if not marks[0] or _reading():
            # An empty web is not a fault -- the roll may have run out, and
            # that is a different thing from a roll that cannot be read.
            last_read[0] = time.time()
            return
        blind = time.time() - last_read[0]
        if blind < args.no_read_secs:
            return
        n = len(marks[0])
        print(f"\n[window] NOTHING IS READING: {n} label(s) in frame and no "
              f"code out of any of them for {blind:.1f}s")
        print(f"[window]   either these labels carry no QR, or the camera "
              f"has stopped seeing them — check the labels first, then the "
              f"lens and the light")
        _raise_fault("unread", seen=time.time(), row=n)
        voice.alert("Check the labels and the camera.",
                    lead="Stopped. Nothing is reading.", key="unread")
        stop_machine(f"no label has read for {blind:.1f}s")

    def _track_bad_label():
        """Follow the offending label from frame to frame.

        Its box is re-anchored to whichever label it now overlaps most, so a
        coil being wound by hand does not shake it off -- and so it stays
        marked on the frames where it reads as the printed code underneath,
        or does not read at all.
        """
        if fault["kind"] not in ("unexpected", "mismatch", "incomplete") \
                or fault["box"] is None:
            return
        best, score = None, 0.20
        for box, _who, _text in marks[0]:
            overlap = _overlap(box, fault["box"])
            if overlap > score:
                best, score = box, overlap
        if best is not None:
            fault["box"] = best
            if fault["kind"] == "incomplete":
                # Still in view, wherever the winding has moved it to.
                fault["in_frame"] = True
                fault["seen"] = time.time()

    zb_reads = [0]           # labels rescued by the zbar fallback
    # [(box, 'qr'|'logo')] for the frame being drawn: what the model found
    # inside the labels.
    parts = [[]]
    # [(box, who, text)] for the frame being drawn: who is 'zxing', 'zbar'
    # or 'fail'. A label skipped by the overlap carry-forward keeps both the
    # verdict and the payload it earned when it was actually decoded, so the
    # overlay holds steady instead of flickering as labels stop being re-read.
    marks = [[]]
    DECODER_COLOUR = {"zxing": (80, 220, 80),      # green
                      "zbar":  (60, 200, 255),     # amber
                      "fail":  (60, 60, 235)}      # red
    # The two things the model finds inside a label. Drawn thin, and in
    # colours that are nobody else's, so the operator can see what the
    # detector has actually got hold of -- which is the only way to tell a
    # label that is short of a part from a detector that dropped a box.
    PART_COLOUR = {"qr":   (255, 190, 60),         # light blue
                   "logo": (230, 120, 255)}        # pink

    def _at_edge(box, frame, margin=2):
        h, w = frame.shape[:2]
        return (box[0] <= margin or box[1] <= margin
                or box[2] >= w - margin or box[3] >= h - margin)

    def _tail(text, n=46):
        """Payloads differ only at the end, so keep the tail, not the head."""
        text = str(text) if text else ""
        return text if len(text) <= n else ".." + text[-(n - 2):]

    def _hold_head_for_recheck(row_idx):
        """The head row came up short. Stop, and keep it open.

        Nothing is written off on the first look. The row stays at the head
        of the window, still matchable, so that when the operator winds the
        coil back its codes tick off exactly as they would have the first
        time. Reading clean clears it; pressing START while it is still short
        is what turns it into a confirmed defect.
        """
        row = sheet.rows[row_idx]
        row_no = row.number
        missing = sorted(window.missing(row_idx))
        cols = ", ".join(up(c) for c in missing)
        recheck["row"] = row_no
        recheck["attempt"] = 1
        print(f"\n[recheck] row {row_no} did not validate: never read {cols}")
        if row.is_dm:
            print(f"[recheck]   this is a DATA MATRIX row — printing "
                  f"{row.printing or '?'} of {args.dm_repeats} of the value "
                  f"listed against row {row.source} of your sheet. The labels "
                  f"to look for carry a datamatrix, not a QR.")

        # What the sheet wanted against what actually came off the camera, for
        # every position of this row — not just the ones that failed. A code
        # that read fine proves the row was in view, which is what makes the
        # missing one interesting.
        req = window.required(row_idx)
        seen = window.seen.get(row_idx, set())
        print(f"[recheck]   {'POS':<5}{'EXPECTED (xlsx)':<48}"
              f"{'DECODED (camera)':<48}RESULT")
        for col, want in enumerate(sheet.rows[row_idx].texts):
            if col not in req:
                got, res = "(not checked)", "-"
            elif col in seen:
                got, res = window.texts.get((row_idx, col), want), "OK"
            else:
                got, res = "*** NOTHING DECODED ***", "MISSING"
            print(f"[recheck]   {up(col):<5}{_tail(want, 46):<48}"
                  f"{_tail(got, 46):<48}{res}")

        # Was the missing code read anywhere at all? If it was, the label is
        # legible and the problem is which row it was credited to. If it was
        # never seen, the label never gave the decoder anything.
        for col in missing:
            want = sheet.rows[row_idx].texts[col]
            if normalize(want) in ever_read:
                print(f"[recheck]   note: {up(col)}'s code HAS been "
                      f"decoded earlier in this run — it was read somewhere, "
                      f"just not credited to row {row_no}")
            else:
                print(f"[recheck]   note: {up(col)}'s code has NEVER "
                      f"decoded in this run — that label is not being read "
                      f"at all")

        if recent:
            print(f"[recheck]   last {min(len(recent), 8)} codes decoded, "
                  f"newest first:")
            for text, verdict in reversed(recent[-8:]):
                print(f"[recheck]     {_tail(text, 44):<46}{verdict}")

        print(f"[recheck] STOPPED — rotate the coil back so row {row_no} "
              f"passes the camera again; its codes are picked up as they "
              f"come.")
        print(f"[recheck] the screen shows which positions are still missing. "
              f"Fill them all and the machine starts itself; press START with "
              f"any still missing and the row is recorded as a label defect.")
        _dump_frame(f"row{row_no}-short")
        if args.no_stop_on_fail:
            # Nothing is going to come and look at this row, so writing it off
            # is the only way the window can keep moving.
            recheck["row"] = None
            recheck["attempt"] = 0
            stale = window.evict_head()
            if stale is not None:
                retire_row(stale, complete=False)
            return
        _raise_fault("short", row=row_no, row_idx=row_idx)
        spoken = " and ".join(f"up {c + 1}" for c in missing)
        voice.alert(f"Row {row_no} did not read {spoken}.",
                    lead="Stopped. Rotate the coil back.",
                    key=f"short-{row_no}")
        stop_machine(f"row {row_no} short of {cols} — held for re-check")

    def _adjudicate_recheck():
        """Called when the operator presses START with a row held open.

        By now the coil has been wound back and the row has had its second
        look. If it filled in, the retire path above has already cleared it
        and there is nothing here to do. If it is still short, the labels have
        been shown to the camera twice and would not read either time — that
        is a defect, not a vision hiccup.
        """
        row_no = recheck["row"]
        if row_no is None:
            return
        row_idx = next((i for i in window.rows()
                        if sheet.rows[i].number == row_no), None)
        if row_idx is None:                  # already retired some other way
            recheck["row"] = None
            return

        missing = sorted(window.missing(row_idx))
        recheck["row"] = None
        recheck["attempt"] = 0
        if not missing:
            return                           # filled in; the retire path has it

        print(f"\n[recheck] row {row_no} is STILL short after re-inspection")
        print(f"[defect]  LABEL HAS ISSUE — sheet row {row_no}")
        voice.alert(f"Row {row_no}. Recorded.", lead="Label defect.",
                    key=f"defect-{row_no}")
        for col in missing:
            print(f"[defect]    {up(col)}: expected "
                  f"'{sheet.rows[row_idx].texts[col]}'")
            print(f"[defect]    {' ' * 9}  read     nothing")
        defects.add(row_no)
        record_row(row_idx, complete=False, status="DEFECT")
        if results is not None:
            results.write(_window_result(row_idx, False))
        while (window.start is not None and window.start <= row_idx
               and not window.exhausted):
            window.evict_head()
        head = (sheet.rows[window.start].number
                if window.start is not None and not window.exhausted
                else "end of sheet")
        print(f"[defect]  accepted — window moves on to row {head}")
        _clear_fault(f"row {row_no} written off as a defect", restart=False)

    def retire_row(row_idx, complete):
        """A sheet row has left the window. Record it and log the verdict.

        A row that came up short normally never reaches here — it is held open
        by _hold_head_for_recheck and settled by _adjudicate_recheck when the
        operator presses START. The short path below is only reached under
        --no-stop-on-fail, the one setting that lets a row be written off
        without a human looking at it.
        """
        row_no = sheet.rows[row_idx].number
        missing = sorted(window.missing(row_idx))
        record_row(row_idx, complete)
        if args.xlsx_every and len(verified) % args.xlsx_every == 0:
            write_annotated_xlsx(background=True)   # off unless asked for
        if complete and not missing:
            print(f"[window] PASS row {row_no}: all "
                  f"{len(window.required(row_idx))} codes read")
            if results is not None:
                results.write(_window_result(row_idx, True))
            return

        cols = ", ".join(up(c) for c in missing)
        print(f"[window] FAIL row {row_no}: never read {cols}")
        if results is not None:
            results.write(_window_result(row_idx, False))

    class _WindowResult:
        """Shaped like a BatchResult so the existing CSV writer can take it."""
        def __init__(self, row, entries, ok):
            self.row, self.entries, self.ok = row, entries, ok
        def summary(self):
            bad = [e.column for e in self.entries if e.status not in ("OK", "SKIPPED")]
            return f"row {self.row} missing {', '.join(bad)}" if bad else f"row {self.row}"

    class _WindowEntry:
        __slots__ = ("pos", "text", "expected", "status")
        def __init__(self, pos, text, expected, status):
            self.pos, self.text, self.expected, self.status = pos, text, expected, status
        @property
        def column(self):
            return up(self.pos)

    def _window_result(row_idx, ok):
        row = sheet.rows[row_idx]
        required = window.required(row_idx)
        seen = window.seen.get(row_idx, set())
        entries = []
        for col, expected in enumerate(row.texts):
            if col not in required:
                status = "SKIPPED"
            elif col in seen:
                status = "OK"
            else:
                status = "NO-READ"
            entries.append(_WindowEntry(col, expected if col in seen else None,
                                        expected, status))
        return _WindowResult(row.number, entries, ok)

    writer = None
    if args.save:
        writer = cv2.VideoWriter(args.save, cv2.VideoWriter_fourcc(*"mp4v"),
                                  args.fps, (disp_w, disp_h))

    # ── file pickers ─────────────────────────────────────────────────────
    # OpenCV's GTK highgui has no file chooser, so these borrow tkinter's. The
    # dialog is modal and runs on this thread, so the capture loop stalls
    # while it is open — which is exactly right: both pickers are only live
    # with the machine idle, so there is nothing to capture meanwhile.
    def _ask(kind):
        try:
            import tkinter
            from tkinter import filedialog
        except ImportError:
            print("[ui] tkinter is not installed - apt install python3-tk to "
                  "use the LOAD SHEET and LABEL FOLDER buttons")
            return None
        root = tkinter.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        try:
            if kind == "sheet":
                near = args.xlsx or prefs.sheet
                return filedialog.askopenfilename(
                    parent=root, title="Choose the validation sheet",
                    initialdir=(os.path.dirname(os.path.abspath(near))
                                if near else APP_DIR),
                    filetypes=[("Excel workbook", "*.xlsx *.xlsm"),
                               ("All files", "*.*")]) or None
            return filedialog.askdirectory(
                parent=root, title="Choose the folder for the label crops",
                initialdir=os.path.abspath(args.label_dir),
                mustexist=False) or None
        finally:
            root.destroy()

    _note = [None]        # what the status bar is saying, for either console

    def _configurable():
        """Both pickers repoint the whole record, so they are only live when
        nothing is part-way through being checked.

        A wrong-sheet fault is the exception: loading the right sheet is the
        fix for it, so the button that does that has to stay live.
        """
        return not validating() and fault["kind"] in (None, "mismatch")

    # Choosing a file and acting on the choice are separate, because under Qt
    # they happen on different threads: the dialog has to run on the GUI
    # thread, and everything it leads to has to run on the one that owns the
    # machine's state.
    def _apply_sheet(path):
        if not _configurable():
            print("[ui] stop the machine before loading a different sheet")
            return
        if not path or (args.xlsx
                        and os.path.abspath(path) == os.path.abspath(args.xlsx)):
            return
        try:
            ValidationSheet(path, args.sheet)     # prove it before committing
        except Exception as exc:
            print(f"[ui] {path} is not a validation sheet: {exc}")
            _note[0] = f"Not a validation sheet: {exc}"
            return
        _bind_run(new_xlsx=path)
        _note[0] = None

    def _apply_label_dir(path):
        if not _configurable():
            print("[ui] stop the machine before changing the label folder")
            return
        if not path or (os.path.abspath(path)
                        == os.path.abspath(args.label_dir)):
            return
        try:
            os.makedirs(path, exist_ok=True)
        except OSError as exc:
            print(f"[ui] cannot write to {path}: {exc}")
            _note[0] = f"Cannot write to that folder: {exc}"
            return
        _bind_run(new_label_dir=path)
        _note[0] = None

    def _pick_sheet():
        if _configurable():
            _apply_sheet(_ask("sheet"))

    def _pick_output():
        if _configurable():
            _apply_label_dir(_ask("folder"))

    # Commands from the Qt console. It never touches the machine's state; it
    # posts a name here and the capture loop picks it up at the top of its
    # next pass, which is what keeps that state single-threaded and keeps a
    # 150ms relay write off the GUI thread.
    commands = queue.Queue()
    quitting = threading.Event()

    def _run_commands():
        while True:
            try:
                name, arg = commands.get_nowait()
            except queue.Empty:
                return
            if name == "start":
                start_machine("start button")
            elif name == "stop":
                stop_machine("stop button")
            elif name == "sheet":
                _apply_sheet(arg)
            elif name == "labeldir":
                _apply_label_dir(arg)
            elif name == "winder":
                set_winder(arg)
            elif name == "direction":
                set_direction(arg)
            elif name == "camera":
                _set_camera(arg)
            elif name == "debug":
                show_debug[0] = not show_debug[0]
                print(f"[ui] diagnostics "
                      f"{'on' if show_debug[0] else 'off'}")
            elif name == "quit":
                quitting.set()

    def ui_state():
        """What the console's status pill is showing."""
        if not loaded():
            return "nosheet"
        if starting_at[0] is not None:
            return "reading"
        if fault["kind"] == "mismatch":
            return "mismatch"
        if fault["kind"] == "unread":
            return "unread"
        if fault["kind"] == "incomplete":
            return "incomplete"
        if fault["kind"] is not None:
            return "rewind"
        if machine_running:
            return "running"
        # Idle means ready to go. With the winder on hand control it is not
        # ready to go, and saying so is the difference between an operator
        # pressing START and wondering why nothing happens, and an operator
        # reaching for the selector.
        return "idle" if winder_is_auto() else "manual"

    # The counters, the window's check status and the decoder tally are
    # diagnostics, not something an operator acts on, so they stay off the
    # screen unless they are asked for. 'd' toggles them mid-run.
    show_debug = [args.debug]

    # Qt owns the main thread, so under it the capture loop is the worker and
    # the console is what the main thread runs. `scale` is how far the frame
    # is shrunk on its way to the screen, and both consoles need it.
    qt_app = qt_window = None
    scale = min(DISP["max_width"] / disp_w, DISP["max_height"] / disp_h, 1.0)

    def display_frame(frame):
        """The frame at the size it will actually be shown.

        The window is capped either way, so pushing the full 5MP frame
        through the toolkit just makes it scale down again on the GUI thread.
        """
        if scale >= 1.0:
            return frame
        return cv2.resize(frame, None, fx=scale, fy=scale,
                          interpolation=cv2.INTER_NEAREST)

    if not args.no_display and args.ui == "qt":
        try:
            from utils.qt_ui import InspectorWindow
            from PyQt5 import QtWidgets
        except ImportError as exc:
            print(f"[ui] PyQt5 is not available ({exc}) - falling back to the "
                  f"OpenCV console. pip install PyQt5 for the faster one.")
            args.ui = "opencv"
        else:
            qt_app = QtWidgets.QApplication.instance() \
                or QtWidgets.QApplication([])
            qt_window = InspectorWindow()
            qt_window.locked = not args.unlock_window
            qt_window.command.connect(lambda name, arg:
                                      commands.put((name, arg)))
            # It fills the screen it is on -- the console is the only thing
            # on that machine and there is nothing to share the glass with.
            # Sizes inside it come from the same screen, so the panel PC and
            # a desk both get a layout meant for them.
            if args.fullscreen:
                qt_window.showFullScreen()
            else:
                qt_window.showMaximized()
            print("[ui] Qt console - WINDER AUTO/MANUAL, START, STOP, "
                  "LOAD SHEET, OPEN RECENT SHEET and LABEL FOLDER are "
                  "buttons. Ctrl+Alt+E is the camera (exposure, gain, "
                  "brightness), Ctrl+Alt+W the diagnostics, F11 full "
                  "screen.")

    if not args.no_display and args.ui == "opencv":
        # WINDOW_NORMAL makes the window resizable; without it, imshow opens
        # at the frame's native resolution which overflows most screens. We
        # set an initial on-screen size, capped to the configured maximum, while the
        # capture/inference itself still runs at full resolution.
        win_name = "Global Shutter Camera - YOLO26 TRT"
        cv2.namedWindow(win_name, cv2.WINDOW_NORMAL)
        cv2.resizeWindow(win_name, int(disp_w * scale), int(disp_h * scale))

        if show_debug[0]:
            view = render_window_view()
            cv2.namedWindow(WINDOW_VIEW, cv2.WINDOW_NORMAL)
            cv2.resizeWindow(WINDOW_VIEW, view.shape[1], view.shape[0])
            print(f"[window] comparison panel in the '{WINDOW_VIEW}' window")

        # The console needs `scale` so its type is sized for what the
        # operator actually sees rather than for the captured pixels.
        panel = ControlPanel(disp_w, disp_h,
                             on_start=lambda: start_machine("start button"),
                             on_stop=lambda: stop_machine("stop button"),
                             on_load_sheet=_pick_sheet,
                             on_output_dir=_pick_output,
                             display_scale=scale)
        # display_frame shrinks the frame by `scale` before imshow, and GTK
        # reports clicks in the coordinates of the image it was handed — so
        # scale them back up before hit-testing the buttons, which are laid
        # out against the full-resolution frame.
        def on_mouse(event, x, y, flags, param):
            panel.on_mouse(event, int(x / scale), int(y / scale), flags, param)

        cv2.setMouseCallback(win_name, on_mouse)
        print("[ui] console buttons in the window (keys: s = start, "
              "x = stop, o = load sheet, f = output folder, d = debug "
              "readout, q = quit)")


    # ── one pass of the machine ──────────────────────────────────────────
    # Capture, infer, decode, validate. Everything the app does per frame
    # except deciding how to show it, so that both consoles drive exactly the
    # same code and the Qt one can drive it from a worker thread.
    fps_state = [time.time(), 0.0]

    def process_frame():
        """Returns (frame, dets), or None when the grab failed."""
        ok, frame = cap.read()
        if not ok:
            print("[camera] frame grab failed, retrying...")
            return None

        if rotate_in_loop:
            frame = rotate_frame(frame, rotate_in_loop)

        # ── inference ────────────────────────────────────────────────────
        inp, ratio, pad = preprocess(frame, model.input_size)
        raw = model.infer(inp)
        dets = postprocess(raw, ratio, pad, frame.shape, args.conf_thres,
                           conf_per_class)

        # ── decode and validate ──────────────────────────────────────────
        # Only while the machine is running, or in the read-in after START
        # was pressed. Stopped, the detector still draws boxes but nothing is
        # decoded or held against the sheet — the operator is handling the
        # coil, and whatever drifts past the lens is not a verdict.
        # Cleared before the scan and set again by it, so it means "the
        # offending code decoded on this frame", not "recently".
        fault["in_frame"] = False
        if scanning():
            last_frame[0] = frame
            scan_frame(frame, dets)
            if (starting_at[0] is not None
                    and time.time() - starting_at[0] >= args.start_delay):
                _finish_start()
        else:
            marks[0] = []        # stopped: no verdicts, so no stale colours
            parts[0] = []
        _track_bad_label()
        _check_reading()

        # Nothing was reading and now something is: the labels that could not
        # be read have been wound out of shot, or whatever was in the way of
        # the camera is gone. Nothing to adjudicate -- a code is a code.
        if fault["kind"] == "unread" and not validating() and _reading():
            last_read[0] = time.time()
            _clear_fault("a label read again")

        # An incomplete label clears the way a wrong one does: there is
        # nothing to re-read, so what settles it is the label being wound out
        # of shot and staying out.
        if fault["kind"] == "incomplete" and not validating():
            gone = time.time() - fault["seen"]
            if not fault["in_frame"] and gone >= args.rewind_clear:
                _clear_fault(f"the incomplete label has been out of frame "
                             f"for {gone:.1f}s")

        # Rewinding after an unexpected code. There is no "it read clean"
        # moment for this fault the way there is for a short row — the code
        # is simply wrong — so what clears it is the offending label being
        # wound out of shot and staying out.
        if fault["kind"] == "unexpected" and not validating():
            gone = time.time() - fault["seen"]
            if not fault["in_frame"] and gone >= args.rewind_clear:
                key = normalize(fault["text"])
                if key in bounced:
                    # It has already been let off once and come straight
                    # back. Restarting again would just bounce the line
                    # between stop and start with nobody in the loop.
                    if not fault["warned"]:
                        fault["warned"] = True
                        print(f"\n[rewind] {_tail(fault['text'], 30)} has "
                              f"stopped the line twice — not restarting on "
                              f"its own. Press START to accept it, or take "
                              f"that label off the web.")
                        voice.alert("The same wrong label again. Press start "
                                    "to accept it, or take it off the coil.",
                                    key="bounced")
                else:
                    bounced.add(key)
                    _clear_fault(f"the unexpected code has been out of frame "
                                 f"for {gone:.1f}s")

        # Running FPS: time between consecutive frames, smoothed with an
        # exponential moving average so the readout doesn't jitter.
        now = time.time()
        dt = now - fps_state[0]
        fps_state[0] = now
        if dt > 0:
            inst = 1.0 / dt
            fps_state[1] = inst if fps_state[1] == 0.0 \
                else (0.9 * fps_state[1] + 0.1 * inst)
        return frame, dets

    def _live_note():
        """What the status bar should be saying right now."""
        if not loaded():
            return (_note[0] or "Load the sheet for this roll — LOAD SHEET, "
                    "or OPEN RECENT SHEET")
        if starting_at[0] is not None:
            left = max(args.start_delay - (time.time() - starting_at[0]), 0)
            return f"Reading labels - {left:.1f}s to relay on"
        if fault["kind"] is not None:
            return "STOPPED: " + _fault_headline()
        return _note[0]

    def _web_line():
        """Which way the web is running and how fast, as the crops see it.

        Here because the crop margin follows it: if a crop still clips a
        code, this is the number that says whether the margin went on the
        right edge, and how much of one there was to give.
        """
        dx, dy = motion[0]
        if abs(dx) < 1 and abs(dy) < 1:
            return "web still"
        if abs(dx) >= abs(dy):
            return f"web {'right' if dx > 0 else 'left'} {abs(dx):.0f} px/frame"
        return f"web {'down' if dy > 0 else 'up'} {abs(dy):.0f} px/frame"

    def _status_line():
        if not loaded():
            return f"no sheet loaded\n{_web_line()}"
        passed = sum(1 for _, ok in window.done if ok)
        head = (sheet.rows[window.start].number
                if window.start is not None and not window.exhausted else "-")
        return (f"{passed} pass / {len(window.done) - passed} short\n"
                f"{window.reads} codes read\nhead row {head}\n{_web_line()}")

    def _headless_line(frame, dets):
        if not loaded():
            print(f"[camera] frame {frame.shape}  fps={fps_state[1]:.1f}  "
                  f"dets={len(dets)}  NO SHEET   ", end="\r")
            return
        passed = sum(1 for _, ok in window.done if ok)
        held = f"  HELD row {recheck['row']}" if recheck["row"] else ""
        if fault["kind"] is not None:
            held += f"  REWIND[{_fault_headline()}]"
        if starting_at[0] is not None:
            left = args.start_delay - (time.time() - starting_at[0])
            held += f"  READ-IN {max(left, 0):.1f}s"
        elif not machine_running:
            held += "  IDLE"
        print(f"[camera] frame {frame.shape}  fps={fps_state[1]:.1f}  "
              f"dets={len(dets)}  pass={passed} "
              f"short={len(window.done) - passed}  "
              f"reads={window.reads}{held}   ", end="\r")

    # ── the Qt console's view of a frame ─────────────────────────────────
    def _rgb(bgr):
        return (bgr[2], bgr[1], bgr[0])

    def _snapshot(frame, dets):
        """Drawing instructions in the coordinates of the displayed image.

        The Qt side knows nothing about sheets, windows or faults, so every
        decision about what to show is made here and handed over as a box, a
        colour and a caption.
        """
        def at(box):
            return tuple(v * scale for v in box)

        boxes = []
        for box, who, text in marks[0]:
            x1, y1, x2, y2 = at(box)
            bad = _bad_label(box)
            colour = (DECODER_COLOUR["fail"] if bad
                      else DECODER_COLOUR.get(who, DECODER_COLOUR["fail"]))
            boxes.append((x1, y1, x2, y2, _rgb(colour),
                          _bad_caption() if bad else
                          _tail(text, 14) if text else "NO READ"))

        part_boxes = [at(box) + (_rgb(PART_COLOUR[kind]),)
                      for box, kind in parts[0]]

        tags = []
        banner = None
        lines = []
        if fault["kind"] is not None:
            for box, _who, text in marks[0]:
                tag = _fault_tag(box, text)
                if tag is None:
                    continue
                caption, colour = tag
                x1, y1, x2, y2 = at(box)
                tags.append((x1, y1, x2, y2, _rgb(colour), caption))
            banner, detail = _fault_report()
            lines = [(t, _rgb(c)) for t, c, _s in detail]

        return {
            "image": display_frame(frame),
            "boxes": boxes, "parts": part_boxes, "tags": tags,
            "banner": banner, "lines": lines,
            "state": ui_state(), "note": _live_note(),
            "sheet": os.path.basename(args.xlsx) if args.xlsx
                     else "no sheet loaded",
            "sheet_dir": (os.path.dirname(os.path.abspath(args.xlsx))
                          if args.xlsx else ""),
            "labeldir": args.label_dir,
            "recent": [p for p in prefs.recent if not args.xlsx
                       or os.path.abspath(p) != os.path.abspath(args.xlsx)],
            "configurable": _configurable(),
            # START is dead until a sheet says what this roll should be, and
            # until the winder is handed over to the console.
            "loaded": loaded(),
            "winder_auto": winder_is_auto(),
            "reverse": checking_reverse(),
            # What the sliders behind 's' are built from: the camera's own
            # limits, and where it is set now.
            "camera": {"ranges": camera.ranges, "values": cam_values[0]},
            "debug": show_debug[0],
            "fps": fps_state[1], "dets": len(dets),
            "status": _status_line(),
            "window_view": render_window_view() if show_debug[0] else None,
        }

    # ── the two consoles ─────────────────────────────────────────────────
    def _loop_opencv():
        """The original: chrome drawn onto the frame, shown with imshow."""
        while True:
            got = process_frame()
            if got is None:
                continue
            frame, dets = got

            frame = draw_detections(frame, dets, class_names)
            if scanning():
                draw_decoders(frame, legend=show_debug[0])
            draw_fault(frame)
            if show_debug[0]:
                draw_window(frame, compact=True)

            if panel is not None:
                panel.sheet_name = (os.path.basename(args.xlsx) if args.xlsx
                                    else "NO SHEET LOADED")
                panel.output_dir = args.label_dir
                panel.configurable = _configurable()
                panel.note = _live_note()
                panel.draw(frame, ui_state())

            if args.no_display:
                _headless_line(frame, dets)
            else:
                if show_debug[0]:
                    cv2.putText(frame,
                                f"FPS {fps_state[1]:.1f}   DETS {len(dets)}",
                                (20, _band_top() + int(30 * _tk())),
                                cv2.FONT_HERSHEY_SIMPLEX, 0.55 * _tk(),
                                (120, 220, 120), 2, cv2.LINE_AA)
                cv2.imshow(win_name, display_frame(frame))
                if show_debug[0]:
                    cv2.imshow(WINDOW_VIEW, render_window_view())

            if writer:
                writer.write(frame)

            if args.no_display:
                continue
            key = cv2.waitKey(1) & 0xFF
            if key == ord("q"):
                # Same rule as the Qt console: quitting stops the line and
                # shuts the run's books, so it waits until the machine is not
                # in the middle of something.
                if not args.unlock_window and (validating()
                                               or fault["kind"] is not None):
                    print("[ui] press STOP before quitting (or Ctrl-C to "
                          "force it)")
                    continue
                break
            if key == ord("d"):
                show_debug[0] = not show_debug[0]
                if show_debug[0]:
                    view = render_window_view()
                    cv2.namedWindow(WINDOW_VIEW, cv2.WINDOW_NORMAL)
                    cv2.resizeWindow(WINDOW_VIEW, view.shape[1],
                                     view.shape[0])
                else:
                    cv2.destroyWindow(WINDOW_VIEW)
                print(f"[ui] debug readout "
                      f"{'on' if show_debug[0] else 'off'}")
            elif panel is not None:
                panel.on_key(key)

    def _loop_qt():
        """Capture on a worker thread, console on the main one.

        The worker never touches a widget: it posts a snapshot and reads the
        command queue. The main thread never touches the machine: it turns
        clicks and keys into commands. That division is what makes it safe to
        run the two at once.
        """
        def work():
            try:
                while not quitting.is_set():
                    _run_commands()
                    got = process_frame()
                    if got is None:
                        continue
                    frame, dets = got
                    qt_window.post(_snapshot(frame, dets))
                    if writer:
                        writer.write(frame)
            except (SystemExit, KeyboardInterrupt):
                pass                              # an ordinary way out
            except BaseException as exc:          # never die in silence
                import traceback
                traceback.print_exc()
                print(f"[camera] capture thread stopped: {exc}")
            finally:
                quitting.set()
                QtCore.QMetaObject.invokeMethod(qt_app, "quit",
                                                QtCore.Qt.QueuedConnection)

        from PyQt5 import QtCore

        # Ctrl+C, under Qt. Python's default handler raises KeyboardInterrupt
        # wherever the interpreter happens to be standing, and under exec_()
        # that is always inside a slot — a paint, or the frame being applied.
        # An exception cannot cross back into C++, so PyQt prints it and
        # carries on: the console keeps running and every Ctrl+C only adds
        # another traceback. So the signal is taken here instead and turned
        # into the same orderly exit the window's own quit takes, which is
        # what lets the shutdown below run and put the relays down. Handlers
        # run on the main thread, the one inside exec_(), so quit() can be
        # called straight out.
        def _interrupt(_sig, _frame):
            if quitting.is_set():
                # The first one has already been asked for and the shutdown
                # is taking its time. Hand the signal back to Python so a
                # third press kills the process outright.
                signal.signal(signal.SIGINT, signal.SIG_DFL)
                print("\n[ui] still shutting down — Ctrl-C again to force it")
                return
            print("\n[ui] interrupted — shutting down")
            quitting.set()
            qt_app.quit()

        previous = signal.signal(signal.SIGINT, _interrupt)
        # Qt's loop is C++, and the interpreter only gets to run that handler
        # when something calls back into Python. Frames do that sixty times a
        # second — but a camera that has hung posts none at all, which is
        # exactly when Ctrl+C gets pressed, so this keeps a beat of its own.
        ticker = QtCore.QTimer()
        ticker.timeout.connect(lambda: None)
        ticker.start(200)

        worker = threading.Thread(target=work, name="capture", daemon=True)
        worker.start()
        try:
            qt_app.exec_()
        finally:
            signal.signal(signal.SIGINT, previous)
        quitting.set()
        worker.join(timeout=5.0)

    try:
        # Deliberately not started here. Nothing is decoded and nothing is
        # held against the sheet until the operator presses START, which is
        # what makes the read-in mean something: it validates the coil at the
        # position it is actually standing in.
        if loaded():
            # 's' is START on the OpenCV console; the Qt one has no bare
            # keys at all, so the prompt says whichever is true here.
            print("[ui] idle — set the winder to AUTO, then press START"
                  if qt_window is not None else
                  "[ui] idle — press START (or 's') to validate and run")
        else:
            print("[ui] waiting for a sheet — LOAD SHEET, or OPEN RECENT "
                  "SHEET for one that has been run before")
        if qt_window is not None:
            _loop_qt()
        else:
            _loop_opencv()

    except KeyboardInterrupt:
        # Ctrl+C is how the line is stopped from the terminal, so it is an
        # ordinary way out, not a crash: swallow it here and let the shutdown
        # below run — a traceback would only bury the run's closing tally.
        print("\n[ui] interrupted — shutting down")
    finally:
        stop_machine("shutting down")
        voice.close()
        relay.close()
        camera.close()
        cap.release()
        if writer:
            writer.release()
        if not args.no_display:
            cv2.destroyAllWindows()
        if journal[0] is not None:
            journal[0].close()
        if verified:
            for _ in range(60):          # let any background save finish first
                with xlsx_lock:
                    if not xlsx_busy[0]:
                        break
                time.sleep(0.1)
            write_annotated_xlsx()
        if results is not None:
            print(f"[results] {results.rows} rows written to {results.path}")
            results.close()
        if saver is not None:
            print(f"[crops] saved {saver.count} label crops to {saver.dir}/")
        if zb_reads[0]:
            print(f"[qr] {zb_reads[0]} label(s) read by the zbar fallback "
                  f"after zxing failed on them")
        if defects:
            print(f"[defect] {len(defects)} row(s) failed twice and were "
                  f"recorded as LABEL ISSUE: "
                  + ", ".join(str(r) for r in sorted(defects)))
        if loaded():
            passed = sum(1 for _, ok in window.done if ok)
            print(f"[validate] done: {passed} rows passed, "
                  f"{len(window.done) - passed} short, "
                  f"{window.reads} codes read")
        else:
            print("[validate] nothing checked — no sheet was loaded")


if __name__ == "__main__":
    main()
