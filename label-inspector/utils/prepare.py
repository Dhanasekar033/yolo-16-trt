"""Turn the sheet the operator loads into the sheet the machine can check.

The two are not the same shape. A sheet describes the roll one spreadsheet
row per row of labels across the web:

    QR DATA1   DATA MATRIX1   HR1   QR DATA2   DATA MATRIX2   HR2   ...
    HTTPS://...403E097F818767              403E097F818767   HTTPS://...
    HTTPS://...403E097F830E08  R27C6087F77  403E097F830E08   HTTPS://...

Almost every row carries QR codes and nothing else. Every so often -- once
every 600 rows on the roll this was written for -- a row also carries a
DATA MATRIX value, and that value is not printed on the same labels as the
QR. It is printed on labels of its own, `REPEATS` of them in a row, one after
the QR labels of the row it is listed against, each up carrying its own
value down the web.

So one spreadsheet row with a datamatrix in it is four physical rows of
labels on the coil: the QR row, then three datamatrix rows. The rolling
window matches a row of labels to a row of the sheet, and there is nothing
in the sheet standing for those three. Left alone they are labels the sheet
does not expect, which is the one fault that stops the line hardest.

This module writes the sheet the window actually runs against: a copy with
each datamatrix expanded into rows of its own, so that every physical row of
labels on the coil has exactly one row in the sheet to be judged against --
and so that a datamatrix printed twice, or four times, or with the third one
missing, is a row that comes up short, which is precisely the check that was
being asked for.

    row 1801  QR DATA1..4          <- the QR labels
    row 1802  R27C6087F77 ..F8F    <- the datamatrix, first printing
    row 1803  R27C6087F77 ..F8F    <-                 second
    row 1804  R27C6087F77 ..F8F    <-                 third

The expansion writes the datamatrix value into the QR DATA column, because
that is the column the window matches on and a code is a code whichever
symbol carried it; CODE TYPE alongside says which symbol it should have been
printed as, which is what the reader is told to look for and what the
overlay tells the operator.

The original file is never opened for writing. The copy lives beside the
record for that sheet -- result/<sheet>/prepared_<sheet>.xlsx -- with a small
.json next to it noting what it was made from, so it is rebuilt when the
operator edits the sheet and reused when they do not.
"""

import json
import os
import re

import openpyxl

QR_COL_RE = re.compile(r"^\s*QR\s*DATA\s*(\d+)\s*$", re.IGNORECASE)
DM_COL_RE = re.compile(r"^\s*DATA\s*MATRIX\s*(\d+)\s*$", re.IGNORECASE)
HR_COL_RE = re.compile(r"^\s*HR\s*(\d+)\s*$", re.IGNORECASE)

# Columns the expansion adds on the right, past whatever the sheet already
# has. ValidationSheet reads CODE TYPE; the other two are there so the
# operator can see what a row is and where it came from.
TYPE_COL   = "CODE TYPE"
PRINT_COL  = "PRINT NO"
SOURCE_COL = "SOURCE ROW"

QR = "QR"
DM = "DATA MATRIX"

REPEATS = 3          # printings of one datamatrix value, down the web


def _columns(header):
    """{n: index} for each family of numbered columns in the header row."""
    qr, dm, hr = {}, {}, {}
    for idx, name in enumerate(header):
        if name is None:
            continue
        for pattern, into in ((QR_COL_RE, qr), (DM_COL_RE, dm),
                              (HR_COL_RE, hr)):
            m = pattern.match(str(name))
            if m:
                into[int(m.group(1))] = idx
                break
    return qr, dm, hr


def _text(value):
    return "" if value is None else str(value).strip()


def signature(src, repeats=REPEATS):
    """What the copy was made from. A change here is a rebuild."""
    st = os.stat(src)
    return {"source": os.path.abspath(src), "size": st.st_size,
            "mtime": int(st.st_mtime), "repeats": int(repeats)}


def prepared_paths(src, dest_dir):
    name = os.path.splitext(os.path.basename(src))[0]
    stem = os.path.join(dest_dir, f"prepared_{name}")
    return stem + ".xlsx", stem + ".json"


def prepare(src, dest_dir, sheet=None, repeats=REPEATS, force=False,
            quiet=False):
    """The path of the sheet to actually run against.

    A sheet with no datamatrix in it is returned untouched — there is nothing
    to expand, and a copy of it would only be a second file to keep in step.
    Otherwise the expanded copy is written into dest_dir and its path comes
    back. The source is opened read-only either way.
    """
    src = os.path.abspath(src)
    dest, meta = prepared_paths(src, dest_dir)
    want = signature(src, repeats)

    if not force and os.path.exists(dest):
        try:
            with open(meta) as fh:
                have = json.load(fh)
        except (OSError, ValueError):
            have = None
        if isinstance(have, dict) and \
                all(have.get(k) == v for k, v in want.items()):
            if not quiet:
                print(f"[prepare] reusing {dest} — "
                      f"{have.get('datamatrix', '?')} datamatrix row(s) "
                      f"expanded x{repeats}")
            return dest

    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet else wb.worksheets[0]
        raw = list(ws.iter_rows(values_only=True))
        title = ws.title
    finally:
        wb.close()

    if not raw:
        raise ValueError(f"{src}: sheet is empty")
    header = list(raw[0])
    if any(str(name).strip().upper() == TYPE_COL for name in header
           if name is not None):
        # A working copy, chosen instead of the sheet it was made from --
        # off the recent menu, or out of the file dialog, since it sits in
        # the record folder in plain sight. Expanding it again would print
        # three rows for every datamatrix row it already has.
        if not quiet:
            print(f"[prepare] {os.path.basename(src)} has already been "
                  f"expanded — running against it as it is")
        return src
    qr_cols, dm_cols, hr_cols = _columns(header)
    if not qr_cols:
        raise ValueError(f"{src}: no 'QR DATA<n>' columns in the header row")
    if not dm_cols:
        if not quiet:
            print(f"[prepare] {os.path.basename(src)} has no DATA MATRIX "
                  f"columns — running against it as it is")
        return src

    width = len(header)

    # Read the sheet through once before writing anything. A workbook is
    # only opened if there turns out to be something to put in it -- an
    # openpyxl write-only workbook keeps a generator open over its own
    # archive from the first row appended, and one built and then abandoned
    # prints a traceback into the operator's log when it is collected.
    body = []
    for number, values in enumerate(raw[1:], start=2):
        values = list(values) + [None] * (width - len(values))
        if not any(_text(v) for v in values):
            continue                                   # blank spacer row
        matrix = {n: _text(values[i]) for n, i in dm_cols.items()
                  if i < width and _text(values[i])}
        body.append((number, values, matrix))

    kept = len(body)
    expanded = sum(1 for _n, _v, m in body if m)
    if not expanded:
        # The columns are there but every one of them is empty. There is
        # nothing to expand, so there is nothing a copy would say that the
        # sheet does not already, and a prepared_ file that is a byte-for-byte
        # restatement of the original is only something else to keep in step.
        if not quiet:
            print(f"[prepare] {os.path.basename(src)} has DATA MATRIX columns "
                  f"but no values in them — running against it as it is")
        return src

    out = openpyxl.Workbook(write_only=True)
    sheet_out = out.create_sheet(title=title[:31] or "Sheet1")
    sheet_out.append(list(header) + [TYPE_COL, PRINT_COL, SOURCE_COL])

    for number, values, matrix in body:
        sheet_out.append(values[:width] + [QR, None, number])
        if not matrix:
            continue
        # One row per printing. The datamatrix goes in the QR DATA column
        # because that is what the window matches on; every other column of
        # the source row is dropped, because none of it is on these labels.
        for printing in range(1, repeats + 1):
            row = [None] * width
            for n, text in matrix.items():
                if n in qr_cols:
                    row[qr_cols[n]] = text
                if n in dm_cols:
                    row[dm_cols[n]] = text
                if n in hr_cols:
                    row[hr_cols[n]] = text     # what is printed beside it
            sheet_out.append(row + [DM, printing, number])

    os.makedirs(dest_dir, exist_ok=True)
    tmp = dest + ".tmp"
    out.save(tmp)
    os.replace(tmp, dest)                      # never a half-written copy

    want["datamatrix"] = expanded
    want["rows"] = kept + expanded * repeats
    try:
        with open(meta, "w") as fh:
            json.dump(want, fh, indent=2)
    except OSError as exc:
        # The copy is what matters; losing the note only costs a rebuild.
        print(f"[prepare] could not write {meta} ({exc}) — the copy will be "
              f"rebuilt next time")

    if not quiet:
        print(f"[prepare] {os.path.basename(src)}: {expanded} row(s) carry a "
              f"datamatrix, each printed {repeats}x down the web")
        print(f"[prepare] wrote {dest} — {kept} label rows + "
              f"{expanded * repeats} datamatrix rows = {want['rows']}")
    return dest
